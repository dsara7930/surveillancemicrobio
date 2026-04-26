import streamlit as st
import json
import csv
import io
import os
import base64
import calendar as cal_module
from datetime import datetime, timedelta, date as date_type
import difflib
import qrcode
import qrcode.image.pil
from io import BytesIO
from datetime import date, datetime, timedelta
import streamlit.components.v1 as components
import uuid
import sys
import re

# ── Gestion accès protégé ──────────────────────────────────────────────────────
if "access_mode" not in st.session_state:
    st.session_state["access_mode"] = None

MOT_DE_PASSE_ADMIN = "pharmaCHBA"

# ── SUPABASE ───────────────────────────────────────────────────────────────────
SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY")
try:
    from supabase import create_client
    _supabase_client = None
except Exception:
    create_client = None
    _supabase_client = None

def get_supabase_client():
    global _supabase_client
    if not create_client or not SUPABASE_URL or not SUPABASE_KEY:
        return None
    if _supabase_client is None:
        try:
            _supabase_client = create_client(SUPABASE_URL, SUPABASE_KEY)
        except Exception:
            _supabase_client = None
    return _supabase_client

# ── JOURS FÉRIÉS & JOURS TRAVAILLÉS ───────────────────────────────────────────
def get_french_holidays(year):
    h = set()
    h.add(date_type(year, 1, 1))
    h.add(date_type(year, 5, 1))
    h.add(date_type(year, 5, 8))
    h.add(date_type(year, 7, 14))
    h.add(date_type(year, 8, 15))
    h.add(date_type(year, 11, 1))
    h.add(date_type(year, 11, 11))
    h.add(date_type(year, 12, 25))
    a = year % 19
    b, c = divmod(year, 100)
    d, e = divmod(b, 4)
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    hh = (19 * a + b - d - g + 15) % 30
    i, k = divmod(c, 4)
    l = (32 + 2 * e + 2 * i - hh - k) % 7
    m = (a + 11 * hh + 22 * l) // 451
    month = (hh + l - 7 * m + 114) // 31
    day = ((hh + l - 7 * m + 114) % 31) + 1
    easter = date_type(year, month, day)
    h.add(easter + timedelta(days=1))
    h.add(easter + timedelta(days=39))
    h.add(easter + timedelta(days=50))
    return h

_HOLIDAY_CACHE = {}

def get_holidays_cached(year):
    if year not in _HOLIDAY_CACHE:
        _HOLIDAY_CACHE[year] = get_french_holidays(year)
    return _HOLIDAY_CACHE[year]

def is_working_day(d):
    if d.weekday() >= 5:
        return False
    return d not in get_holidays_cached(d.year)

def next_working_day_offset(start_date, n_days):
    result = start_date + timedelta(days=n_days)
    while result.weekday() >= 5 or result in get_holidays_cached(result.year):
        result += timedelta(days=1)
    return result

# ── PAGE CONFIG ────────────────────────────────────────────────────────────────
st.set_page_config(layout="wide", page_title="MicroSurveillance URC", page_icon="🦠")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=Syne:wght@600;700;800&display=swap');
html,body,[class*="css"]{font-family:'Syne',sans-serif;font-size:17px}
.stApp,[data-testid="stAppViewContainer"],[data-testid="stMain"]{background:#f8fafc!important}
[data-testid="stHeader"]{background:#f8fafc!important;border-bottom:1px solid #e2e8f0!important}
#MainMenu,footer{visibility:hidden}
.block-container{max-width:100%!important;padding-left:1rem!important;padding-right:1rem!important}
p,li,div{color:#1e293b;font-size:1rem}
h1,h2,h3,h4{color:#0f172a!important;font-size:1.15rem!important}
.stMarkdown p,.stMarkdown li{color:#0f172a;font-size:1rem}
[data-testid="stSidebar"]{background:#ffffff!important;border-right:2px solid #e2e8f0!important;box-shadow:2px 0 12px rgba(0,0,0,.07)!important}
[data-testid="stSidebar"] *{color:#1e293b!important;font-size:.95rem!important}
[data-testid="stSidebar"] p{color:#0f172a!important;font-size:.95rem!important}
[data-testid="collapsedControl"]{background:#2563eb!important;border-radius:0 12px 12px 0!important;width:32px!important;min-width:32px!important;box-shadow:4px 0 16px rgba(37,99,235,.5)!important;border:none!important}
[data-testid="collapsedControl"]:hover{background:#1d4ed8!important}
[data-testid="collapsedControl"] svg{fill:#ffffff!important;stroke:#ffffff!important;width:20px!important;height:20px!important;opacity:1!important}
[data-testid="stSidebarCollapsedControl"]{background:#2563eb!important;border-radius:0 12px 12px 0!important;width:32px!important;box-shadow:4px 0 16px rgba(37,99,235,.5)!important}
[data-testid="stSidebarCollapsedControl"] svg{fill:#ffffff!important;stroke:#ffffff!important}
[data-testid="stSidebarCollapsedControl"]:hover{background:#1d4ed8!important}
.stButton>button{border-radius:8px!important;font-size:.95rem!important;font-weight:600!important;border:1.5px solid #e2e8f0!important;color:#1e293b!important;background:#ffffff!important;transition:all .15s;padding:8px 16px!important}
.stButton>button:hover{background:#f1f5f9!important;border-color:#0f172a!important}
.stButton>button[kind="primary"]{background:#2563eb!important;color:#fff!important;border-color:#2563eb!important}
.stButton>button[kind="primary"]:hover{background:#1d4ed8!important;border-color:#1d4ed8!important}
.stTextInput>div>div>input,.stTextArea>div>div>textarea,.stSelectbox>div>div,.stNumberInput>div>div>input{background:#ffffff!important;color:#1e293b!important;border:1.5px solid #cbd5e1!important;border-radius:8px!important;font-size:1rem!important}
.stTextInput>div>div>input:focus,.stTextArea>div>div>textarea:focus{border-color:#2563eb!important;box-shadow:0 0 0 3px rgba(37,99,235,.15)!important}
label{color:#374151!important;font-size:.95rem!important;font-weight:600!important}
.stCheckbox label{color:#374151!important;font-size:.95rem!important}
.stCheckbox span{color:#374151!important;font-size:.95rem!important}
div[data-testid="stExpander"]{background:#ffffff!important;border:1.5px solid #e2e8f0!important;border-radius:10px!important;box-shadow:0 1px 4px rgba(0,0,0,.06)!important}
div[data-testid="stExpander"] summary,div[data-testid="stExpander"] summary *{color:#1e293b!important;font-size:1rem!important}
div[data-testid="stExpander"] summary svg{fill:#0f172a!important;stroke:#0f172a!important}
.stTabs [data-baseweb="tab-list"]{background:#f1f5f9!important;border-radius:10px!important;padding:3px!important;border:1.5px solid #e2e8f0!important}
.stTabs [data-baseweb="tab"]{background:transparent!important;color:#0f172a!important;border-radius:8px!important;font-weight:600!important;font-size:1rem!important;padding:8px 18px!important}
.stTabs [aria-selected="true"]{background:#ffffff!important;color:#2563eb!important;box-shadow:0 1px 4px rgba(0,0,0,.1)!important}
[data-testid="stMetric"]{background:#fff!important;border:1.5px solid #e2e8f0!important;border-radius:10px!important;padding:16px!important}
[data-testid="stMetricValue"]{color:#0f172a!important;font-size:1.6rem!important}
[data-testid="stMetricLabel"]{color:#0f172a!important;font-size:.9rem!important}
[data-testid="stMetricDelta"]{font-size:.85rem!important}
.stAlert{border-radius:10px!important}
.stSuccess>div{background:#f0fdf4!important;border:1px solid #86efac!important;color:#166534!important;font-size:1rem!important}
.stWarning>div{background:#fffbeb!important;border:1px solid #fcd34d!important;color:#92400e!important;font-size:1rem!important}
.stInfo>div{background:#eff6ff!important;border:1px solid #93c5fd!important;color:#1e40af!important;font-size:1rem!important}
.stError>div{background:#fef2f2!important;border:1px solid #fca5a5!important;color:#991b1b!important;font-size:1rem!important}
hr{border-color:#e2e8f0!important}
[data-testid="stSidebar"] .stButton>button[kind="primary"]{background:#2563eb!important;color:#fff!important;border-color:#2563eb!important}
[data-testid="stSidebar"] .stButton>button{background:#f8fafc!important;color:#374151!important;border:1.5px solid #e2e8f0!important;font-size:1rem!important;padding:12px 8px!important}
[data-testid="stSidebar"] .stButton>button:hover{background:#eff6ff!important;border-color:#93c5fd!important}
[data-testid="stSidebar"] p{font-size:.95rem!important}
.stSlider [data-testid="stThumbValue"]{color:#2563eb!important}
.stNumberInput button{background:#f1f5f9!important;border-color:#e2e8f0!important;color:#1e293b!important;font-size:1rem!important}
.stNumberInput button:hover{background:#e2e8f0!important}
.stDownloadButton>button{color:#2563eb!important;border-color:#93c5fd!important;background:#eff6ff!important;font-size:.95rem!important}
.stDownloadButton>button:hover{background:#dbeafe!important}
.stCaption{font-size:.88rem!important}
[data-testid="stSelectbox"] div,[data-testid="stSelectbox"] span{font-size:1rem!important}
[data-testid="stRadio"] label,[data-testid="stRadio"] span{font-size:1rem!important}
[data-testid="stDateInput"] input{font-size:1rem!important}
</style>""", unsafe_allow_html=True)

# ── CONSTANTES ─────────────────────────────────────────────────────────────────
RISK_COLORS = {1:"#22c55e",2:"#84cc16",3:"#f59e0b",4:"#f97316",5:"#ef4444"}
RISK_LABELS = {1:"Limité",2:"Modéré",3:"Important",4:"Majeur",5:"Critique"}
CSV_FILE          = "surveillance_data.csv"
GERMS_FILE        = "germs_data.json"
THRESHOLDS_FILE   = "thresholds_config.json"
MEASURES_FILE     = "measures_config.json"
POINTS_FILE       = "points.json"
PRELEVEMENTS_FILE = "prelevements.json"
SCHEDULES_FILE    = "schedules.json"
PENDING_FILE      = "pending_identifications.json"
ARCHIVED_FILE     = "archived_samples.json"
OPERATORS_FILE    = "operators.json"
PLANS_FILE        = "plans_data.json"

DEFAULT_THRESHOLDS = {
    5: {"alert": 1,  "action": 1},
    4: {"alert": 10, "action": 25},
    3: {"alert": 25, "action": 40},
    2: {"alert": 25, "action": 40},
    1: {"alert": 40, "action": 50},
}

DEFAULT_MEASURES = {
    5: {
        "alert": "• Informer immédiatement le responsable qualité\n• Vérifier et renforcer le bionettoyage de la zone\n• Contrôler l'intégrité des filtres HEPA\n• Augmenter la fréquence de surveillance\n• Documenter l'événement dans le registre qualité",
        "action": "• ARRÊT IMMÉDIAT des activités si possible\n• Alerter le pharmacien responsable et la direction\n• Isoler la zone contaminée\n• Décontamination renforcée avec désinfectant adapté\n• Recherche de la source de contamination\n• Bilan mycologique complet\n• Ne pas reprendre l'activité avant résultat conforme\n• Déclaration d'événement indésirable"
    },
    4: {
        "alert": "• Informer le responsable qualité\n• Renforcer le bionettoyage de la zone concernée\n• Vérifier les procédures d'habillage et d'hygiène\n• Contrôler les flux d'air et la pression différentielle\n• Programmer un prélèvement de contrôle sous 48h",
        "action": "• Alerter le pharmacien responsable\n• Suspendre les préparations critiques si nécessaire\n• Nettoyage et désinfection renforcés de la zone\n• Vérification complète de l'installation\n• Enquête sur l'origine de la contamination\n• Prélèvements de contrôle avant reprise\n• Enregistrement et analyse des causes"
    },
    3: {
        "alert": "• Informer le responsable d'équipe\n• Renforcer les mesures d'hygiène du personnel\n• Vérifier le respect des procédures de bionettoyage\n• Programmer un prélèvement de contrôle\n• Surveiller l'évolution",
        "action": "• Informer le responsable qualité\n• Nettoyage et désinfection de la zone\n• Vérification des procédures en vigueur\n• Renforcer la formation du personnel si nécessaire\n• Prélèvements de contrôle sous 72h\n• Documentation et analyse de tendance"
    },
    2: {
        "alert": "• Surveiller l'évolution des résultats\n• Vérifier le respect des procédures de bionettoyage\n• Contrôler l'hygiène du personnel\n• Programmer un prélèvement de contrôle",
        "action": "• Informer le responsable d'équipe\n• Renforcer le bionettoyage de la zone\n• Vérifier les procédures d'habillage\n• Prélèvements de contrôle sous 5 jours\n• Analyse de tendance sur les derniers résultats"
    },
    1: {
        "alert": "• Surveiller l'évolution\n• Vérifier les procédures de nettoyage\n• Prélèvement de contrôle à planifier",
        "action": "• Renforcer le bionettoyage\n• Vérifier l'hygiène du personnel\n• Prélèvement de contrôle sous 1 semaine\n• Documentation dans le registre de surveillance"
    },
}

ALL_ORIGINS = [
    "Air","Humidité","Flore fécale","Oropharynx / Gouttelettes",
    "Peau / Muqueuse","Sol / Carton / Surface sèche",
]

# ── DONNÉES FAQ ────────────────────────────────────────────────────────────────
FAQ_CATEGORIES = [
    "Général", "Score & Seuils", "Prélèvements",
    "Paramètres", "Données", "Mesures correctives",
]

DEFAULT_FAQ = [
    {
        "id": "faq_001", "category": "Général", "order": 0,
        "question": "À quoi sert cette application ?",
        "answer": (
            "Cette application permet de **suivre la surveillance microbiologique** "
            "de votre environnement pharmaceutique. Elle gère les prélèvements d'air "
            "et de surface, l'identification des germes, le calcul automatique des scores "
            "de criticité et le déclenchement des alertes et actions correctives."
        ),
    },
    {
        "id": "faq_002", "category": "Score & Seuils", "order": 1,
        "question": "Comment est calculé le score de criticité ?",
        "answer": (
            "Le score total est calculé selon la formule :\n\n"
            "**Score = Criticité lieu (1–4) × Pathogénicité (1–3) × Résistance (1–3) × Dissémination (1–3)**\n\n"
            "- Score **< seuil alerte** → ✅ Conforme\n"
            "- Score entre les deux seuils → ⚠️ Alerte\n"
            "- Score **> seuil action** → 🚨 Action immédiate\n\n"
            "Les seuils sont configurables dans **Paramètres → Seuils d'alerte**."
        ),
    },
    {
        "id": "faq_003", "category": "Score & Seuils", "order": 2,
        "question": "Que signifie la criticité du lieu (1–3) ?",
        "answer": (
            "La criticité du lieu qualifie l'importance environnementale du point de prélèvement :\n\n"
            "- **Niveau 1** 🟢 — Limité\n"
            "- **Niveau 2** 🔵 — Modéré\n"
            "- **Niveau 3** 🟠 — Important\n"
            "- **Niveau 4** 🔴 — Critique\n"
            "Ce niveau est défini dans **Paramètres → Points de prélèvement** et est "
            "automatiquement repris lors de l'identification microbiologique."
        ),
    },
    {
        "id": "faq_004", "category": "Prélèvements", "order": 3,
        "question": "Comment ajouter un nouveau prélèvement ?",
        "answer": (
            "Rendez-vous dans **Surveillance → Nouveau prélèvement**, puis :\n\n"
            "1. Sélectionnez le point de prélèvement dans le menu déroulant\n"
            "2. Choisissez l'opérateur et la date\n"
            "4. Validez le formulaire\n\n"
            "Le prélèvement apparaîtra dans la liste en attente d'identification microbiologique."
        ),
    },
    {
        "id": "faq_005", "category": "Prélèvements", "order": 4,
        "question": "Qu'est-ce qu'une identification microbiologique ?",
        "answer": (
            "Après la lecture des géloses, vous saisissez les germes identifiés pour chaque prélèvement. "
            "L'application calcule alors automatiquement le **score de criticité** pour chaque germe "
            "trouvé et déclenche les alertes ou actions appropriées.\n\n"
            "Accès : **Surveillance → Identifier** sur un prélèvement en attente."
        ),
    },
    {
        "id": "faq_006", "category": "Paramètres", "order": 5,
        "question": "Comment configurer les points de prélèvement ?",
        "answer": (
            "Dans **Paramètres → Points de prélèvement** :\n\n"
            "- Cliquez sur **➕ Ajouter** pour créer un nouveau point\n"
            "- Renseignez le nom, le type (Air / Surface), la classe ISO/GMP, "
            "la criticité du lieu, la gélose utilisée et la fréquence de prélèvement\n"
            "- Modifiez (✏️) ou supprimez (🗑️) les points existants\n\n"
            "Les points sont synchronisés automatiquement avec Supabase si configuré."
        ),
    },
    {
        "id": "faq_008", "category": "Données", "order": 7,
        "question": "Comment sauvegarder et restaurer mes données ?",
        "answer": (
            "Dans **Paramètres → Sauvegarde** :\n\n"
            "- **⬇️ Exporter** : télécharge un fichier JSON contenant toutes vos données "
            "(germes, prélèvements, points, opérateurs, historique...)\n"
            "- **⬆️ Restaurer** : importez un fichier de sauvegarde pour remplacer "
            "toutes les données actuelles\n\n"
            "⚠️ La restauration est irréversible — effectuez toujours une exportation avant."
        ),
    },
    {
        "id": "faq_009", "category": "Données", "order": 8,
        "question": "Pourquoi configurer Supabase ?",
        "answer": (
            "Sans Supabase, les données sont **perdues à chaque redémarrage** de l'application "
            "(mise à jour du code, timeout Streamlit Cloud...).\n\n"
            "Supabase est une base de données cloud **gratuite** qui garantit la persistance. "
            "Configuration dans **Paramètres → Base de données** — il suffit de coller "
            "votre `SUPABASE_URL` et `SUPABASE_KEY` dans les secrets Streamlit."
        ),
    },
    {
        "id": "faq_010", "category": "Mesures correctives", "order": 9,
        "question": "Comment fonctionnent les mesures correctives ?",
        "answer": (
            "Les mesures correctives sont des actions prédéfinies suggérées automatiquement "
            "lors d'une alerte ou d'une action microbiologique.\n\n"
            "Chaque mesure est associée à :\n"
            "- Une **origine** (Air, Surface, Flore fécale...)\n"
            "- Un **niveau de criticité** (1 à 5)\n"
            "- Un **type** (⚠️ Alerte ou 🚨 Action)\n\n"
            "Gérez-les dans **Paramètres → Mesures correctives**."
        ),
    },
]

DEFAULT_ORIGIN_MEASURES = [
    {"id":"m001","text":"Documenter l'événement dans le registre qualité","scope":"all","risk":"all","type":"alert"},
    {"id":"m002","text":"Programmer un prélèvement de contrôle","scope":"all","risk":"all","type":"alert"},
    {"id":"m003","text":"Surveiller l'évolution des prochains résultats","scope":"all","risk":"all","type":"alert"},
    {"id":"m004","text":"Informer le responsable qualité","scope":"all","risk":[3,4,5],"type":"alert"},
    {"id":"m005","text":"Augmenter la fréquence de surveillance","scope":"all","risk":[3,4,5],"type":"alert"},
    {"id":"m006","text":"Renforcer le bionettoyage de la zone concernée","scope":"all","risk":[3,4,5],"type":"alert"},
    {"id":"m010","text":"Alerter le pharmacien responsable et la direction","scope":"all","risk":[4,5],"type":"action"},
    {"id":"m011","text":"Isoler la zone contaminée","scope":"all","risk":[4,5],"type":"action"},
    {"id":"m012","text":"ARRÊT des préparations critiques si nécessaire","scope":"all","risk":[4,5],"type":"action"},
    {"id":"m013","text":"Décontamination renforcée avec désinfectant adapté (Surfa'Safe / APA)","scope":"all","risk":[4,5],"type":"action"},
    {"id":"m014","text":"Ne pas reprendre l'activité avant résultat conforme","scope":"all","risk":[4,5],"type":"action"},
    {"id":"m015","text":"Déclaration d'événement indésirable (fiche EI)","scope":"all","risk":[4,5],"type":"action"},
    {"id":"m016","text":"Renforcer le bionettoyage de la zone","scope":"all","risk":[1,2,3],"type":"action"},
    {"id":"m017","text":"Vérifier les procédures de bionettoyage en vigueur","scope":"all","risk":[1,2,3],"type":"action"},
    {"id":"m018","text":"Informer le responsable d'équipe","scope":"all","risk":[1,2,3],"type":"action"},
    {"id":"m020","text":"Contrôler l'intégrité des filtres HEPA","scope":"Air","risk":"all","type":"alert"},
    {"id":"m021","text":"Vérifier les flux d'air et la pression différentielle","scope":"Air","risk":"all","type":"alert"},
    {"id":"m022","text":"Contrôler les entrées / sorties de matériel (cartons, vêtements)","scope":"Air","risk":[3,4,5],"type":"alert"},
    {"id":"m023","text":"Contrôler l'étanchéité des jonctions de filtres HEPA","scope":"Air","risk":[4,5],"type":"action"},
    {"id":"m024","text":"Bilan fongique complet (Aspergillus, Fusarium, Penicillium)","scope":"Air","risk":[4,5],"type":"action"},
    {"id":"m025","text":"Décontamination par nébulisation H2O2 si moisissure critique","scope":"Air","risk":[5],"type":"action"},
    {"id":"m030","text":"Identifier et traiter les sources d'humidité","scope":"Humidité","risk":"all","type":"alert"},
    {"id":"m031","text":"Vérifier l'étanchéité des canalisations et conduites d'eau","scope":"Humidité","risk":"all","type":"alert"},
    {"id":"m032","text":"Contrôler le taux d'humidité relative de la salle","scope":"Humidité","risk":[3,4,5],"type":"alert"},
    {"id":"m033","text":"Nettoyage et désinfection renforcés des surfaces humides","scope":"Humidité","risk":"all","type":"action"},
    {"id":"m034","text":"Décontamination au peroxyde d'hydrogène si Pseudomonas/Mycobactérie","scope":"Humidité","risk":[4,5],"type":"action"},
    {"id":"m035","text":"Recherche et élimination de tout biofilm résiduel","scope":"Humidité","risk":[4,5],"type":"action"},
    {"id":"m040","text":"Contrôler les entrées de matières premières et emballages","scope":"Sol / Carton / Surface sèche","risk":"all","type":"alert"},
    {"id":"m041","text":"Renforcer le bionettoyage des sols et surfaces","scope":"Sol / Carton / Surface sèche","risk":"all","type":"alert"},
    {"id":"m042","text":"Vérifier le protocole de dé-cartonnage à l'entrée","scope":"Sol / Carton / Surface sèche","risk":[3,4,5],"type":"alert"},
    {"id":"m043","text":"Retrait et destruction des cartons et emballages suspects","scope":"Sol / Carton / Surface sèche","risk":"all","type":"action"},
    {"id":"m044","text":"Décontamination sporicide si spores détectées (Bacillus, Clostridium)","scope":"Sol / Carton / Surface sèche","risk":[4,5],"type":"action"},
    {"id":"m045","text":"Bilan sporal complet de la zone","scope":"Sol / Carton / Surface sèche","risk":[5],"type":"action"},
    {"id":"m050","text":"Vérifier les procédures d'habillage et port des EPI","scope":"Peau / Muqueuse","risk":"all","type":"alert"},
    {"id":"m051","text":"Contrôler la technique de friction hydro-alcoolique","scope":"Peau / Muqueuse","risk":"all","type":"alert"},
    {"id":"m052","text":"Renforcer la formation du personnel (hygiène des mains)","scope":"Peau / Muqueuse","risk":[3,4,5],"type":"action"},
    {"id":"m053","text":"Vérifier l'absence de lésion cutanée chez le personnel","scope":"Peau / Muqueuse","risk":[4,5],"type":"action"},
    {"id":"m054","text":"Enquête sur le personnel intervenant dans la zone","scope":"Peau / Muqueuse","risk":[4,5],"type":"action"},
    {"id":"m060","text":"Vérifier les procédures de lavage des mains","scope":"Flore fécale","risk":"all","type":"alert"},
    {"id":"m061","text":"Contrôler la chaîne de décontamination des équipements","scope":"Flore fécale","risk":"all","type":"alert"},
    {"id":"m062","text":"Recherche de source de contamination fécale","scope":"Flore fécale","risk":[3,4,5],"type":"action"},
    {"id":"m063","text":"Nettoyage désinfectant à spectre large (entérobactéries/ERV)","scope":"Flore fécale","risk":"all","type":"action"},
    {"id":"m064","text":"Test de portage pour le personnel si E. coli / Entérocoque multirésistant","scope":"Flore fécale","risk":[4,5],"type":"action"},
    {"id":"m070","text":"Vérifier le port correct du masque FFP2 ou chirurgical","scope":"Oropharynx / Gouttelettes","risk":"all","type":"alert"},
    {"id":"m071","text":"Rappeler l'interdiction de parler dans la ZAC","scope":"Oropharynx / Gouttelettes","risk":"all","type":"alert"},
    {"id":"m072","text":"Enquête sur le personnel présent lors du prélèvement positif","scope":"Oropharynx / Gouttelettes","risk":[3,4,5],"type":"action"},
    {"id":"m073","text":"Contrôle de santé du personnel (angine, rhino-pharyngite)","scope":"Oropharynx / Gouttelettes","risk":[3,4,5],"type":"action"},
    {"id":"m074","text":"Éviction temporaire du personnel symptomatique","scope":"Oropharynx / Gouttelettes","risk":[4,5],"type":"action"},
]

DEFAULT_GERMS = [
    dict(name="Staphylococcus spp.",path=["Germes","Bactéries","Humains","Peau / Muqueuse"],notes=None,comment=None),
    dict(name="Corynebacterium spp.",path=["Germes","Bactéries","Humains","Peau / Muqueuse"],notes=None,comment=None),
    dict(name="Cutibacterium acnes",path=["Germes","Bactéries","Humains","Peau / Muqueuse"],notes=None,comment=None),
    dict(name="Micrococcus spp.",path=["Germes","Bactéries","Humains","Peau / Muqueuse"],notes=None,comment=None),
    dict(name="Dermabacter hominis",path=["Germes","Bactéries","Humains","Peau / Muqueuse"],notes=None,comment=None),
    dict(name="Brevibacterium epidermidis",path=["Germes","Bactéries","Humains","Peau / Muqueuse"],notes=None,comment=None),
    dict(name="Streptococcus mitis/salivarius/sanguinis/anginosus",path=["Germes","Bactéries","Humains","Oropharynx / Gouttelettes"],notes=None,comment=None),
    dict(name="Streptococcus pyogenes/agalactiae/pneumoniae",path=["Germes","Bactéries","Humains","Oropharynx / Gouttelettes"],notes=None,comment=None),
    dict(name="Escherichia coli",path=["Germes","Bactéries","Humains","Flore fécale"],notes=None,comment=None),
    dict(name="Enterococcus spp.",path=["Germes","Bactéries","Humains","Flore fécale"],notes=None,comment=None),
    dict(name="Enterobacter spp.",path=["Germes","Bactéries","Humains","Flore fécale"],notes=None,comment=None),
    dict(name="Citrobacter spp.",path=["Germes","Bactéries","Humains","Flore fécale"],notes=None,comment=None),
    dict(name="Klebsiella pneumoniae",path=["Germes","Bactéries","Humains","Flore fécale"],notes=None,comment=None),
    dict(name="Proteus spp.",path=["Germes","Bactéries","Humains","Flore fécale"],notes=None,comment=None),
    dict(name="Morganella spp.",path=["Germes","Bactéries","Humains","Flore fécale"],notes=None,comment=None),
    dict(name="Providencia spp.",path=["Germes","Bactéries","Humains","Flore fécale"],notes=None,comment=None),
    dict(name="Salmonella spp.",path=["Germes","Bactéries","Humains","Flore fécale"],notes=None,comment=None),
    dict(name="Shigella spp.",path=["Germes","Bactéries","Humains","Flore fécale"],notes=None,comment=None),
    dict(name="Yersinia enterocolitica",path=["Germes","Bactéries","Humains","Flore fécale"],notes=None,comment=None),
    dict(name="Pseudomonas spp.",path=["Germes","Bactéries","Environnemental","Humidité"],notes=None,comment=None),
    dict(name="Acinetobacter spp.",path=["Germes","Bactéries","Environnemental","Humidité"],notes=None,comment=None),
    dict(name="Paenibacillus spp. (SPORULES)",path=["Germes","Bactéries","Environnemental","Sol / Carton / Surface sèche"],notes=None,comment=None),
    dict(name="Hafnia alvei",path=["Germes","Bactéries","Humains","Flore fécale"],notes=None,comment=None),
    dict(name="Sphingomonas paucimobilis",path=["Germes","Bactéries","Environnemental","Humidité"],notes=None,comment=None),
    dict(name="Sphingobium spp.",path=["Germes","Bactéries","Environnemental","Humidité"],notes=None,comment=None),
    dict(name="Methylobacterium spp.",path=["Germes","Bactéries","Environnemental","Humidité"],notes=None,comment=None),
    dict(name="Caulobacter crescentus",path=["Germes","Bactéries","Environnemental","Humidité"],notes=None,comment=None),
    dict(name="Mycobacterium non tuberculeux",path=["Germes","Bactéries","Environnemental","Humidité"],notes=None,comment=None),
    dict(name="Burkholderia cepacia",path=["Germes","Bactéries","Environnemental","Sol / Carton / Surface sèche"],notes=None,comment=None),
    dict(name="Massilia spp.",path=["Germes","Bactéries","Environnemental","Sol / Carton / Surface sèche"],notes=None,comment=None),
    dict(name="Bacillus spp. (SPORULES)",path=["Germes","Bactéries","Environnemental","Sol / Carton / Surface sèche"],notes="Sporulé",comment=None),
    dict(name="Clostridium spp. (SPORULES)",path=["Germes","Bactéries","Environnemental","Sol / Carton / Surface sèche"],notes="Sporulé",comment=None),
    dict(name="Geobacillus stearothermophilus (SPORULES)",path=["Germes","Bactéries","Environnemental","Sol / Carton / Surface sèche"],notes="Sporulé thermophile",comment=None),
    dict(name="Arthrobacter spp.",path=["Germes","Bactéries","Environnemental","Sol / Carton / Surface sèche"],notes=None,comment=None),
    dict(name="Cellulomonas spp.",path=["Germes","Bactéries","Environnemental","Sol / Carton / Surface sèche"],notes=None,comment=None),
    dict(name="Curtobacterium spp.",path=["Germes","Bactéries","Environnemental","Sol / Carton / Surface sèche"],notes=None,comment=None),
    dict(name="Agrococcus spp.",path=["Germes","Bactéries","Environnemental","Sol / Carton / Surface sèche"],notes=None,comment=None),
    dict(name="Microbacterium spp.",path=["Germes","Bactéries","Environnemental","Sol / Carton / Surface sèche"],notes=None,comment=None),
    dict(name="Brevibacterium linens/casei",path=["Germes","Bactéries","Environnemental","Sol / Carton / Surface sèche"],notes=None,comment=None),
    dict(name="Georgenia spp.",path=["Germes","Bactéries","Environnemental","Sol / Carton / Surface sèche"],notes=None,comment=None),
    dict(name="Candida spp.",path=["Germes","Champignons","Humain","Peau / Muqueuse"],notes="Levure",comment="Levures = pas de production de spores"),
    dict(name="Trichosporon spp.",path=["Germes","Champignons","Humain","Peau / Muqueuse"],notes="Levure",comment="Levures = pas de production de spores"),
    dict(name="Rhodotorula spp.",path=["Germes","Champignons","Environnemental","Humidité"],notes="Levure",comment="Levures = pas de production de spores"),
    dict(name="Fusarium spp.",path=["Germes","Champignons","Environnemental","Sol / Carton / Surface sèche"],notes="Conidies 2-4um",comment="Production de spores (conidies) => dissémination dans l'air facilitée par petite taille (2-4 µm) + Résistance aux agents oxydants"),
    dict(name="Aureobasidium spp.",path=["Germes","Champignons","Environnemental","Humidité"],notes="Blastospores + biofilm",comment="Production de blastospores (moins résistante aux agents oxydants et ne se dissémine pas dans l'air comme les conidies d'Aspergillus ou Fusarium) et production de biofilm"),
    dict(name="Mucorales",path=["Germes","Champignons","Environnemental","Sol / Carton / Surface sèche"],notes="Sporangiospores",comment="Production de sporangiospores (moins résistant aux agents oxydants que les conidies d'Aspergillus ou Fusarium)"),
    dict(name="Alternaria spp.",path=["Germes","Champignons","Environnemental","Air"],notes="Conidies grandes",comment="Production de spores (conidies) mais de plus grande taille, moins déshydratées et moins mélanisées que celles de Fusarium ou Aspergillus => Moins résistante à l'oxydation"),
    dict(name="Aspergillus spp.",path=["Germes","Champignons","Environnemental","Air"],notes="Conidies 2-4um",comment="Production de spores (conidies) => dissémination dans l'air facilitée par petite taille (2-4 µm) + Résistance aux agents oxydants"),
    dict(name="Cladosporium spp.",path=["Germes","Champignons","Environnemental","Air"],notes="Conidies grandes",comment="Production de spores (conidies) mais de plus grande taille, moins déshydratées et moins mélanisées => Moins résistante à l'oxydation"),
    dict(name="Penicillium spp.",path=["Germes","Champignons","Environnemental","Air"],notes="Conidies",comment="Production de spores (conidies) mais moins déshydratées et moins mélanisées que celles de Fusarium ou Aspergillus => Moins résistantes à l'oxydation"),
    dict(name="Wallemia sebi",path=["Germes","Champignons","Environnemental","Air"],notes="Arthroconidies",comment=None),
]

DEFAULT_GERM_NAMES = {g["name"] for g in DEFAULT_GERMS}

# ── PERSISTENCE ────────────────────────────────────────────────────────────────
def _supa_upsert(key, value_json):
    supa = get_supabase_client()
    if supa is None:
        return False
    try:
        supa.table('app_state').upsert(
            {'key': key, 'value': value_json}, on_conflict='key'
        ).execute()
        return True
    except Exception as e:
        print(f"[SUPA ERROR] key={key} : {e}")
        return False

def _supa_get(key):
    supa = get_supabase_client()
    if supa is None:
        return None
    try:
        res = supa.table('app_state').select('value').eq('key', key).execute()
        if res and getattr(res, 'data', None):
            row = res.data[0]
            return row.get('value') if isinstance(row, dict) else row['value']
    except Exception:
        pass
    return None

def save_germs(germs):
    known_default_names = sorted(DEFAULT_GERM_NAMES)
    payload = {"germs": germs, "known_defaults": known_default_names}
    js = json.dumps(payload, ensure_ascii=False)
    _supa_upsert('germs', js)
    try:
        with open(GERMS_FILE, "w") as f:
            f.write(js)
    except Exception:
        pass

def load_germs():
    defaults_by_name = {d["name"]: d for d in DEFAULT_GERMS}
    saved_germs = []
    known_defaults = set()
    raw_json = _supa_get('germs')
    if raw_json:
        try:
            raw = json.loads(raw_json)
            if isinstance(raw, dict):
                saved_germs  = raw.get("germs", [])
                known_defaults = set(raw.get("known_defaults", []))
            elif isinstance(raw, list):
                saved_germs = raw
        except Exception:
            saved_germs = []
    if not saved_germs and os.path.exists(GERMS_FILE):
        try:
            with open(GERMS_FILE) as f:
                raw = json.load(f)
            if isinstance(raw, dict):
                saved_germs    = raw.get("germs", [])
                known_defaults = set(raw.get("known_defaults", []))
            elif isinstance(raw, list):
                saved_germs = raw
        except Exception:
            saved_germs = []
    if not saved_germs:
        return [dict(d) for d in DEFAULT_GERMS], len(DEFAULT_GERMS)
    saved_by_name = {g.get("name", ""): g for g in saved_germs}
    merged = []
    new_defaults_added = 0
    for dflt in DEFAULT_GERMS:
        name = dflt["name"]
        if name in saved_by_name:
            merged.append(dict(saved_by_name[name]))
        elif name not in known_defaults:
            merged.append(dict(dflt))
            new_defaults_added += 1
    for g in saved_germs:
        name = g.get("name", "")
        if name and name not in defaults_by_name:
            merged.append(dict(g))
    return merged, new_defaults_added

def load_thresholds():
    saved = {}
    raw_json = _supa_get('thresholds')
    if raw_json:
        try:
            raw = json.loads(raw_json)
            saved = {int(k): v for k, v in raw.items() if k != 'measures'}
        except Exception:
            saved = {}
    if not saved and os.path.exists(THRESHOLDS_FILE):
        try:
            with open(THRESHOLDS_FILE) as f:
                raw = json.load(f)
            saved = {int(k): v for k, v in raw.items() if k != 'measures'}
        except Exception:
            saved = {}
    merged = {k: dict(v) for k, v in DEFAULT_THRESHOLDS.items()}
    merged.update(saved)
    return merged

def load_measures():
    if os.path.exists(THRESHOLDS_FILE):
        try:
            with open(THRESHOLDS_FILE) as f:
                raw = json.load(f)
            if "measures" in raw:
                return {int(k): v for k, v in raw["measures"].items()}
        except Exception:
            pass
    return {k: dict(v) for k, v in DEFAULT_MEASURES.items()}

def save_thresholds_and_measures(thresholds, measures):
    data = {str(k): v for k, v in thresholds.items()}
    data["measures"] = {str(k): v for k, v in measures.items()}
    js = json.dumps(data, ensure_ascii=False)
    _supa_upsert('thresholds', js)
    try:
        with open(THRESHOLDS_FILE, "w") as f:
            f.write(js)
    except Exception:
        pass

def get_thresholds_for_risk(risk, thresholds):
    return thresholds.get(risk, {"alert": 25, "action": 40})

def load_origin_measures():
    raw_json = _supa_get('measures')
    if raw_json:
        try:
            raw = json.loads(raw_json)
            if isinstance(raw, list) and raw:
                return raw
        except Exception:
            pass
    if os.path.exists(THRESHOLDS_FILE):
        try:
            with open(THRESHOLDS_FILE) as f:
                raw = json.load(f)
            if "measures" in raw and isinstance(raw["measures"], list) and raw["measures"]:
                return raw["measures"]
        except Exception:
            pass
    return [dict(m) for m in DEFAULT_ORIGIN_MEASURES]

def save_origin_measures(measures, supa=True):
    if supa:
        try:
            result = _supa_upsert('measures', json.dumps(measures, ensure_ascii=False))
            if not result:
                st.warning("⚠️ Supabase non connecté — sauvegarde locale uniquement.")
        except Exception as e:
            st.error(f"❌ Erreur Supabase : {e}")
    if os.path.exists(THRESHOLDS_FILE):
        try:
            with open(THRESHOLDS_FILE) as f:
                raw = json.load(f)
        except Exception:
            raw = {}
    else:
        raw = {}
    raw["measures"] = measures
    try:
        with open(THRESHOLDS_FILE, "w") as f:
            json.dump(raw, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def _load_json_key(key, local_file):
    raw_json = _supa_get(key)
    if raw_json:
        try:
            raw = json.loads(raw_json)
            if isinstance(raw, list):
                return [dict(x) for x in raw]
        except Exception:
            pass
    if os.path.exists(local_file):
        try:
            with open(local_file, 'r', encoding='utf-8') as f:
                raw = json.load(f)
            if isinstance(raw, list):
                return [dict(x) for x in raw]
        except Exception:
            pass
    return []

def _save_json_key(key, data, local_file):
    js = json.dumps(data, ensure_ascii=False)
    _supa_upsert(key, js)
    try:
        with open(local_file, 'w', encoding='utf-8') as f:
            f.write(js)
    except Exception:
        pass

def load_points():  return _load_json_key('points', POINTS_FILE)
def save_points(d, supa=True):
    _save_json_key('points', d, POINTS_FILE)
    if supa: _supa_upsert('points', json.dumps(d, ensure_ascii=False))

def load_prelevements():  return _load_json_key('prelevements', PRELEVEMENTS_FILE)
def save_prelevements(d, supa=True):
    _save_json_key('prelevements', d, PRELEVEMENTS_FILE)
    if supa: _supa_upsert('prelevements', json.dumps(d, ensure_ascii=False))

def load_schedules():  return _load_json_key('schedules', SCHEDULES_FILE)
def save_schedules(d, supa=True):
    _save_json_key('schedules', d, SCHEDULES_FILE)
    if supa: _supa_upsert('schedules', json.dumps(d, ensure_ascii=False))

def load_pending_identifications():  return _load_json_key('pending_identifications', PENDING_FILE)
def save_pending_identifications(d, supa=True):
    _save_json_key('pending_identifications', d, PENDING_FILE)
    if supa: _supa_upsert('pending_identifications', json.dumps(d, ensure_ascii=False))

def load_archived_samples():  return _load_json_key('archived_samples', ARCHIVED_FILE)
def save_archived_samples(d, supa=True):
    _save_json_key('archived_samples', d, ARCHIVED_FILE)
    if supa: _supa_upsert('archived_samples', json.dumps(d, ensure_ascii=False))

def load_operators():  return _load_json_key('operators', OPERATORS_FILE)
def save_operators(d, supa=True):
    _save_json_key('operators', d, OPERATORS_FILE)
    if supa: _supa_upsert('operators', json.dumps(d, ensure_ascii=False))

def load_plans():  return _load_json_key('plans', PLANS_FILE)
def save_plans(d, supa=True):
    _save_json_key('plans', d, PLANS_FILE)
    if supa: _supa_upsert('plans', json.dumps(d, ensure_ascii=False))
def compute_germ_score(g):
    gobj = next((x for x in st.session_state.germs if x['name'] == g["germ"]), None)
    if gobj:
        return (int(gobj.get('pathogenicity', 1))
                * int(gobj.get('resistance', 1))
                * int(gobj.get('dissemination', 1)))
    return 1

def load_faq():
    raw_json = _supa_get('faq')
    if raw_json:
        try:
            raw = json.loads(raw_json)
            if isinstance(raw, list) and raw:
                return raw
        except Exception:
            pass
    return [dict(f) for f in DEFAULT_FAQ]

def save_faq(faq_items, supa=True):
    if supa:
        _supa_upsert('faq', json.dumps(faq_items, ensure_ascii=False))

def load_surveillance():
    raw_json = _supa_get('surveillance')
    if raw_json:
        try:
            return json.loads(raw_json)
        except Exception:
            pass
    if os.path.exists(CSV_FILE):
        try:
            rows = []
            with open(CSV_FILE, newline='', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    rows.append(row)
            return rows
        except Exception:
            pass
    return []

def save_surveillance(records):
    js = json.dumps(records, ensure_ascii=False)
    _supa_upsert('surveillance', js)
    try:
        if not records:
            if os.path.exists(CSV_FILE):
                os.remove(CSV_FILE)
        else:
            all_keys = list(dict.fromkeys(k for r in records for k in r.keys()))
            with open(CSV_FILE, "w", newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=all_keys, extrasaction="ignore")
                writer.writeheader()
                writer.writerows(records)
    except Exception:
        pass

def export_all_data():
    return {
        "_meta": {
            "version": "5.0",
            "exported_at": datetime.now().isoformat(),
            "app": "MicroSurveillance URC"
        },
        "germs":                   st.session_state.germs,
        "thresholds":              {str(k): v for k, v in st.session_state.thresholds.items()},
        "measures":                {str(k): v for k, v in st.session_state.measures.items()},
        "origin_measures":         st.session_state.origin_measures,
        "points":                  st.session_state.points,
        "operators":               st.session_state.operators,
        "plans":                   st.session_state.plans,
        "prelevements":            st.session_state.prelevements,
        "schedules":               st.session_state.schedules,
        "pending_identifications": st.session_state.pending_identifications,
        "archived_samples":        st.session_state.archived_samples,
        "surveillance":            st.session_state.surveillance,
        "planning_overrides":      st.session_state.get("planning_overrides", {}),
    }

def import_all_data(data: dict):
    try:
        required = ["germs", "points", "operators", "prelevements", "schedules", "surveillance"]
        for key in required:
            if key not in data:
                return False, f"Clé manquante dans le fichier : '{key}'"
        st.session_state.germs = [dict(g) for g in data["germs"]]
        _germs_sync, _ = load_germs()
        _default_names = {d["name"] for d in DEFAULT_GERMS}
        _custom = [g for g in st.session_state.germs if g.get("name") not in _default_names]
        st.session_state.germs = _germs_sync + [g for g in _custom if g.get("name") not in {x["name"] for x in _germs_sync}]
        st.session_state.origin_measures         = [dict(m) for m in data.get("origin_measures", [])] or [dict(m) for m in DEFAULT_ORIGIN_MEASURES]
        st.session_state.points                  = [dict(p) for p in data.get("points", [])]
        st.session_state.operators               = [dict(o) for o in data.get("operators", [])]
        st.session_state.plans                   = [dict(p) for p in data.get("plans", [])]
        st.session_state.prelevements            = [dict(p) for p in data.get("prelevements", [])]
        st.session_state.schedules               = [dict(s) for s in data.get("schedules", [])]
        st.session_state.pending_identifications = [dict(x) for x in data.get("pending_identifications", [])]
        st.session_state.archived_samples        = [dict(x) for x in data.get("archived_samples", [])]
        st.session_state.surveillance            = list(data.get("surveillance", []))
        if "thresholds" in data:
            st.session_state.thresholds = {int(k): v for k, v in data["thresholds"].items()}
        if "measures" in data:
            st.session_state.measures   = {int(k): v for k, v in data["measures"].items()}
        save_germs(st.session_state.germs)
        save_origin_measures(st.session_state.origin_measures)
        save_points(st.session_state.points)
        save_operators(st.session_state.operators)
        save_plans(st.session_state.plans)
        save_prelevements(st.session_state.prelevements)
        save_schedules(st.session_state.schedules)
        save_pending_identifications(st.session_state.pending_identifications)
        save_archived_samples(st.session_state.archived_samples)
        save_surveillance(st.session_state.surveillance)
        save_thresholds_and_measures(st.session_state.thresholds, st.session_state.measures)
        n_p = len(st.session_state.prelevements)
        n_s = len(st.session_state.schedules)
        n_g = len(st.session_state.germs)
        n_h = len(st.session_state.surveillance)
        return True, (f"Restauration réussie — {n_g} germes, {n_p} prélèvements, "
                      f"{n_s} lectures planifiées, {n_h} entrées historique.")
    except Exception as e:
        return False, f"Erreur lors de la restauration : {e}"

def find_germ_match(query, germs):
    query_low  = query.lower().strip()
    query_genus = query_low.split()[0] if query_low else ""
    best_score = 0
    best_match = None
    for g in germs:
        name_low = g["name"].lower()
        genus    = name_low.split()[0]
        if query_genus and query_genus == genus:
            score = 0.9
        else:
            score       = difflib.SequenceMatcher(None, query_low, name_low).ratio()
            genus_score = difflib.SequenceMatcher(None, query_genus, genus).ratio()
            score       = max(score, genus_score * 0.85)
        if score > best_score:
            best_score = score
            best_match = g
    return best_match, best_score

def load_planning_skips():
    raw_json = _supa_get('planning_skips')
    if raw_json:
        try:
            raw = json.loads(raw_json)
            if isinstance(raw, dict):
                return raw
        except Exception:
            pass
    return {}

def save_planning_skips(skips):
    _supa_upsert('planning_skips', json.dumps(skips, ensure_ascii=False))

# ── Helpers scoring ────────────────────────────────────────────────────────────
def _get_location_criticality(sample):
    if "location_criticality" in sample:
        try:
            return int(sample["location_criticality"])
        except Exception:
            pass
    pt = next((p for p in st.session_state.points
               if p.get("label") == sample.get("label")), None)
    if pt and "location_criticality" in pt:
        try:
            return int(pt["location_criticality"])
        except Exception:
            pass
    rc = str(sample.get("room_class", "")).strip().upper()
    return {"A": 3, "B": 2, "C": 2, "D": 1}.get(rc, 1)

def _get_germ_score(germ):
    if all(k in germ for k in ("pathogenicity", "resistance", "dissemination")):
        return int(germ["pathogenicity"]) * int(germ["resistance"]) * int(germ["dissemination"])
    old = germ.get("risk", 1)
    return {1: 1, 2: 2, 3: 6, 4: 12, 5: 18}.get(old, old)

def _evaluate_score(total):
    _sa = st.session_state.get("_seuil_alerte", 24)
    _sc = st.session_state.get("_seuil_action", 36)
    if total > _sc:  return "action", "🚨 ACTION",  "#ef4444"
    if total >= _sa: return "alert",  "⚠️ ALERTE",  "#f59e0b"
    return "ok", "✅ Conforme", "#22c55e"

def _loc_crit_label(n):
    return {1: "Limité", 2: "Modéré", 3: "Important", 4: "Critique"}.get(n, str(n))

# ── HELPER : valider une lecture comme NÉGATIVE ────────────────────────────────
def _valider_negatif(proc_id):
    proc = next((x for x in st.session_state.schedules if x['id'] == proc_id), None)
    if not proc:
        return
    smp = next((p for p in st.session_state.prelevements if p['id'] == proc['sample_id']), None)

    proc["status"]    = "done"
    proc["colonies"]  = 0
    proc["date_read"] = str(datetime.today().date())

    if proc["when"] == "J2":
        pass

    elif proc["when"] == "J7" and smp and not smp.get("archived"):
        # ── Calcul criticité inline (évite dépendance à _get_location_criticality) ──
        _lc = 1
        _label_smp = smp.get("label", "")
        if _label_smp:
            _pt_fr = next(
                (p for p in st.session_state.points if p.get("label") == _label_smp), None
            )
            if _pt_fr:
                try:
                    _lc = int(_pt_fr.get("location_criticality", 1))
                except (ValueError, TypeError):
                    _lc = 1
        if _lc == 1:
            try:
                _lc = int(smp.get("location_criticality", 1) or 1)
            except (ValueError, TypeError):
                _lc = 1

        smp["archived"] = True
        st.session_state.archived_samples.append(smp)
        save_archived_samples(st.session_state.archived_samples)
        save_prelevements(st.session_state.prelevements)

        st.session_state.surveillance.append({
            "date":                 str(datetime.today().date()),
            "prelevement":          smp.get("label", "?"),
            "sample_id":            proc["sample_id"],
            "germ_match":           "Négatif",
            "germ_saisi":           "Négatif",
            "ufc":                  0,
            "ufc_total":            0,
            "germ_score":           0,
            "location_criticality": _lc,
            "total_score":          0,
            "status":               "ok",
            "operateur":            smp.get("operateur", "?"),
            "remarque":             "",
            "readings":             "J7",
            "room_class":           smp.get("room_class", ""),
        })
        save_surveillance(st.session_state.surveillance)

    save_schedules(st.session_state.schedules)
# ── CONTRÔLE D'ACCÈS PROTÉGÉ ───────────────────────────────────────────────────
def check_access_protege(onglet_nom: str) -> bool:
    key_mode = f"access_mode_{onglet_nom}"
    key_pwd  = f"pwd_input_{onglet_nom}"
    key_err  = f"pwd_error_{onglet_nom}"
    if st.session_state.get(key_mode) == "admin":
        col_info, col_lock = st.columns([5, 1])
        with col_info:
            st.success("🔓 Mode administrateur — modifications autorisées")
        with col_lock:
            if st.button("🔒 Verrouiller", key=f"lock_{onglet_nom}", use_container_width=True):
                st.session_state[key_mode] = None
                st.rerun()
        return True
    if st.session_state.get(key_mode) == "lecture":
        col_info, col_conn = st.columns([5, 1])
        with col_info:
            st.info("👁️ Mode lecture seule — aucune modification possible")
        with col_conn:
            if st.button("🔑 Se connecter", key=f"connect_{onglet_nom}", use_container_width=True):
                st.session_state[key_mode] = None
                st.rerun()
        return False
    st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)
    st.markdown(
        "<div style='background:#fff;border:1.5px solid #e2e8f0;border-radius:14px;"
        "padding:28px 32px;max-width:460px;margin:0 auto;box-shadow:0 4px 20px rgba(0,0,0,.08)'>"
        "<div style='text-align:center;font-size:2.2rem;margin-bottom:8px'>🔐</div>"
        f"<div style='text-align:center;font-weight:800;font-size:1.15rem;color:#0f172a'>"
        f"Accès protégé — {onglet_nom}</div>"
        "<div style='text-align:center;font-size:.85rem;color:#64748b;margin:8px 0 20px'>"
        "Cet onglet est restreint. Connectez-vous pour modifier,<br>ou continuez en lecture seule.</div>"
        "</div>",
        unsafe_allow_html=True,
    )
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**🔑 Accès administrateur**")
        pwd = st.text_input(
            "Mot de passe", type="password", key=key_pwd,
            placeholder="Entrez le mot de passe", label_visibility="collapsed"
        )
        if st.button("✅ Connexion", key=f"btn_admin_{onglet_nom}",
                     use_container_width=True, type="primary"):
            if pwd == MOT_DE_PASSE_ADMIN:
                st.session_state[key_mode] = "admin"
                st.session_state[key_err]  = False
                st.rerun()
            else:
                st.session_state[key_err] = True
                st.rerun()
        if st.session_state.get(key_err):
            st.error("❌ Mot de passe incorrect")
    with col2:
        st.markdown("**👁️ Lecture seule**")
        st.markdown(
            "<div style='font-size:.82rem;color:#64748b;margin-bottom:10px'>"
            "Consultez le contenu sans pouvoir effectuer de modifications.</div>",
            unsafe_allow_html=True,
        )
        if st.button("👁️ Continuer en lecture", key=f"btn_lecture_{onglet_nom}",
                     use_container_width=True):
            st.session_state[key_mode] = "lecture"
            st.rerun()
    st.stop()
    return False

# ── SESSION STATE ──────────────────────────────────────────────────────────────
if "germs" not in st.session_state:
    _germs, _new = load_germs()
    st.session_state.germs = _germs
    st.session_state.germs_synced_count = _new
    if _new > 0:
        save_germs(st.session_state.germs)
if "thresholds" not in st.session_state:
    st.session_state.thresholds = load_thresholds()
if "measures" not in st.session_state:
    st.session_state.measures = load_measures()
if "surveillance" not in st.session_state:
    st.session_state.surveillance = load_surveillance()
if "show_add" not in st.session_state:
    st.session_state.show_add = False
if "edit_idx" not in st.session_state:
    st.session_state.edit_idx = None
if "map_points" not in st.session_state:
    st.session_state.map_points = []
if "map_image" not in st.session_state:
    st.session_state.map_image = None
if "active_tab" not in st.session_state:
    st.session_state.active_tab = "accueil"
if "origin_measures" not in st.session_state:
    st.session_state.origin_measures = load_origin_measures()
if "show_new_measure" not in st.session_state:
    st.session_state.show_new_measure = False
if "prelevements" not in st.session_state:
    st.session_state.prelevements = load_prelevements()
if "schedules" not in st.session_state:
    st.session_state.schedules = load_schedules()
if "pending_identifications" not in st.session_state:
    st.session_state.pending_identifications = load_pending_identifications()
if "archived_samples" not in st.session_state:
    st.session_state.archived_samples = load_archived_samples()
if "points" not in st.session_state:
    st.session_state.points = load_points()
if "operators" not in st.session_state:
    st.session_state.operators = load_operators()
if "plans" not in st.session_state:
    st.session_state.plans = load_plans()
if "planning_skips" not in st.session_state:
    st.session_state["planning_skips"] = load_planning_skips()
if "faq_items" not in st.session_state:
    st.session_state.faq_items = load_faq()
if "_seuil_alerte" not in st.session_state:
    _raw_seuils = _supa_get('seuils')
    if _raw_seuils:
        try:
            _s = json.loads(_raw_seuils)
            st.session_state["_seuil_alerte"] = int(_s.get("alerte", 24))
            st.session_state["_seuil_action"] = int(_s.get("action", 36))
        except Exception:
            st.session_state["_seuil_alerte"] = 24
            st.session_state["_seuil_action"] = 36
    else:
        st.session_state["_seuil_alerte"] = 24
        st.session_state["_seuil_action"] = 36
if "due_alert_shown" not in st.session_state:
    st.session_state.due_alert_shown = False
if "current_process" not in st.session_state:
    st.session_state.current_process = None
if "cal_year" not in st.session_state:
    st.session_state.cal_year = datetime.today().year
if "cal_month" not in st.session_state:
    st.session_state.cal_month = datetime.today().month
if "planning_overrides" not in st.session_state:
    raw = _supa_get('planning_overrides')
    st.session_state["planning_overrides"] = json.loads(raw) if raw else {}
if "class_constraints_loaded" not in st.session_state:
    raw_cc = _supa_get('class_constraints')
    if raw_cc:
        try:
            for cls, val in json.loads(raw_cc).items():
                st.session_state[f"class_max_{cls}"] = int(val)
        except Exception:
            pass
    st.session_state["class_constraints_loaded"] = True

# ── SIDEBAR ────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="MicroSurveillance URC",
    layout="wide",
    initial_sidebar_state="collapsed"
)
with st.sidebar:
    st.markdown(
        '<p style="font-size:.85rem;letter-spacing:.1em;text-transform:uppercase;'
        'color:#94a3b8;margin-bottom:12px;font-weight:700">NAVIGATION</p>',
        unsafe_allow_html=True,
    )
    tabs_cfg = [
        ("accueil",      "🏠", "Accueil"),
        ("Base de données",   "🦠", "Base de données"),
        ("surveillance", "🔍", "Identification & Surveillance"),
        ("planning",     "📅", "Planning"),
        ("analyse",   "📋", "Analyse"),
        ("parametres",   "⚙️", "Paramètres & Seuils"),
    ]
    for key, icon, label in tabs_cfg:
        if st.button(f"{icon}  {label}", key=f"nav_{key}", use_container_width=True):
            st.session_state.active_tab = key
            st.rerun()

    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
    col_gif, col_btn = st.sidebar.columns([1, 2])
    with col_gif:
        st.components.v1.html("""
<div style="display:flex;justify-content:center;align-items:center;padding-top:4px">
  <iframe src="https://giphy.com/embed/bSEkPdQfsSHCMYn7fD"
          width="90" height="90"
          style="border:none;pointer-events:none;display:block"
          frameBorder="0"></iframe>
</div>
""", height=100, scrolling=False)

    with col_btn:
        st.markdown("<div style='height:22px'></div>", unsafe_allow_html=True)
        if st.button(
            "Si tu as besoin\nd'aide, je suis là !",
            key="faq_open_btn",
            use_container_width=True,
        ):
            st.session_state.active_tab = "faq"
            st.rerun()

    st.divider()

    supa_ok   = get_supabase_client() is not None
    supa_icon = "🟢" if supa_ok else "🔴"
    supa_txt  = "Supabase connecté" if supa_ok else "Mode local (fichiers)"
    st.markdown(
        f'<p style="font-size:.7rem;color:#94a3b8;text-align:center">{supa_icon} {supa_txt}</p>',
        unsafe_allow_html=True,
    )
    st.divider()

    st.markdown(
        '<p style="font-size:.7rem;color:#f59e0b;font-weight:700;text-align:center;'
        'text-transform:uppercase;letter-spacing:.08em">💾 Sauvegarde données</p>',
        unsafe_allow_html=True,
    )
    _backup_data = json.dumps(export_all_data(), ensure_ascii=False, indent=2)
    _backup_name = f"backup_URC_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
    st.download_button(
        label="⬇️ Exporter toutes les données",
        data=_backup_data,
        file_name=_backup_name,
        mime="application/json",
        use_container_width=True,
        key="sidebar_export",
        help="Téléchargez ce fichier avant toute modification du code.",
    )
    if not supa_ok:
        st.markdown(
            '<p style="font-size:.6rem;color:#f59e0b;text-align:center;margin-top:4px">'
            '⚠️ Sans Supabase, exportez régulièrement vos données !</p>',
            unsafe_allow_html=True,
        )

    st.divider()

# ── HEADER ─────────────────────────────────────────────────────────────────────
active = st.session_state.active_tab
today  = datetime.today().date()

st.markdown(
    '<h1 style="font-size:1.3rem;letter-spacing:.1em;text-transform:uppercase;'
    'color:#1e40af!important;margin-bottom:0">🦠 MicroSurveillance URC</h1>',
    unsafe_allow_html=True,
)
st.caption("Surveillance microbiologique — Unité de Reconstitution des Chimiothérapies")

# ── ONGLET FAQ ─────────────────────────────────────────────────────────────────
if active == "faq":
    from collections import defaultdict
    import re as _re

    faq_items = sorted(
        st.session_state.get("faq_items", DEFAULT_FAQ),
        key=lambda x: x.get("order", 999),
    )

    if st.button("← Retour", key="faq_back"):
        st.session_state.active_tab = "Base de données"
        st.rerun()

    st.markdown(
        "<div style='background:linear-gradient(135deg,#7c3aed,#a855f7);"
        "border-radius:14px;padding:20px 28px;margin:12px 0 16px 0'>"
        "<div style='color:#fff;font-size:1.3rem;font-weight:900'>🍄 Centre d'aide — FAQ</div>"
        f"<div style='color:#e9d5ff;font-size:.8rem;margin-top:3px'>"
        f"{len(faq_items)} questions & réponses disponibles</div>"
        "</div>",
        unsafe_allow_html=True)

    fc1, fc2 = st.columns([3, 1])
    with fc1:
        faq_query = st.text_input(
            "search", placeholder="🔍 Rechercher...",
            label_visibility="collapsed", key="faq_page_search")
    with fc2:
        all_cats_p = ["Toutes les catégories"] + sorted(
            set(f.get("category", "Général") for f in faq_items))
        sel_cat_p = st.selectbox(
            "cat", all_cats_p,
            label_visibility="collapsed", key="faq_page_cat")

    st.markdown(
        "<hr style='margin:8px 0 16px;border-color:#e2e8f0'>",
        unsafe_allow_html=True)

    q_p = faq_query.strip().lower()

    def _faq_match(item):
        if sel_cat_p != "Toutes les catégories" and item.get("category") != sel_cat_p:
            return False
        if q_p:
            return (q_p in item["question"].lower()
                    or q_p in item["answer"].lower()
                    or q_p in item.get("category", "").lower())
        return True

    filtered = [f for f in faq_items if _faq_match(f)]

    if not filtered:
        st.markdown(
            "<div style='text-align:center;padding:48px 0;color:#94a3b8'>"
            "<div style='font-size:2.5rem;margin-bottom:10px'>🔍</div>"
            f"<div style='font-size:.9rem'>Aucun résultat pour "
            f"<strong>« {faq_query} »</strong></div></div>",
            unsafe_allow_html=True)
    else:
        CAT_COLORS = {
            "Général":            "#2563eb",
            "Score & Seuils":     "#7c3aed",
            "Prélèvements":       "#0891b2",
            "Paramètres":         "#059669",
            "Données":            "#d97706",
            "Mesures correctives":"#dc2626",
        }
        grouped = defaultdict(list)
        for item in filtered:
            grouped[item.get("category", "Général")].append(item)

        cats_sorted = sorted(grouped.keys())
        mid = (len(cats_sorted) + 1) // 2
        col_left, col_right = st.columns(2)

        for ci, cat in enumerate(cats_sorted):
            col = col_left if ci < mid else col_right
            c   = CAT_COLORS.get(cat, "#475569")
            with col:
                st.markdown(
                    f"<div style='display:inline-block;background:{c}12;color:{c};"
                    f"border:1px solid {c}44;border-radius:20px;padding:3px 14px;"
                    f"font-size:.68rem;font-weight:700;text-transform:uppercase;"
                    f"letter-spacing:.06em;margin:12px 0 6px'>📂 {cat}</div>",
                    unsafe_allow_html=True)
                for item in grouped[cat]:
                    q_display = item["question"]
                    if q_p:
                        q_display = _re.sub(
                            f"({_re.escape(faq_query)})",
                            r"<mark style='background:#fef08a;border-radius:3px;"
                            r"padding:0 2px'>\1</mark>",
                            q_display, flags=_re.IGNORECASE)
                    with st.expander(item["question"]):
                        st.markdown(item["answer"])

    st.markdown(
        "<div style='margin-top:24px;padding:12px 16px;background:#f8fafc;"
        "border:1px solid #e2e8f0;border-radius:10px;"
        "font-size:.72rem;color:#94a3b8;text-align:center'>"
        "Vous ne trouvez pas votre réponse ? Contactez votre pharmacien référent.</div>",
        unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB : ACCUEIL
# ═══════════════════════════════════════════════════════════════════════════════

if active == "accueil":
    st.markdown("""
    <style>
    .home-grid {
        display: grid;
        grid-template-columns: 1fr 1fr;
        gap: 0;
        min-height: 420px;
    }
    .home-card {
        display: flex;
        flex-direction: column;
        align-items: center;
        justify-content: center;
        padding: 2.5rem 1.5rem;
        border: 1px solid #e2e8f0;
        text-align: center;
    }
    .home-card:first-child { border-radius: 12px 0 0 12px; border-right: none; }
    .home-card:last-child  { border-radius: 0 12px 12px 0; }
    .home-card img { width: 440px; height: 440px; object-fit: contain; border-radius: 8px; margin-bottom: 1rem; }
    .home-card h2 { font-size: 1.5rem; font-weight: 500; margin: 0 0 0.4rem; }
    .home-card p  { font-size: 0.85rem; color: #64748b; margin: 0; }
    </style>

    <div class="home-grid">
      <div class="home-card">
        <img src="https://media4.giphy.com/media/3oKIPj0RCvQEGuNQFG/giphy.webp" />
        <h2>Planning</h2>
        <p>Gérer le calendrier des tâches</p>
      </div>
      <div class="home-card">
        <img src="https://i.pinimg.com/originals/b2/4c/9e/b24c9e272f886980409584f6299b12ed.gif" />
        <h2>Prélèvements</h2>
        <p>Identification et surveillance</p>
      </div>
    </div>
    """, unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        if st.button("📅 Accéder au Planning", use_container_width=True):
            st.session_state.active_tab = "planning"
            st.rerun()
    with col2:
        if st.button("🔍 Accéder aux Prélèvements", use_container_width=True):
            st.session_state.active_tab = "surveillance"
            st.rerun()

# ═══════════════════════════════════════════════════════════════════════════════
# TAB : BASE DE DONNEES
# ═══════════════════════════════════════════════════════════════════════════════

if active == "Base de données":
    can_edit = check_access_protege("Base de données")

    PATHO_OPTS = [
        "1 — Non pathogène",
        "2 — Pathogène opportuniste",
        "3 — Pathogène opportuniste multirésistant / Pathogène primaire",
    ]
    RESIST_OPTS = [
        "1 — Sensible",
        "2 — Résistant au Surfa'Safe",
        "3 — Résistant au Surfa'Safe et Acide Peracétique",
    ]
    DISSEM_OPTS = [
        "1 — Environnemental",
        "2 — Manuporté",
        "3 — Aéroporté",
    ]

    def _risk_color(score):
        if score <= 4:   return "#22c55e"
        if score <= 8:   return "#84cc16"
        if score <= 12:  return "#f59e0b"
        if score <= 18:  return "#f97316"
        return "#ef4444"

    def _risk_label(score):
        if score <= 4:   return "Faible"
        if score <= 8:   return "Modéré"
        if score <= 12:  return "Important"
        if score <= 18:  return "Majeur"
        return "Critique"

    def _infer_resistance(surfa, apa):
        s = (surfa or "").lower()
        a = (apa  or "").lower()
        if "risque" in s and "risque" in a:
            return 3
        if "risque" in s:
            return 2
        return 1

    def _germ_score(g):
        if all(k in g for k in ("pathogenicity", "resistance", "dissemination")):
            return int(g["pathogenicity"]) * int(g["resistance"]) * int(g["dissemination"])
        old = g.get("risk", 1)
        return {1: 1, 2: 2, 3: 6, 4: 12, 5: 18}.get(old, old)

    _synced        = st.session_state.get("germs_synced_count", 0)
    _total_default = len(DEFAULT_GERMS)
    _total_germs   = len(st.session_state.germs)
    _custom_count  = max(0, _total_germs - _total_default)

    st.markdown(f"""
    <div style="background:linear-gradient(135deg,#eff6ff,#dbeafe);border:1.5px solid #93c5fd;
    border-radius:12px;padding:12px 18px;margin-bottom:14px;display:flex;align-items:center;
    justify-content:space-between;flex-wrap:wrap;gap:10px">
      <div style="display:flex;align-items:center;gap:14px;flex-wrap:wrap">
        <span style="font-size:1.4rem">🦠</span>
        <div>
          <div style="font-weight:800;color:#1e40af;font-size:.88rem">Base de données germes</div>
          <div style="font-size:.7rem;color:#3b82f6;margin-top:2px">
            {_total_default} standards · {_custom_count} personnalisé(s) · {_total_germs} au total
          </div>
          <div style="font-size:.65rem;color:#64748b;margin-top:1px">
            Score criticité = Pathogénicité × Résistance × Dissémination (1–27)
          </div>
        </div>
      </div>
      <div style="display:flex;gap:8px;flex-wrap:wrap">
        <div style="background:#fff;border:1px solid #93c5fd;border-radius:8px;padding:6px 12px;text-align:center">
          <div style="font-size:.6rem;color:#1e40af;text-transform:uppercase;font-weight:700">Standard</div>
          <div style="font-size:1.1rem;font-weight:800;color:#1e40af">{_total_default}</div>
        </div>
        <div style="background:#fff;border:1px solid #86efac;border-radius:8px;padding:6px 12px;text-align:center">
          <div style="font-size:.6rem;color:#166534;text-transform:uppercase;font-weight:700">Custom</div>
          <div style="font-size:1.1rem;font-weight:800;color:#166534">{_custom_count}</div>
        </div>
        <div style="background:#fff;border:1px solid #fcd34d;border-radius:8px;padding:6px 12px;text-align:center">
          <div style="font-size:.6rem;color:#92400e;text-transform:uppercase;font-weight:700">Total</div>
          <div style="font-size:1.1rem;font-weight:800;color:#92400e">{_total_germs}</div>
        </div>
      </div>
    </div>""", unsafe_allow_html=True)

    if _synced > 0:
        st.success(f"✅ Synchronisation : {_synced} germe(s) mis à jour depuis la base de référence.")

    col_btn1, col_btn2 = st.columns(2)
    with col_btn1:
        if can_edit:
            if st.button("➕ Ajouter un germe", use_container_width=True):
                st.session_state.show_add = not st.session_state.show_add
                st.session_state.edit_idx = None
    with col_btn2:
        if can_edit:
            if st.button("💾 Sauvegarder", use_container_width=True, key="save_germs_btn"):
                save_germs(st.session_state.germs)
                st.success("✅ Germes sauvegardés !")

    if not can_edit:
        st.info("👁️ Mode lecture seule — connectez-vous pour modifier les germes.")

    def germ_form(existing=None, idx=None):
        is_edit = existing is not None
        with st.container():
            st.markdown(
                f"<div style='background:#f8fafc;border:1.5px solid #e2e8f0;border-radius:12px;"
                f"padding:18px;margin-bottom:12px'>",
                unsafe_allow_html=True)
            st.markdown(f"### {'✏️ Modifier' if is_edit else '➕ Ajouter'} un germe")

            c1, c2, c3 = st.columns(3)

            with c1:
                st.markdown(
                    "<div style='font-size:.72rem;font-weight:800;color:#1e40af;"
                    "letter-spacing:.06em;text-transform:uppercase;margin-bottom:6px'>"
                    "📋 Identification</div>", unsafe_allow_html=True)

                new_name = st.text_input(
                    "Nom du germe *",
                    value=existing["name"] if is_edit else "",
                    placeholder="Ex: Listeria spp.",
                    disabled=not can_edit)

                new_famille = st.selectbox(
                    "Famille *", ["Bactéries", "Champignons"],
                    index=["Bactéries", "Champignons"].index(existing["path"][1])
                          if is_edit else 0,
                    disabled=not can_edit)

                new_origine = st.selectbox(
                    "Origine *", ["Humains ", "Environnemental"],
                    index=0 if (not is_edit or existing["path"][2] in ["Humains", "Humain"]) else 1,
                    disabled=not can_edit)

                if new_famille == "Bactéries":
                    cats = (["Peau / Muqueuse", "Oropharynx / Gouttelettes", "Flore fécale"]
                            if "Humain" in new_origine
                            else ["Humidité", "Sol / Carton / Surface sèche"])
                else:
                    cats = (["Peau / Muqueuse"]
                            if "Humain" in new_origine
                            else ["Humidité", "Sol / Carton / Surface sèche", "Air"])

                cur_cat = existing["path"][3] if is_edit and existing["path"][3] in cats else cats[0]
                new_cat = st.selectbox(
                    "Catégorie *", cats,
                    index=cats.index(cur_cat) if cur_cat in cats else 0,
                    disabled=not can_edit)

            with c2:
                st.markdown(
                    "<div style='font-size:.72rem;font-weight:800;color:#1e40af;"
                    "letter-spacing:.06em;text-transform:uppercase;margin-bottom:6px'>"
                    "🔬 Critères de criticité</div>", unsafe_allow_html=True)

                cur_patho = int(existing.get("pathogenicity", 1)) if is_edit else 1
                patho_lbl = st.selectbox(
                    "🧬 Pathogénicité *", PATHO_OPTS,
                    index=min(cur_patho - 1, 2),
                    key="form_patho", disabled=not can_edit)
                patho_num = int(patho_lbl[0])

                if is_edit:
                    cur_resist = int(existing.get(
                        "resistance",
                        _infer_resistance(existing.get("surfa"), existing.get("apa"))))
                else:
                    cur_resist = 1
                resist_lbl = st.selectbox(
                    "🧴 Résistance aux désinfectants *", RESIST_OPTS,
                    index=min(cur_resist - 1, 2),
                    key="form_resist", disabled=not can_edit)
                resist_num = int(resist_lbl[0])

                cur_dissem = int(existing.get("dissemination", 1)) if is_edit else 1
                dissem_lbl = st.selectbox(
                    "💨 Mode de dissémination *", DISSEM_OPTS,
                    index=min(cur_dissem - 1, 2),
                    key="form_dissem", disabled=not can_edit)
                dissem_num = int(dissem_lbl[0])

                risk_num = patho_num * resist_num * dissem_num
                rc = _risk_color(risk_num)
                rl = _risk_label(risk_num)

                if risk_num > 36:
                    status_txt = "🚨 Action probable si lieu ≥ critique"
                    status_bg  = "#fef2f2"
                    status_c   = "#991b1b"
                elif risk_num >= 24:
                    status_txt = "⚠️ Alerte probable si lieu ≥ modéré"
                    status_bg  = "#fffbeb"
                    status_c   = "#92400e"
                else:
                    status_txt = "✅ Conforme selon lieu de prélèvement"
                    status_bg  = "#f0fdf4"
                    status_c   = "#166534"

                st.markdown(f"""
                <div style="background:{rc}12;border:2px solid {rc}55;border-radius:12px;
                padding:14px;margin-top:8px;text-align:center">
                  <div style="font-size:.62rem;color:#475569;text-transform:uppercase;
                  letter-spacing:.1em;margin-bottom:6px;font-weight:700">Score de criticité</div>
                  <div style="font-size:2.8rem;font-weight:900;color:{rc};line-height:1">
                    {risk_num}
                  </div>
                  <div style="font-size:.78rem;font-weight:700;color:{rc};margin-top:4px">
                    {rl}
                  </div>
                  <div style="font-size:.62rem;color:#64748b;margin-top:8px;
                  background:#fff;border-radius:6px;padding:5px 8px;display:inline-block">
                    {patho_num} (patho) × {resist_num} (résist.) × {dissem_num} (dissém.) = {risk_num}
                  </div>
                  <div style="font-size:.65rem;font-weight:600;padding:5px 8px;border-radius:6px;
                  margin-top:8px;background:{status_bg};color:{status_c}">
                    {status_txt}
                  </div>
                </div>""", unsafe_allow_html=True)

            with c3:
                st.markdown(
                    "<div style='font-size:.72rem;font-weight:800;color:#1e40af;"
                    "letter-spacing:.06em;text-transform:uppercase;margin-bottom:6px'>"
                    "Information supplémentaire</div>", unsafe_allow_html=True)

                new_notes = st.text_area(
                    "📝 Notes",
                    value=existing.get("notes", "") or "" if is_edit else "",
                    height=70, disabled=not can_edit)

                new_comment = st.text_area(
                    "💬 Commentaire détaillé",
                    value=existing.get("comment", "") or "" if is_edit else "",
                    height=70, disabled=not can_edit)

            st.markdown("</div>", unsafe_allow_html=True)
            cb1, cb2 = st.columns(2)
            with cb1:
                if can_edit:
                    if st.button(
                        "✅ " + ("Modifier" if is_edit else "Ajouter"),
                        use_container_width=True, key="form_submit"):
                        if not new_name.strip():
                            st.error("Le nom est obligatoire.")
                            return
                        origine_node = (
                            ("Humains" if new_famille == "Bactéries" else "Humain")
                            if "Humain" in new_origine else "Environnemental")
                        new_germ = dict(
                            name=new_name.strip(),
                            path=["Germes", new_famille, origine_node, new_cat],
                            pathogenicity=patho_num,
                            resistance=resist_num,
                            dissemination=dissem_num,
                            risk=risk_num,
                            notes=new_notes.strip() or None,
                            comment=new_comment.strip() or None,
                        )
                        if is_edit:
                            st.session_state.germs[idx] = new_germ
                            st.session_state.edit_idx   = None
                        else:
                            if any(g["name"].lower() == new_name.strip().lower()
                                   for g in st.session_state.germs):
                                st.error("Ce germe existe déjà.")
                                return
                            st.session_state.germs.append(new_germ)
                            st.session_state.show_add = False
                        save_germs(st.session_state.germs)
                        st.rerun()
            with cb2:
                if can_edit:
                    if st.button("Annuler", use_container_width=True, key="form_cancel"):
                        st.session_state.show_add = False
                        st.session_state.edit_idx = None
                        st.rerun()

    if can_edit and st.session_state.show_add and st.session_state.edit_idx is None:
        germ_form()
    if can_edit and st.session_state.edit_idx is not None:
        germ_form(
            existing=st.session_state.germs[st.session_state.edit_idx],
            idx=st.session_state.edit_idx)

    st.divider()

    germs_json         = json.dumps(st.session_state.germs, ensure_ascii=False)
    default_names_json = json.dumps(sorted(DEFAULT_GERM_NAMES), ensure_ascii=False)

    tree_html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{background:#f8fafc;color:#1e293b;font-family:'Segoe UI',sans-serif;height:95vh;overflow:hidden}}
.app{{display:flex;height:95vh}}
.tree-wrap{{flex:1;overflow:auto;padding:8px;scrollbar-width:thin;scrollbar-color:#cbd5e1 transparent}}
svg{{min-width:900px;width:100%;height:100%}}
.node rect{{fill:#fff;stroke:#e2e8f0;stroke-width:1.5;transition:all .2s;cursor:pointer}}
.node.highlighted rect{{stroke-width:2.5;filter:drop-shadow(0 0 4px rgba(0,0,0,.15))}}
.node text{{font-size:11px;fill:#0f172a;pointer-events:none;font-family:'Courier New',monospace}}
.link{{fill:none;stroke:#e2e8f0;stroke-width:1.5;transition:all .3s}}
.link.highlighted{{stroke-width:2.5}}
.right-panel{{width:490px;border-left:2px solid #e2e8f0;display:flex;flex-direction:column;background:#f1f5f9;flex-shrink:0}}
.sbox{{padding:12px 14px;border-bottom:1px solid #e2e8f0}}
.sbox input{{width:100%;background:#fff;border:1.5px solid #e2e8f0;border-radius:10px;padding:10px 14px;color:#1e293b;font-size:.9rem;outline:none}}
.sbox input:focus{{border-color:#2563eb;box-shadow:0 0 0 3px rgba(37,99,235,.1)}}
.germ-list{{flex:1;overflow-y:auto;padding:6px;scrollbar-width:thin}}
.germ-item{{display:flex;align-items:center;gap:10px;padding:9px 12px;border-radius:8px;cursor:pointer;transition:background .15s;font-size:.85rem;color:#0f172a;border:1px solid transparent;margin-bottom:3px}}
.germ-item:hover{{background:#fff;box-shadow:0 1px 4px rgba(0,0,0,.08)}}
.germ-item.active{{background:#fff;border-color:#2563eb;box-shadow:0 1px 6px rgba(37,99,235,.2)}}
.risk-dot{{width:10px;height:10px;border-radius:50%;flex-shrink:0}}
.germ-count{{font-size:.72rem;color:#64748b;padding:6px 14px;text-align:center;border-bottom:1px solid #e2e8f0;background:#f8fafc;font-weight:600}}
.group-header{{font-size:.65rem;font-weight:800;letter-spacing:.12em;text-transform:uppercase;color:#475569;padding:8px 12px 3px;background:#e2e8f0;margin:4px -6px 3px;border-radius:3px}}
.info-panel{{border-top:2px solid #e2e8f0;padding:14px;background:#fff;display:none;max-height:600px;overflow-y:auto}}
.info-panel.visible{{display:block}}
.info-name{{font-size:.92rem;font-weight:700;font-style:italic;color:#1e293b;margin-bottom:3px}}
.info-path{{font-size:.6rem;color:#2563eb;opacity:.85;margin-bottom:8px;font-family:monospace}}
.info-badge{{display:inline-flex;align-items:center;gap:5px;font-size:.7rem;padding:3px 10px;border-radius:20px;border:1px solid;margin-bottom:10px;font-weight:600}}
.info-lbl{{font-size:.58rem;letter-spacing:.1em;text-transform:uppercase;color:#64748b;margin-bottom:2px;margin-top:8px;font-weight:700}}
.info-val{{font-size:.78rem;color:#1e293b;line-height:1.4}}
.score-grid{{display:grid;grid-template-columns:1fr 1fr 1fr;gap:5px;margin-top:6px}}
.score-cell{{background:#f8fafc;border:1px solid #e2e8f0;border-radius:7px;padding:7px 5px;text-align:center}}
.score-cell .lbl{{font-size:.52rem;color:#64748b;text-transform:uppercase;margin-bottom:3px;line-height:1.3}}
.score-cell .val{{font-size:1.2rem;font-weight:900}}
.score-total{{background:#1e293b;border-radius:8px;padding:8px;text-align:center;margin-top:6px}}
.alert-row{{padding:5px 10px;border-radius:6px;font-size:.68rem;font-weight:700;margin-top:5px}}
.notes-box{{margin-top:6px;padding:7px 10px;border-radius:6px;background:rgba(37,99,235,.04);border:1px solid rgba(37,99,235,.18);font-size:.72rem;color:#0f172a;line-height:1.5}}
</style></head><body>
<div class="app">
  <div class="tree-wrap"><svg id="svg"></svg></div>
  <div class="right-panel">
    <div class="sbox"><input type="text" id="sbox" placeholder="🔍 Rechercher un germe..." oninput="filterList()"></div>
    <div class="germ-count" id="germCount">Chargement...</div>
    <div class="germ-list" id="germList"></div>
    <div class="info-panel" id="infoPanel"></div>
  </div>
</div>
<script>
const GERMS={germs_json};
const DEFAULT_NAMES=new Set({default_names_json});
const NODE_W=190,NODE_H=28,H_GAP=28,V_GAP=10;
const LEVEL_COLS=["#38bdf8","#818cf8","#fb923c","#34d399","#a3e635"];
const PATHO_LBL=["","Non pathogène","Pathogène opportuniste","Pathogène opp. MR / Primaire"];
const RESIST_LBL=["","Sensible","Résistant Surfa'Safe","Résistant Surfa'Safe + APA"];
const DISSEM_LBL=["","Environnemental","Manuporté","Aéroporté"];

function riskColor(s){{
  if(s<=4)return"#22c55e";if(s<=8)return"#84cc16";
  if(s<=12)return"#f59e0b";if(s<=18)return"#f97316";return"#ef4444";
}}
function riskLabel(s){{
  if(s<=4)return"Faible";if(s<=8)return"Modéré";
  if(s<=12)return"Important";if(s<=18)return"Majeur";return"Critique";
}}
function germScore(g){{
  if(g.pathogenicity&&g.resistance&&g.dissemination)
    return g.pathogenicity*g.resistance*g.dissemination;
  const m={{1:1,2:2,3:6,4:12,5:18}};return m[g.risk]||g.risk||1;
}}

function buildTree(){{
  const root={{name:"Germes",children:[]}};
  GERMS.forEach(g=>{{
    let cur=root;
    g.path.slice(1).forEach(n=>{{
      let c=cur.children&&cur.children.find(x=>x.name===n);
      if(!c){{c={{name:n,children:[]}};if(!cur.children)cur.children=[];cur.children.push(c);}}
      cur=c;
    }});
  }});
  function clean(n){{if(n.children&&!n.children.length)delete n.children;else if(n.children)n.children.forEach(clean);}}
  clean(root);return root;
}}
function computeLayout(node,depth=0,y=0){{
  node.depth=depth;node.x=depth*(NODE_W+H_GAP);
  if(!node.children||!node.children.length){{node.y=y;return y+NODE_H;}}
  let cy=y;node.children.forEach(c=>{{cy=computeLayout(c,depth+1,cy);cy+=V_GAP;}});
  cy-=V_GAP;node.y=(y+cy)/2;return cy+V_GAP;
}}
function allNodes(n){{return[n,...(n.children||[]).flatMap(allNodes)];}}
function allLinks(n){{return(n.children||[]).flatMap(c=>[{{source:n,target:c}},...allLinks(c)]);}}
function buildPaths(n,p=[]){{n.fullPath=[...p,n.name];(n.children||[]).forEach(c=>buildPaths(c,n.fullPath));}}

function renderTree(){{
  const tree=buildTree();computeLayout(tree);buildPaths(tree);
  const nodes=allNodes(tree),links=allLinks(tree);
  const maxY=Math.max(...nodes.map(n=>n.y))+NODE_H+20;
  const maxX=Math.max(...nodes.map(n=>n.x))+NODE_W+20;
  const svg=document.getElementById('svg');
  svg.innerHTML='';
  svg.setAttribute('viewBox',`0 0 ${{maxX}} ${{maxY}}`);
  svg.setAttribute('height',maxY);svg.setAttribute('width',maxX);
  links.forEach(l=>{{
    const p=document.createElementNS('http://www.w3.org/2000/svg','path');
    const x1=l.source.x+NODE_W,y1=l.source.y+NODE_H/2,x2=l.target.x,y2=l.target.y+NODE_H/2,mx=(x1+x2)/2;
    p.setAttribute('d',`M${{x1}},${{y1}} C${{mx}},${{y1}} ${{mx}},${{y2}} ${{x2}},${{y2}}`);
    p.setAttribute('class','link');
    p.dataset.sourcefull=l.source.fullPath.join('|||');
    p.dataset.targetfull=l.target.fullPath.join('|||');
    svg.appendChild(p);
  }});
  nodes.forEach(node=>{{
    const g=document.createElementNS('http://www.w3.org/2000/svg','g');
    g.setAttribute('class','node');
    g.setAttribute('transform',`translate(${{node.x}},${{node.y}})`);
    g.dataset.fullpath=node.fullPath.join('|||');
    const col=LEVEL_COLS[node.depth]||"#0f172a";
    const rect=document.createElementNS('http://www.w3.org/2000/svg','rect');
    rect.setAttribute('width',NODE_W);rect.setAttribute('height',NODE_H);
    rect.setAttribute('rx',5);rect.setAttribute('stroke',col);
    const text=document.createElementNS('http://www.w3.org/2000/svg','text');
    text.setAttribute('x',NODE_W/2);text.setAttribute('y',NODE_H/2+4);
    text.setAttribute('text-anchor','middle');
    text.textContent=node.name.length>27?node.name.substring(0,25)+'...':node.name;
    g.appendChild(rect);g.appendChild(text);
    g.addEventListener('mouseenter',()=>highlightPath(node.fullPath));
    g.addEventListener('mouseleave',clearHighlight);
    svg.appendChild(g);
  }});
}}
let selectedPath=null;
function highlightPath(exactPath){{
  const vp=new Set();for(let i=1;i<=exactPath.length;i++)vp.add(exactPath.slice(0,i).join('|||'));
  document.querySelectorAll('.node').forEach(n=>n.classList.toggle('highlighted',vp.has(n.dataset.fullpath||'')));
  document.querySelectorAll('.link').forEach(l=>{{
    const on=vp.has(l.dataset.sourcefull||'')&&vp.has(l.dataset.targetfull||'');
    l.classList.toggle('highlighted',on);
    if(on){{const d=(l.dataset.sourcefull||'').split('|||').length-1;l.style.stroke=LEVEL_COLS[d]||'#38bdf8';}}
    else l.style.stroke='';
  }});
}}
function clearHighlight(){{
  if(selectedPath){{highlightPath(selectedPath);return;}}
  document.querySelectorAll('.node').forEach(n=>n.classList.remove('highlighted'));
  document.querySelectorAll('.link').forEach(l=>{{l.classList.remove('highlighted');l.style.stroke=''}});
}}
function renderList(filter=''){{
  const list=document.getElementById('germList');
  const countEl=document.getElementById('germCount');
  list.innerHTML='';
  const filtered=GERMS.filter(g=>g.name.toLowerCase().includes(filter.toLowerCase()));
  countEl.textContent=filter?`${{filtered.length}} / ${{GERMS.length}} germe(s)`:`${{GERMS.length}} germe(s) — cliquez pour détails`;
  const groups={{}};
  filtered.forEach(g=>{{const cat=(g.path&&g.path[3])||'Autres';if(!groups[cat])groups[cat]=[];groups[cat].push(g);}});
  Object.keys(groups).sort().forEach(cat=>{{
    const header=document.createElement('div');header.className='group-header';header.textContent=cat;list.appendChild(header);
    groups[cat].forEach(g=>{{
      const score=germScore(g);const col=riskColor(score);
      const isCustom=!DEFAULT_NAMES.has(g.name);
      const isAir=(g.dissemination||0)===3;
      const hasResist=(g.resistance||0)>=2;
      let badges='';
      if(isAir)badges+='<span style="font-size:.48rem;background:#eff6ff;color:#1e40af;border-radius:3px;padding:0 4px;margin-left:2px">✈</span>';
      if(hasResist)badges+='<span style="font-size:.48rem;background:#fee2e2;color:#991b1b;border-radius:3px;padding:0 4px;margin-left:2px">⚠</span>';
      if(isCustom)badges+='<span style="font-size:.48rem;background:rgba(56,189,248,.15);color:#2563eb;border-radius:3px;padding:0 4px;margin-left:2px">★</span>';
      const div=document.createElement('div');div.className='germ-item';div.dataset.name=g.name;
      div.innerHTML=`<span class="risk-dot" style="background:${{col}}"></span><span style="flex:1;line-height:1.3">${{g.name}}${{badges}}</span><span style="font-size:.62rem;color:${{col}};font-weight:800;flex-shrink:0">${{score}}</span>`;
      div.addEventListener('click',()=>selectGerm(g));
      list.appendChild(div);
    }});
  }});
}}
function filterList(){{renderList(document.getElementById('sbox').value);}}
function selectGerm(g){{
  selectedPath=g.path;highlightPath(g.path);
  document.querySelectorAll('.germ-item').forEach(el=>el.classList.toggle('active',el.dataset.name===g.name));
  showInfo(g);
}}
function showInfo(g){{
  const panel=document.getElementById('infoPanel');panel.classList.add('visible');
  const score=germScore(g);const col=riskColor(score);
  const alertRow=score>24?`<div class="alert-row" style="background:#fef2f2;color:#991b1b">🚨 Action (score germe seul &gt; 24)</div>`
    :score>=16?`<div class="alert-row" style="background:#fffbeb;color:#92400e">⚠️ Alerte probable selon criticité du lieu</div>`
    :`<div class="alert-row" style="background:#f0fdf4;color:#166534">✅ Conforme à score germe seul</div>`;
  const nh=g.notes?`<div class="info-lbl">📝 Notes</div><div class="notes-box">${{g.notes}}</div>`:'';
  const ch=g.comment?`<div class="info-lbl" style="color:#fb923c">💬 Commentaire</div><div class="notes-box" style="color:#ea580c;background:rgba(251,146,60,.06);border-color:rgba(251,146,60,.3)">${{g.comment}}</div>`:'';
  panel.innerHTML=`
    <div class="info-name">${{g.name}}</div>
    <div class="info-path">${{g.path.join(' › ')}}</div>
    <div class="info-badge" style="color:${{col}};background:${{col}}22;border-color:${{col}}55">
      <span style="width:8px;height:8px;border-radius:50%;background:${{col}};display:inline-block"></span>
      Score ${{score}} — ${{riskLabel(score)}}
    </div>
    <div class="info-lbl">Critères de criticité</div>
    <div class="score-grid">
      <div class="score-cell">
        <div class="lbl">🧬 Pathogénicité</div>
        <div class="val" style="color:${{riskColor(g.pathogenicity||1)}}">${{g.pathogenicity||1}}</div>
        <div style="font-size:.5rem;color:#64748b;margin-top:2px">${{PATHO_LBL[g.pathogenicity||1]||'—'}}</div>
      </div>
      <div class="score-cell">
        <div class="lbl">🧴 Résistance</div>
        <div class="val" style="color:${{riskColor(g.resistance||1)}}">${{g.resistance||1}}</div>
        <div style="font-size:.5rem;color:#64748b;margin-top:2px">${{RESIST_LBL[g.resistance||1]||'—'}}</div>
      </div>
      <div class="score-cell">
        <div class="lbl">💨 Dissémination</div>
        <div class="val" style="color:${{riskColor(g.dissemination||1)}}">${{g.dissemination||1}}</div>
        <div style="font-size:.5rem;color:#64748b;margin-top:2px">${{DISSEM_LBL[g.dissemination||1]||'—'}}</div>
      </div>
    </div>
    <div class="score-total">
      <div style="font-size:.58rem;color:#94a3b8;text-transform:uppercase;letter-spacing:.1em">Score calculé</div>
      <div style="font-size:1.8rem;font-weight:900;color:${{col}};line-height:1.1">${{score}}</div>
      <div style="font-size:.6rem;color:#94a3b8">${{g.pathogenicity||'?'}} × ${{g.resistance||'?'}} × ${{g.dissemination||'?'}}</div>
    </div>
    ${{alertRow}}
    ${{nh}}${{ch}}`;
}}
renderTree();
renderList();
</script></body></html>"""

    st.components.v1.html(tree_html, height=1200, scrolling=False)

    st.markdown("### ✏️ Gérer les germes")
    search_edit = st.text_input(
        "Filtrer", placeholder="Rechercher un germe...", label_visibility="collapsed")
    filtered = ([g for g in st.session_state.germs
                 if search_edit.lower() in g["name"].lower()]
                if search_edit else st.session_state.germs)

    for i, g in enumerate(filtered):
        real_idx = st.session_state.germs.index(g)
        score    = _germ_score(g)
        c        = _risk_color(score)
        rl       = _risk_label(score)

        col_n, col_s, col_rl, col_e, col_d = st.columns([4, 0.6, 1, 0.6, 0.6])
        with col_n:
            st.markdown(
                f'<span style="color:{c};font-size:.75rem">●</span> '
                f'<span style="font-size:.82rem;font-style:italic">{g["name"]}</span>',
                unsafe_allow_html=True)
        with col_s:
            st.markdown(
                f'<span style="font-size:.75rem;color:{c};font-weight:800">{score}</span>',
                unsafe_allow_html=True)
        with col_rl:
            st.markdown(
                f'<span style="font-size:.65rem;color:{c};font-weight:700">{rl}</span>',
                unsafe_allow_html=True)
        with col_e:
            if can_edit:
                if st.button("✏️", key=f"edit_{real_idx}_{i}"):
                    st.session_state.edit_idx = real_idx
                    st.session_state.show_add = False
                    st.rerun()
        with col_d:
            if can_edit:
                if st.button("🗑️", key=f"del_{real_idx}_{i}"):
                    st.session_state.germs.pop(real_idx)
                    save_germs(st.session_state.germs)
                    st.rerun()

# ═══════════════════════════════════════════════════════════════════════════════
# TAB : SURVEILLANCE
# ═══════════════════════════════════════════════════════════════════════════════


# ── Helper criticité lieu ──────────────────────────────────────────────────────
def _get_location_criticality(smp: dict) -> int:
    label = smp.get("label", "")
    if label:
        pt = next((p for p in st.session_state.points if p.get("label") == label), None)
        if pt:
            try:
                return int(pt.get("location_criticality", 1))
            except (ValueError, TypeError):
                pass
    raw = smp.get("location_criticality") or smp.get("criticality")
    if raw is not None:
        try:
            return int(raw)
        except (ValueError, TypeError):
            pass
    return 1


def _loc_crit_label(lc: int) -> str:
    return {1: "Limité", 2: "Modéré", 3: "Important", 4: "Critique"}.get(lc, "?")


# ═══════════════════════════════════════════════════════════════════════════════
# FONCTION UNIFIÉE — _render_mesures_correctives
#
# Appelée depuis :
#   • la liste surveillance  → _render_mesures_correctives(entry, entry_idx)
#   • le popup post-ident    → _render_mesures_correctives(pop_data,
#                                                           entry_idx=None,
#                                                           popup_mode=True)
# ═══════════════════════════════════════════════════════════════════════════════
def _render_mesures_correctives(
    entry: dict,
    entry_idx: int | None = None,
    popup_mode: bool = False,
):
    from datetime import datetime as _dt

    STATUS    = entry.get("status", "ok")
    MC_STATUT = entry.get("mc_statut", "")
    MC_DATE   = entry.get("mc_date",   "")

    if STATUS not in ("alert", "action"):
        return

    # ── Clé de suffixe unique pour les widgets Streamlit ─────────────────────
    key_suffix = (
        str(entry_idx)
        if entry_idx is not None
        else str(entry.get("sample_id", entry.get("label", "pop")))
    )

    # ── Palette selon l'état des mesures ─────────────────────────────────────
    if MC_STATUT == "fait":
        _brd      = "#86efac"; _bg = "#f0fdf4"; _title_col = "#166534"
        _badge_bg = "#22c55e"; _badge_txt = "MESURES CORRECTIVES FAITES ✅"
    else:
        _brd      = "#fca5a5" if STATUS == "action" else "#fcd34d"
        _bg       = "#fef2f2" if STATUS == "action" else "#fffbeb"
        _title_col = "#991b1b" if STATUS == "action" else "#92400e"
        _badge_bg  = "#ef4444" if STATUS == "action" else "#f59e0b"
        _badge_txt = "MESURES CORRECTIVES EN ATTENTE"

    st.markdown(
        f"<div style='background:{_bg};border:1.5px solid {_brd};"
        f"border-radius:12px;padding:14px 18px;margin-top:10px'>"
        f"<div style='font-size:.82rem;font-weight:800;color:{_title_col};"
        f"margin-bottom:10px;display:flex;justify-content:space-between;align-items:center'>"
        f"<span>🔧 Mesures correctives</span>"
        f"<span style='background:{_badge_bg};color:#fff;border-radius:6px;"
        f"padding:2px 10px;font-size:.68rem;font-weight:700'>{_badge_txt}</span>"
        f"</div>",
        unsafe_allow_html=True,
    )

    # ── Helper : trouver l'index réel dans st.session_state.surveillance ──────
    def _find_surv_index() -> int | None:
        if entry_idx is not None and 0 <= entry_idx < len(st.session_state.surveillance):
            return entry_idx
        _sid  = entry.get("sample_id") or entry.get("label") or entry.get("prelevement")
        _date = entry.get("date", "")
        _germ = (entry.get("germ_saisi", "")
                 or entry.get("germ_match", "")
                 or entry.get("germ", ""))
        _ufc  = entry.get("ufc")
        for _i, _sr in enumerate(st.session_state.surveillance):
            if (
                (_sr.get("prelevement") == _sid
                 or _sr.get("sample_id") == _sid
                 or _sr.get("label")     == _sid)
                and ((not _date) or _sr.get("date", "") == _date)
                and ((not _germ) or _sr.get("germ_saisi", "") == _germ
                     or _sr.get("germ_match", "") == _germ)
                and (_ufc is None or _sr.get("ufc") == _ufc)
                and _sr.get("status") in ("alert", "action")
            ):
                return _i
        return None

    def _persist():
        save_surveillance(st.session_state.surveillance)
        _supa_upsert(
            "surveillance",
            json.dumps(st.session_state.surveillance, ensure_ascii=False),
        )

    def _valider_mc():
        _i = _find_surv_index()
        if _i is not None:
            st.session_state.surveillance[_i]["mc_statut"] = "fait"
            st.session_state.surveillance[_i]["mc_detail"] = ""
            st.session_state.surveillance[_i]["mc_date"]   = _dt.now().strftime("%d/%m/%Y %H:%M")
        _persist()

    def _annuler_mc():
        _i = _find_surv_index()
        if _i is not None:
            st.session_state.surveillance[_i]["mc_statut"] = ""
            st.session_state.surveillance[_i]["mc_detail"] = ""
            st.session_state.surveillance[_i]["mc_date"]   = ""
        _persist()

    # ── Génération PDF ────────────────────────────────────────────────────────
    def _gen_pdf_mesures(data: dict, mesures: list) -> bytes:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units     import cm as rl_cm
        from reportlab.lib           import colors as rlc
        from reportlab.platypus      import (
            BaseDocTemplate, Frame, PageTemplate,
            Paragraph, Spacer, HRFlowable, Table, TableStyle,
        )
        from reportlab.lib.styles import ParagraphStyle
        from io import BytesIO

        buf    = BytesIO()
        A4_W, A4_H = A4
        MARGIN = 1.8 * rl_cm

        s_title  = ParagraphStyle("mc_t",  fontName="Helvetica-Bold", fontSize=14,
                                  leading=18, textColor=rlc.HexColor("#1e40af"), spaceAfter=6)
        s_sub    = ParagraphStyle("mc_s",  fontName="Helvetica",      fontSize=9,
                                  leading=12, textColor=rlc.HexColor("#64748b"), spaceAfter=10)
        s_val    = ParagraphStyle("mc_v",  fontName="Helvetica",      fontSize=9,
                                  leading=12, textColor=rlc.HexColor("#475569"), spaceAfter=8)
        s_mhead  = ParagraphStyle("mc_mh", fontName="Helvetica-Bold", fontSize=10,
                                  leading=13, textColor=rlc.HexColor("#991b1b"),
                                  spaceBefore=12, spaceAfter=6)
        s_item   = ParagraphStyle("mc_i",  fontName="Helvetica",      fontSize=9,
                                  leading=13, textColor=rlc.HexColor("#0f172a"),
                                  leftIndent=10, spaceAfter=5)
        s_footer = ParagraphStyle("mc_f",  fontName="Helvetica",      fontSize=7,
                                  leading=9,  textColor=rlc.HexColor("#94a3b8"))

        doc   = BaseDocTemplate(buf, pagesize=A4,
                                leftMargin=MARGIN, rightMargin=MARGIN,
                                topMargin=MARGIN,  bottomMargin=MARGIN)
        frame = Frame(MARGIN, MARGIN, A4_W - 2*MARGIN, A4_H - 2*MARGIN,
                      leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0)
        doc.addPageTemplates([PageTemplate(id="main", frames=[frame])])

        _is_action  = data.get("status") == "action"
        _status_txt = "🚨 ACTION REQUISE" if _is_action else "⚠️ ALERTE"
        _status_col = rlc.HexColor("#dc2626") if _is_action else rlc.HexColor("#d97706")
        _now_str    = _dt.now().strftime("%d/%m/%Y %H:%M")

        story = [
            Paragraph("FICHE MESURES CORRECTIVES", s_title),
            Paragraph(f"MicroSurveillance URC — Généré le {_now_str}", s_sub),
            HRFlowable(width="100%", thickness=1.5, color=_status_col, spaceAfter=10),
        ]

        tbl_data = [
            ["Statut",   _status_txt],
            ["Point",    data.get("label",   "—")],
            ["Germe",    data.get("germ",    data.get("germ_match", "—"))],
            ["UFC",      str(data.get("ufc", "—"))],
            ["Score",    str(data.get("total_score", "—"))],
            ["Lieu Nv.", str(data.get("loc_criticality",
                                      data.get("location_criticality", "—")))],
        ]
        tbl = Table(tbl_data, colWidths=[4*rl_cm, A4_W - 2*MARGIN - 4*rl_cm])
        tbl.setStyle(TableStyle([
            ("FONTNAME",       (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME",       (1, 0), (1, -1), "Helvetica"),
            ("FONTSIZE",       (0, 0), (-1, -1), 9),
            ("TEXTCOLOR",      (0, 0), (0, -1), rlc.HexColor("#64748b")),
            ("TEXTCOLOR",      (1, 0), (1,  0), _status_col),
            ("FONTNAME",       (1, 0), (1,  0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 0), (-1, -1), [rlc.HexColor("#f8fafc"), rlc.white]),
            ("LEFTPADDING",    (0, 0), (-1, -1), 8),
            ("RIGHTPADDING",   (0, 0), (-1, -1), 8),
            ("TOPPADDING",     (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING",  (0, 0), (-1, -1), 5),
            ("GRID",           (0, 0), (-1, -1), 0.5, rlc.HexColor("#e2e8f0")),
        ]))
        story += [tbl, Spacer(1, 14)]

        story.append(Paragraph("📋 Mesures correctives applicables", s_mhead))
        if mesures:
            for idx_m, m in enumerate(mesures, 1):
                story.append(Paragraph(f"{idx_m}. {m['text']}", s_item))
        else:
            story.append(Paragraph("Aucune mesure corrective configurée.", s_val))

        story += [
            Spacer(1, 20),
            HRFlowable(width="100%", thickness=0.5,
                       color=rlc.HexColor("#cbd5e1"), spaceAfter=8),
            Paragraph("Préleveur / Responsable : ________________________________", s_val),
            Spacer(1, 6),
            Paragraph("Date de traitement :    ________________________________",   s_val),
            Spacer(1, 6),
            Paragraph("Signature :              ________________________________",  s_val),
            Spacer(1, 20),
            Paragraph("URC — MicroSurveillance · Document généré automatiquement", s_footer),
        ]
        doc.build(story)
        buf.seek(0)
        return buf.getvalue()

    # ── Rendu selon mc_statut ─────────────────────────────────────────────────
    if MC_STATUT == "fait":
        st.markdown(
            f"<div style='font-size:.78rem;color:#166534;margin-bottom:6px'>"
            f"✅ Validé le <b>{MC_DATE}</b></div>",
            unsafe_allow_html=True,
        )
        if st.button("↩️ Annuler la validation",
                     key=f"mc_annuler_{key_suffix}",
                     use_container_width=False):
            _annuler_mc()
            st.rerun()

    else:
        mesures = getattr(st.session_state, "mesures_correctives", [])

        # Résumé compact en mode popup
        if popup_mode:
            _germ  = entry.get("germ", entry.get("germ_match", "—"))
            _ufc_v = entry.get("ufc", "—")
            _score = entry.get("total_score", "—")
            _label = entry.get("label", "—")
            _trig  = entry.get("triggered_by", "")
            st.markdown(
                f"<div style='font-size:.75rem;color:{_title_col};"
                f"margin-bottom:10px;line-height:1.6'>"
                f"<b>Point :</b> {_label} &nbsp;·&nbsp; "
                f"<b>Germe :</b> {_germ} &nbsp;·&nbsp; "
                f"<b>UFC :</b> {_ufc_v} &nbsp;·&nbsp; "
                f"<b>Score :</b> {_score}"
                + (f"<br><span style='font-size:.7rem;color:#64748b'>{_trig}</span>"
                   if _trig else "")
                + "</div>",
                unsafe_allow_html=True,
            )

        _pdf_key = f"_pdf_mc_{key_suffix}"
        cols = st.columns([3, 2, 1]) if popup_mode else st.columns([3, 2])
        _b1, _b2 = cols[0], cols[1]
        _b3 = cols[2] if popup_mode else None

        with _b1:
            btn_label = (
                "✅ Compris — Mesures prises en charge"
                if popup_mode
                else "✅ Prise en compte des mesures correctives"
            )
            if st.button(btn_label, key=f"mc_valider_{key_suffix}",
                         type="primary", use_container_width=True):
                _valider_mc()
                if popup_mode:
                    st.session_state["_last_mesures_popup"] = entry
                    st.session_state["_show_mesures_popup"] = None
                else:
                    st.session_state["mc_success"] = True
                st.rerun()

        with _b2:
            if st.button("🖨️ Imprimer (PDF)", use_container_width=True,
                         key=f"mc_print_{key_suffix}"):
                try:
                    st.session_state[_pdf_key] = _gen_pdf_mesures(entry, mesures)
                except ImportError:
                    st.error("❌ ReportLab non installé.")
                except Exception as _pe:
                    st.error(f"Erreur PDF : {_pe}")
            if st.session_state.get(_pdf_key):
                _fname = (
                    entry.get("label", entry.get("prelevement", "mesures"))[:20]
                    .replace(" ", "_")
                )
                st.download_button(
                    "⬇️ Télécharger la fiche PDF",
                    data=st.session_state[_pdf_key],
                    file_name=f"mesures_correctives_{_fname}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                    key=f"mc_dl_{key_suffix}",
                )

        if popup_mode and _b3:
            with _b3:
                if st.button("✕ Ignorer", use_container_width=True,
                             key=f"mc_dismiss_{key_suffix}"):
                    st.session_state["_show_mesures_popup"] = None
                    st.rerun()

    st.markdown("</div>", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# TAB SURVEILLANCE — LOGIQUE PRINCIPALE
# ═══════════════════════════════════════════════════════════════════════════════
if active == "surveillance":
    st.markdown("### 🔍 Identification & Surveillance microbiologique")

    # ── Fix global des dates J2 / J7 ─────────────────────────────────────────
    updated = False
    for s in st.session_state.schedules:
        smp = next((p for p in st.session_state.prelevements
                    if p["id"] == s["sample_id"]), None)
        if not smp or not smp.get("date"):
            continue
        p_date = date.fromisoformat(smp["date"])
        if s["when"] == "J2":
            new_due = next_working_day_offset(p_date, 2).isoformat()
            if s.get("due_date") != new_due:
                s["due_date"] = new_due; updated = True
        elif s["when"] == "J7":
            new_due = next_working_day_offset(p_date, 7).isoformat()
            if s.get("due_date") != new_due:
                s["due_date"] = new_due; updated = True
    if updated:
        save_schedules(st.session_state.schedules)

    # ── Helper : jours depuis le prélèvement ─────────────────────────────────
    def _days_since_sample(s):
        smp = next((p for p in st.session_state.prelevements
                    if p["id"] == s["sample_id"]), None)
        if not smp or not smp.get("date"):
            return 999
        return (today - date.fromisoformat(smp["date"])).days

    _active_sids_surv = {p["id"] for p in st.session_state.prelevements
                         if not p.get("archived")}

    _j2_red    = sum(1 for s in st.session_state.schedules
                     if s["when"] == "J2" and s["status"] == "pending"
                     and s.get("sample_id") in _active_sids_surv
                     and datetime.fromisoformat(s["due_date"]).date() <= today)
    _j2_orange = sum(1 for s in st.session_state.schedules
                     if s["when"] == "J2" and s["status"] == "pending"
                     and s.get("sample_id") in _active_sids_surv
                     and datetime.fromisoformat(s["due_date"]).date() > today)
    _j7_red    = sum(1 for s in st.session_state.schedules
                     if s["when"] == "J7" and s["status"] == "pending"
                     and s.get("sample_id") in _active_sids_surv
                     and datetime.fromisoformat(s["due_date"]).date() <= today
                     and _days_since_sample(s) >= 7)
    _j7_orange = sum(1 for s in st.session_state.schedules
                     if s["when"] == "J7" and s["status"] == "pending"
                     and s.get("sample_id") in _active_sids_surv
                     and (datetime.fromisoformat(s["due_date"]).date() > today
                          or _days_since_sample(s) < 7))
    _id_red    = sum(1 for p in st.session_state.pending_identifications
                     if p.get("status") == "pending")

    _dot_j2 = " 🔴" if _j2_red > 0 else (" 🟠" if _j2_orange > 0 else "")
    _dot_j7 = " 🔴" if _j7_red > 0 else (" 🟠" if _j7_orange > 0 else "")
    _dot_id = " 🔴" if _id_red > 0 else ""

    tab_nouveau, tab_j2, tab_j7, tab_ident = st.tabs([
        "🧪 Nouveau prélèvement",
        f"📖 Lecture J2{_dot_j2}",
        f"📗 Lecture J7{_dot_j7}",
        f"🔴 Identifications en attente{_dot_id}",
    ])

    # ──────────────────────────────────────────────────────────────────────────
    # HELPERS GLOBAUX
    # ──────────────────────────────────────────────────────────────────────────
    def _fix_qr_input(text: str) -> str:
        def _fix_encoding(s):
            try:
                return s.encode("latin-1").decode("utf-8")
            except (UnicodeDecodeError, UnicodeEncodeError):
                return s

        def _fix_azerty(s):
            AZERTY_MAP = {
                "q": "a", "z": "w", "a": "q", "w": "z",
                "Q": "A", "Z": "W", "A": "Q", "W": "Z",
                "M": ":", ";": "m",
                "&": "1", "é": "2", '"': "3", "'": "4",
                "(": "5", "-": "6", "è": "7", "_": "8",
                "ç": "9", "à": "0",
            }
            return "".join(AZERTY_MAP.get(c, c) for c in s)

        text = _fix_encoding(text)
        if not text.startswith("{"):
            text = _fix_azerty(text)
        return text

    if "prelev_mode" not in st.session_state:
        st.session_state["prelev_mode"] = "manuel"
    if "qr_counter" not in st.session_state:
        st.session_state["qr_counter"] = 0

    # ══════════════════════════════════════════════════════════════════════════
    # ONGLET 1 — NOUVEAU PRÉLÈVEMENT
    # ══════════════════════════════════════════════════════════════════════════
    with tab_nouveau:
        tog1, tog2 = st.columns(2)
        with tog1:
            if st.button("✏️  Saisie manuelle", use_container_width=True,
                         type="primary" if st.session_state["prelev_mode"] == "manuel" else "secondary",
                         key="tog_manuel"):
                st.session_state["prelev_mode"] = "manuel"
                st.rerun()
        with tog2:
            if st.button("📷  Scan QR code", use_container_width=True,
                         type="primary" if st.session_state["prelev_mode"] == "scan" else "secondary",
                         key="tog_scan"):
                st.session_state["prelev_mode"] = "scan"
                st.rerun()

        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

        # ── MODE MANUEL ───────────────────────────────────────────────────────
        if st.session_state["prelev_mode"] == "manuel":
            if not st.session_state.points:
                st.info("Aucun point de prélèvement défini — allez dans **Paramètres → Points de prélèvement**.")
            else:
                p_col1, p_col2 = st.columns([3, 2])
                with p_col1:
                    point_labels = [
                        f"{pt['label']} — {pt.get('type','?')} — "
                        f"{'Critique' if pt.get('location_criticality',1)==4 else 'Important' if pt.get('location_criticality',1)==3 else 'Modéré' if pt.get('location_criticality',1)==2 else 'Limité'}"
                        for pt in st.session_state.points
                    ]
                    sel_idx = st.selectbox(
                        "Point de prélèvement",
                        list(range(len(point_labels))),
                        format_func=lambda i: point_labels[i],
                        key="new_prelev_point")
                    selected_point = st.session_state.points[sel_idx]
                    pt_type     = selected_point.get("type", "—")
                    pt_loc_crit = int(selected_point.get("location_criticality", 1))
                    pt_gelose   = selected_point.get("gelose", "—")
                    pt_room     = selected_point.get("room_class", "—")
                    type_icon   = "💨" if pt_type == "Air" else "🧴"
                    lc_col      = {"1": "#22c55e", "2": "#f59e0b", "3": "#ef4444"}.get(str(pt_loc_crit), "#94a3b8")
                    lc_lbl      = _loc_crit_label(pt_loc_crit)
                    st.markdown(f"""
                    <div style="background:#f0f9ff;border:1px solid #bae6fd;border-radius:8px;padding:12px;margin-top:4px">
                    <div style="font-size:.75rem;font-weight:700;color:#0369a1;margin-bottom:8px">{type_icon} Détails du point sélectionné</div>
                    <div style="display:grid;grid-template-columns:1fr 1fr;gap:6px">
                        <div style="background:#fff;border-radius:6px;padding:8px;border:1px solid #e0f2fe">
                        <div style="font-size:.6rem;color:#64748b;text-transform:uppercase">Type</div>
                        <div style="font-size:.85rem;font-weight:700;color:#0f172a;margin-top:2px">{pt_type}</div>
                        </div>
                        <div style="background:#dbeafe;border-radius:6px;padding:8px;border:1px solid #93c5fd">
                        <div style="font-size:.6rem;color:#1e40af;text-transform:uppercase">Classe ISO / GMP</div>
                        <div style="font-size:.85rem;font-weight:800;color:#1e40af;margin-top:2px">{pt_room if pt_room and pt_room != '—' else '—'}</div>
                        </div>
                        <div style="background:{lc_col}11;border-radius:6px;padding:8px;border:1px solid {lc_col}44">
                        <div style="font-size:.6rem;color:#64748b;text-transform:uppercase">Criticité lieu</div>
                        <div style="font-size:.85rem;font-weight:700;color:{lc_col};margin-top:2px">Nv.{pt_loc_crit} — {lc_lbl}</div>
                        </div>
                        <div style="background:#fff;border-radius:6px;padding:8px;border:1px solid #e0f2fe">
                        <div style="font-size:.6rem;color:#64748b;text-transform:uppercase">Gélose</div>
                        <div style="font-size:.85rem;font-weight:700;color:#1d4ed8;margin-top:2px">🧫 {pt_gelose}</div>
                        </div>
                    </div>
                    </div>""", unsafe_allow_html=True)

                with p_col2:
                    oper_list = [
                        o["nom"] + (" — " + o.get("profession", "") if o.get("profession") else "")
                        for o in st.session_state.operators
                    ]
                    if oper_list:
                        oper_sel = st.selectbox("Opérateur", ["— Sélectionner —"] + oper_list,
                                                key="new_prelev_oper_sel")
                        p_oper = oper_sel if oper_sel != "— Sélectionner —" else ""
                    else:
                        st.info("Aucun opérateur — ajoutez-en dans Paramètres")
                        p_oper = st.text_input("Opérateur (manuel)", placeholder="Nom",
                                               key="new_prelev_oper_manual")
                    p_date = st.date_input("Date prélèvement", value=datetime.today(),
                                           key="new_prelev_date")
                    j2_date_calc = next_working_day_offset(p_date, 2)
                    j7_date_calc = next_working_day_offset(p_date, 7)
                    st.markdown(f"""
                    <div style="background:#f0fdf4;border:1px solid #86efac;border-radius:8px;padding:8px;margin-top:6px;font-size:.7rem;color:#166534">
                    📅 J2 (2 jours ouvrés) : <strong>{j2_date_calc.strftime('%d/%m/%Y')}</strong><br>
                    📅 J7 (7 jours calendaires) : <strong>{j7_date_calc.strftime('%d/%m/%Y')}</strong>
                    </div>""", unsafe_allow_html=True)
                    p_commentaire = st.text_area("💬 Commentaire", placeholder="Remarque, contexte...",
                                                 height=70, key="new_prelev_commentaire")

                p_isolateur = ""
                p_poste     = "Poste 1"
                if str(pt_room).strip().upper() == "A":
                    st.markdown(
                        "<div style='background:#fef9c3;border:1px solid #fde047;border-radius:8px;"
                        "padding:10px 14px;margin-top:8px'>"
                        "<div style='font-size:.7rem;font-weight:700;color:#854d0e;margin-bottom:8px'>"
                        "🔬 Paramètres Zone Classe A</div>",
                        unsafe_allow_html=True)
                    iso_col, poste_col = st.columns(2)
                    with iso_col:
                        p_isolateur = st.radio("Isolateur",
                                               ["Iso 16/0724", "Iso 14/07169"],
                                               index=None, horizontal=True,
                                               key="new_prelev_isolateur")
                    with poste_col:
                        p_poste = st.radio("Poste",
                                           ["Poste 1", "Poste 2", "Commun"],
                                           index=None, horizontal=True,
                                           key="new_prelev_poste")
                    st.markdown("</div>", unsafe_allow_html=True)

                if st.button("💾 Enregistrer le prélèvement", use_container_width=True,
                            key="save_prelev", type="primary"):
                    if not p_oper:
                        st.error("⚠️ Veuillez sélectionner un opérateur.")
                    elif str(pt_room).strip().upper() == "A" and not p_isolateur:
                        st.error("⚠️ Veuillez sélectionner un isolateur.")
                    elif str(pt_room).strip().upper() == "A" and not p_poste:
                        st.error("⚠️ Veuillez sélectionner un poste.")
                    else:
                        pid = f"s{len(st.session_state.prelevements)+1}_{int(datetime.now().timestamp())}"
                        is_zone_a = str(pt_room).strip().upper() == "A"

                        sample = {
                            "id":                   pid,
                            "label":                selected_point["label"],
                            "type":                 selected_point.get("type"),
                            "gelose":               selected_point.get("gelose", "—"),
                            "room_class":           selected_point.get("room_class", ""),
                            "location_criticality": pt_loc_crit,
                            "operateur":            p_oper if p_oper else "Non renseigné",
                            "date":                 str(p_date) if p_date else str(today),
                            "archived":             False,
                            "num_isolateur":        p_isolateur if is_zone_a else "",
                            "poste":                p_poste     if is_zone_a else "",
                            "commentaire":          p_commentaire if p_commentaire else "",
                            "created_via":          "manuel",
                        }
                        st.session_state.prelevements.append(sample)
                        save_prelevements(st.session_state.prelevements)
                        for when, due in [("J2", j2_date_calc), ("J7", j7_date_calc)]:
                            st.session_state.schedules.append({
                                "id":        f"sch_{pid}_{when}",
                                "sample_id": pid,
                                "label":     sample["label"],
                                "due_date":  due.isoformat(),
                                "when":      when,
                                "status":    "pending",
                            })
                        save_schedules(st.session_state.schedules)
                        st.success(
                            f"✅ **{sample['label']}** enregistré !\n"
                            f"J2 → {j2_date_calc.strftime('%d/%m/%Y')} | "
                            f"J7 → {j7_date_calc.strftime('%d/%m/%Y')}"
                        )

                st.divider()
                st.markdown("#### 📋 Prélèvements en cours")
                actifs = [s for s in st.session_state.prelevements if not s.get("archived")]
                with st.expander(f"📋 Prélèvements en cours ({len(actifs)})", expanded=False):
                    if not actifs:
                        st.info("Aucun prélèvement en cours.")
                    for idx, samp in enumerate(st.session_state.prelevements):
                        if samp.get("archived"):
                            continue
                        col_info, col_edit, col_del = st.columns([5, 1, 1])
                        with col_info:
                            loc_c    = int(samp.get("location_criticality", 1))
                            lc_col_r = {"1": "#22c55e", "2": "#f59e0b", "3": "#ef4444"}.get(str(loc_c), "#94a3b8")
                            room_cl  = samp.get("room_class", "") or ""
                            room_badge = (
                                f"<span style='background:#dbeafe;color:#1e40af;border:1px solid #93c5fd;"
                                f"border-radius:4px;padding:1px 6px;font-size:.72rem;font-weight:800;"
                                f"margin-left:4px'>Cl.{room_cl}</span>"
                                if room_cl else ""
                            )
                            via_badge = (
                                "<span style='background:#ede9fe;color:#5b21b6;border-radius:4px;"
                                "padding:1px 6px;font-size:.65rem;margin-left:4px'>QR</span>"
                                if samp.get("created_via") == "qr_scan" else ""
                            )
                            _comment_html = (
                                f"<div style='font-size:.72rem;color:#6366f1;margin-top:3px'>"
                                f"💬 {samp['commentaire']}</div>"
                                if samp.get("commentaire") else ""
                            )
                            st.markdown(
                                f"<div style='background:#fff;border:1.5px solid #e2e8f0;"
                                f"border-radius:10px;padding:10px 16px;margin-bottom:6px'>"
                                f"<span style='font-weight:700'>{samp['label']}</span>"
                                f"{room_badge}{via_badge} "
                                f"<span style='color:#64748b;font-size:.8rem'>— {samp.get('type','—')} · "
                                f"<span style='color:{lc_col_r};font-weight:600'>Nv.{loc_c}</span> · "
                                f"{samp.get('date','—')} · {samp.get('operateur','—')}</span>"
                                f"{_comment_html}</div>",
                                unsafe_allow_html=True)
                        with col_edit:
                            if st.button("✏️", key=f"edit_prelev_btn_{samp['id']}",
                                         use_container_width=True):
                                st.session_state["edit_prelev_id"] = samp["id"]
                                st.rerun()
                        with col_del:
                            if st.button("🗑️", key=f"del_prelev_btn_{samp['id']}",
                                         use_container_width=True):
                                sid = samp["id"]
                                st.session_state.schedules    = [x for x in st.session_state.schedules    if x.get("sample_id") != sid]
                                st.session_state.prelevements = [x for x in st.session_state.prelevements if x["id"] != sid]
                                st.session_state.pending_identifications = [
                                    x for x in st.session_state.pending_identifications
                                    if x.get("sample_id") != sid
                                ]
                                save_schedules(st.session_state.schedules)
                                save_prelevements(st.session_state.prelevements)
                                save_pending_identifications(st.session_state.pending_identifications)
                                st.success(f"🗑️ Prélèvement **{samp['label']}** supprimé.")
                                st.rerun()

        # ── MODE SCAN QR ──────────────────────────────────────────────────────
        else:
            st.markdown("""
            <div style="background:linear-gradient(135deg,#eff6ff,#dbeafe);border:1.5px solid #93c5fd;
            border-radius:14px;padding:14px 18px;margin-bottom:14px">
            <div style="font-weight:700;color:#1e40af;font-size:.9rem;margin-bottom:6px">🔳 Comment scanner</div>
            <div style="font-size:.8rem;color:#1e293b;line-height:1.8">
                <strong>1.</strong> Imprimez les étiquettes depuis <em>Planning → Planning mensuel</em> 🖨️<br>
                <strong>2.</strong> Cliquez dans le champ ci-dessous, puis scannez l'étiquette<br>
                <strong>3.</strong> Le formulaire se pré-remplit — confirmez et enregistrez
            </div></div>""", unsafe_allow_html=True)

            qr_raw_input = st.text_input(
                "Zone de scan",
                key=f"qr_scan_input_{st.session_state['qr_counter']}",
                placeholder="Cliquez ici puis scannez l'étiquette QR...",
                label_visibility="collapsed",
                help="Si des caractères spéciaux apparaissent, configurez la douchette en layout EN-US.")

            qr_raw = qr_raw_input.strip() if qr_raw_input else ""
            if qr_raw:
                qr_fixed = _fix_qr_input(qr_raw)
                if qr_fixed != qr_raw:
                    st.caption("🔄 Correction automatique appliquée.")
                

            _scanned_data = None
            if qr_raw:
                _found = next((p for p in st.session_state.points
                               if str(p["id"]) == qr_raw), None)
                if _found:
                    _scanned_data = _found
                else:
                    st.markdown(
                        "<div style='background:#fffbeb;border:1.5px solid #fcd34d;"
                        "border-radius:10px;padding:12px 16px;margin-top:8px'>"
                        "<div style='font-weight:700;color:#92400e'>⚠️ Numéro non reconnu</div>"
                        "</div>", unsafe_allow_html=True)

            if _scanned_data:
                _lbl        = _scanned_data.get("label", "")
                _type       = _scanned_data.get("type", "")
                _rc         = _scanned_data.get("room_class", "")
                _lc         = int(_scanned_data.get("location_criticality", 1))
                _gel        = _scanned_data.get("gelose", "")
                _is_classea = _rc.strip().upper() == "A"
                lc_col_s    = {"1": "#22c55e", "2": "#f59e0b", "3": "#ef4444"}.get(str(_lc), "#94a3b8")

                st.markdown(
                    f"<div style='background:linear-gradient(135deg,#f0fdf4,#dcfce7);"
                    f"border:2.5px solid #22c55e;border-radius:14px;padding:16px 20px;margin:10px 0'>"
                    f"<div style='font-size:1rem;font-weight:700;color:#166534;margin-bottom:10px'>"
                    f"✅ Point reconnu — {_lbl}</div>"
                    f"<div style='display:grid;grid-template-columns:repeat(4,1fr);gap:8px'>"
                    f"<div style='background:#fff;border-radius:8px;padding:8px;text-align:center;"
                    f"border:1px solid #86efac'><div style='font-size:.58rem;color:#64748b;"
                    f"text-transform:uppercase'>Type</div>"
                    f"<div style='font-size:.85rem;font-weight:700'>{'💨' if _type=='Air' else '🧴'} {_type}</div></div>"
                    f"<div style='background:#dbeafe;border-radius:8px;padding:8px;text-align:center;"
                    f"border:1px solid #93c5fd'><div style='font-size:.58rem;color:#1e40af;"
                    f"text-transform:uppercase'>Classe</div>"
                    f"<div style='font-size:.85rem;font-weight:800;color:#1e40af'>{_rc or '—'}</div></div>"
                    f"<div style='background:{lc_col_s}11;border-radius:8px;padding:8px;text-align:center;"
                    f"border:1px solid {lc_col_s}44'><div style='font-size:.58rem;color:#64748b;"
                    f"text-transform:uppercase'>Criticité</div>"
                    f"<div style='font-size:.85rem;font-weight:700;color:{lc_col_s}'>Nv.{_lc}</div></div>"
                    f"<div style='background:#fff;border-radius:8px;padding:8px;text-align:center;"
                    f"border:1px solid #86efac'><div style='font-size:.58rem;color:#64748b;"
                    f"text-transform:uppercase'>Gélose</div>"
                    f"<div style='font-size:.85rem;font-weight:700;color:#1d4ed8'>🧫 {_gel[:14]}</div></div>"
                    f"</div></div>", unsafe_allow_html=True)

                sf1, sf2 = st.columns(2)
                with sf1:
                    oper_list_s = [
                        o["nom"] + (" — " + o.get("profession", "") if o.get("profession") else "")
                        for o in st.session_state.operators
                    ]
                    if oper_list_s:
                        scan_oper = st.selectbox("👤 Opérateur *",
                                                 ["— Sélectionner —"] + oper_list_s,
                                                 key="scan_oper_sel")
                        scan_oper = scan_oper if scan_oper != "— Sélectionner —" else ""
                    else:
                        scan_oper = st.text_input("👤 Opérateur *", placeholder="Nom",
                                                  key="scan_oper_manual")
                    scan_date = st.date_input("📅 Date", value=datetime.today(), key="scan_date")

                with sf2:
                    scan_isolateur = ""
                    scan_poste     = ""
                    if _is_classea:
                        st.markdown(
                            "<div style='background:#fef9c3;border:1px solid #fde047;"
                            "border-radius:8px;padding:8px 12px;margin-bottom:6px;"
                            "font-size:.7rem;font-weight:700;color:#854d0e'>🔬 Classe A</div>",
                            unsafe_allow_html=True)
                        scan_isolateur = st.radio("Isolateur",
                                                  ["Iso 16/0724", "Iso 14/07169"],
                                                  index=None, horizontal=True,
                                                  key="scan_iso_sel")
                        scan_poste = st.radio("Poste",
                                              ["Poste 1", "Poste 2", "Commun"],
                                              index=None, horizontal=True,
                                              key="scan_poste_sel")
                    scan_comment = st.text_area("💬 Commentaire", placeholder="Remarques...",
                                               height=80, key="scan_comment")

                _scan_j2 = next_working_day_offset(scan_date, 2)
                _scan_j7 = next_working_day_offset(scan_date, 7)
                st.markdown(
                    f"<div style='background:#f0fdf4;border:1px solid #86efac;border-radius:8px;"
                    f"padding:8px 14px;font-size:.72rem;color:#166534;margin-bottom:10px'>"
                    f"📅 J2 → <strong>{_scan_j2.strftime('%d/%m/%Y')}</strong> &nbsp;·&nbsp; "
                    f"📅 J7 → <strong>{_scan_j7.strftime('%d/%m/%Y')}</strong></div>",
                    unsafe_allow_html=True)

                sbc1, sbc2 = st.columns([3, 1])
                with sbc1:
                    if st.button(f"💾 Enregistrer — {_lbl}", use_container_width=True,
                                 type="primary", key="scan_save_btn"):
                        if not scan_oper:
                            st.error("⚠️ Veuillez sélectionner un opérateur.")
                        elif _is_classea and not scan_isolateur:
                            st.error("⚠️ Veuillez sélectionner un isolateur.")
                        elif _is_classea and not scan_poste:
                            st.error("⚠️ Veuillez sélectionner un poste.")
                        else:
                            _pid = f"s{len(st.session_state.prelevements)+1}_{int(datetime.now().timestamp())}"
                            _sample_scan = {
                                "id": _pid, "label": _lbl, "type": _type, "gelose": _gel,
                                "room_class": _rc, "location_criticality": _lc,
                                "operateur": scan_oper, "date": str(scan_date),
                                "archived": False,
                                "num_isolateur": scan_isolateur if _is_classea else "",
                                "poste":         scan_poste     if _is_classea else "",
                                "commentaire": scan_comment or "",
                                "created_via": "qr_scan",
                            }
                            st.session_state.prelevements.append(_sample_scan)
                            save_prelevements(st.session_state.prelevements)
                            for _when, _due in [("J2", _scan_j2), ("J7", _scan_j7)]:
                                st.session_state.schedules.append({
                                    "id": f"sch_{_pid}_{_when}", "sample_id": _pid,
                                    "label": _lbl, "due_date": _due.isoformat(),
                                    "when": _when, "status": "pending",
                                })
                            save_schedules(st.session_state.schedules)
                            iso_info = f" · {scan_isolateur} · {scan_poste}" if _is_classea else ""
                            st.success(
                                f"✅ **{_lbl}** enregistré{iso_info}\n"
                                f"J2 → {_scan_j2.strftime('%d/%m/%Y')} · "
                                f"J7 → {_scan_j7.strftime('%d/%m/%Y')}"
                            )
                            st.session_state["qr_counter"] += 1
                            st.rerun()
                with sbc2:
                    if st.button("✕ Annuler", use_container_width=True, key="scan_cancel_btn"):
                        st.session_state["qr_counter"] += 1
                        st.rerun()

    # ──────────────────────────────────────────────────────────────────────────
    # HELPERS LECTURES
    # ──────────────────────────────────────────────────────────────────────────
    def _sort_schedules(schedule_list, sort_key):
        def _get_smp(s):
            return next((p for p in st.session_state.prelevements
                         if p["id"] == s["sample_id"]), {})
        if sort_key == "label":
            return sorted(schedule_list, key=lambda s: s.get("label", "").lower())
        elif sort_key == "operateur":
            return sorted(schedule_list, key=lambda s: _get_smp(s).get("operateur", "").lower())
        elif sort_key == "date_prelevement":
            return sorted(schedule_list, key=lambda s: _get_smp(s).get("date", "9999-12-31"))
        elif sort_key == "echeance":
            return sorted(schedule_list, key=lambda s: s.get("due_date", "9999-12-31"))
        return schedule_list

    def _render_lecture_card(s, tab_prefix=""):
        sched_date = datetime.fromisoformat(s["due_date"]).date()
        days_since = _days_since_sample(s)
        min_days   = 7 if s["when"] == "J7" else 0
        is_late    = sched_date <= today and days_since >= min_days

        border_col = "#ef4444" if is_late else "#3b82f6"
        bg_col     = "#fef2f2" if is_late else "#eff6ff"
        badge_col  = "#dc2626" if is_late else "#1d4ed8"

        if is_late:
            status_txt = "EN RETARD"
        elif s["when"] == "J7" and days_since < 7:
            status_txt = f"disponible dans {7 - days_since}j"
        else:
            status_txt = f"dans {(sched_date - today).days}j"

        smp          = next((p for p in st.session_state.prelevements if p["id"] == s["sample_id"]), None)
        pt_type      = smp.get("type",       "?") if smp else "?"
        pt_gelose    = smp.get("gelose",     "?") if smp else "?"
        pt_oper      = smp.get("operateur",  "?") if smp else "?"
        pt_date_p    = smp.get("date",       "—") if smp else "—"
        pt_comment   = smp.get("commentaire","") if smp else ""
        comment_short = (pt_comment[:40] + "…") if len(pt_comment) > 40 else (pt_comment or "—")

        with st.container():
            st.markdown(f"""
            <div style="background:{bg_col};border:1.5px solid {border_col};border-radius:10px;
                        padding:14px 16px;margin-bottom:8px">
            <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:8px">
                <div>
                <span style="font-weight:700;font-size:.9rem;color:#0f172a">{s['label']}</span>
                <span style="background:{border_col};color:#fff;font-size:.6rem;font-weight:700;
                            padding:2px 8px;border-radius:10px;margin-left:8px">{s['when']}</span>
                <span style="color:{badge_col};font-size:.65rem;font-weight:600;margin-left:6px">{status_txt}</span>
                </div>
                <span style="font-size:.75rem;color:#475569">📅 Échéance : {s['due_date'][:10]}</span>
            </div>
            <div style="display:grid;grid-template-columns:repeat(5,1fr);gap:6px">
                <div style="background:#fff;border-radius:6px;padding:6px 8px;border:1px solid #e2e8f0">
                <div style="font-size:.55rem;color:#64748b;text-transform:uppercase">Type</div>
                <div style="font-size:.75rem;font-weight:600;color:#0f172a">{pt_type}</div>
                </div>
                <div style="background:#f0fdf4;border-radius:6px;padding:6px 8px;border:1px solid #86efac">
                <div style="font-size:.55rem;color:#166534;text-transform:uppercase">Date prélèv.</div>
                <div style="font-size:.75rem;font-weight:700;color:#166534">📅 {pt_date_p}</div>
                </div>
                <div style="background:#fff;border-radius:6px;padding:6px 8px;border:1px solid #e2e8f0">
                <div style="font-size:.55rem;color:#64748b;text-transform:uppercase">Gélose</div>
                <div style="font-size:.75rem;font-weight:600;color:#1d4ed8">🧫 {pt_gelose}</div>
                </div>
                <div style="background:#fff;border-radius:6px;padding:6px 8px;border:1px solid #e2e8f0">
                <div style="font-size:.55rem;color:#64748b;text-transform:uppercase">Opérateur</div>
                <div style="font-size:.75rem;font-weight:600;color:#0f172a">{pt_oper}</div>
                </div>
                <div style="background:#fffbeb;border-radius:6px;padding:6px 8px;border:1px solid #fde047">
                <div style="font-size:.55rem;color:#92400e;text-transform:uppercase">Commentaire</div>
                <div style="font-size:.72rem;font-weight:600;color:#78350f"
                    title="{pt_comment}">💬 {comment_short}</div>
                </div>
            </div>
            </div>""", unsafe_allow_html=True)

            bc1, bc2 = st.columns([3, 1])
            with bc1:
                if st.button(f"🔬 Traiter cette lecture ({s['when']})",
                             key=f"{tab_prefix}proc_{s['id']}", use_container_width=True):
                    st.session_state.current_process = s["id"]
                    st.rerun()
            with bc2:
                if st.button("🗑️ Supprimer", key=f"{tab_prefix}del_sch_{s['id']}",
                             use_container_width=True):
                    sid = s.get("sample_id")
                    st.session_state.schedules    = [x for x in st.session_state.schedules    if x["sample_id"] != sid]
                    st.session_state.prelevements = [x for x in st.session_state.prelevements if x["id"]         != sid]
                    st.session_state.pending_identifications = [
                        x for x in st.session_state.pending_identifications
                        if x.get("sample_id") != sid
                    ]
                    save_schedules(st.session_state.schedules)
                    save_prelevements(st.session_state.prelevements)
                    save_pending_identifications(st.session_state.pending_identifications)
                    st.success("Prélèvement supprimé.")
                    st.rerun()

    def _render_traitement_lecture(proc_id):
        proc = next((x for x in st.session_state.schedules if x["id"] == proc_id), None)
        if not proc:
            return
        smp      = next((p for p in st.session_state.prelevements if p["id"] == proc["sample_id"]), None)

        if proc["when"] == "J7":
            st.markdown(
                "<div style='background:#fff7ed;border:1px solid #fed7aa;border-radius:8px;"
                "padding:8px 14px;margin-bottom:8px;font-size:.78rem;color:#9a3412'>"
                "💡 Vous traitez la lecture <b>J7</b>.</div>",
                unsafe_allow_html=True)
            if st.button("↩️ Revenir à la lecture J2", key=f"back_j2_{proc_id}"):
                _j2 = next((x for x in st.session_state.schedules
                            if x["sample_id"] == proc["sample_id"] and x["when"] == "J2"), None)
                if _j2:
                    _j2["status"] = "pending"
                proc["status"] = "pending"
                save_schedules(st.session_state.schedules)
                st.session_state.pending_identifications = [
                    x for x in st.session_state.pending_identifications
                    if x.get("sample_id") != proc["sample_id"]
                ]
                save_pending_identifications(st.session_state.pending_identifications)
                if smp and smp.get("archived"):
                    smp["archived"] = False
                    st.session_state.archived_samples = [
                        x for x in st.session_state.archived_samples if x.get("id") != smp["id"]
                    ]
                    save_archived_samples(st.session_state.archived_samples)
                    save_prelevements(st.session_state.prelevements)
                st.session_state.current_process = None
                st.success("↩️ J2 et J7 remises en attente.")
                st.rerun()
            st.markdown("---")

        lc1, lc2 = st.columns([2, 2])
        with lc1:
            res = st.radio(
                "Résultat",
                ["✅ Négatif (0 colonie)", "🔴 Positif (colonies détectées)"],
                index=0, key=f"res_{proc_id}")
        with lc2:
            ncol = st.number_input("Nombre de colonies (UFC)", min_value=1, value=1,
                                   key=f"ncol_{proc_id}") if "Positif" in res else 0

        btn1, _, btn3 = st.columns([3, 1, 1])
        with btn1:
            if st.button("💾 Valider la lecture", use_container_width=True,
                         type="primary", key=f"valider_{proc_id}"):
                proc["status"]    = "done"
                proc["colonies"]  = ncol
                proc["date_read"] = str(datetime.today().date())
                save_schedules(st.session_state.schedules)

                if "Positif" in res and ncol > 0:
                    label_smp = smp.get("label", "?") if smp else "?"
                    st.session_state.pending_identifications.append({
                        "id":        str(uuid.uuid4()),
                        "sample_id": proc["sample_id"],
                        "label":     label_smp,
                        "when":      proc["when"],
                        "colonies":  ncol,
                        "date":      str(datetime.today().date()),
                        "status":    "pending",
                    })
                    save_pending_identifications(st.session_state.pending_identifications)

                    if proc["when"] == "J2":
                        _j7 = next((x for x in st.session_state.schedules
                                    if x["sample_id"] == proc["sample_id"] and x["when"] == "J7"), None)
                        if _j7:
                            _j7["status"] = "skipped"
                            save_schedules(st.session_state.schedules)
                        st.success(f"🔴 J2 positive — {ncol} UFC · identification en attente.")
                    elif proc["when"] == "J7" and smp and not smp.get("archived"):
                        smp["archived"] = True
                        st.session_state.archived_samples.append(smp)
                        save_archived_samples(st.session_state.archived_samples)
                        save_prelevements(st.session_state.prelevements)
                        st.success(f"🔴 J7 positive — {ncol} UFC · identification en attente.")

                else:
                    if proc["when"] == "J2":
                        j7_sch = next((x for x in st.session_state.schedules
                                       if x["sample_id"] == proc["sample_id"] and x["when"] == "J7"), None)
                        st.success(
                            f"✅ J2 négative — J7 prévue le "
                            f"{j7_sch['due_date'][:10] if j7_sch else '?'}."
                        )
                    elif proc["when"] == "J7" and smp and not smp.get("archived"):
                        smp["archived"] = True
                        st.session_state.archived_samples.append(smp)
                        save_archived_samples(st.session_state.archived_samples)
                        save_prelevements(st.session_state.prelevements)
                        _lc_neg = 1
                        try:
                            _lc_neg = int(smp.get("location_criticality", 1) or 1)
                        except (ValueError, TypeError):
                            pass
                        st.session_state.surveillance.append({
                            "date":                 str(datetime.today().date()),   # date de lecture (inchangé)
                            "date_prelevement":     smp.get("date", ""),            # ← AJOUTER cette ligne
                            "prelevement":          smp.get("label", "?"),
                            "sample_id":            proc["sample_id"],
                            "germ_match":           "Négatif",
                            "germ_saisi":           "Négatif",
                            "ufc":                  0,
                            "ufc_total":            0,
                            "germ_score":           0,
                            "location_criticality": _lc_neg,
                            "total_score":          0,
                            "status":               "ok",
                            "operateur":            smp.get("operateur", "?"),
                            "remarque":             "",
                            "readings":             "J7",
                            "room_class":           smp.get("room_class", ""),
                        })
                        save_surveillance(st.session_state.surveillance)
                        st.success("✅ J7 négative — prélèvement archivé.")

                st.session_state.current_process = None
                st.rerun()

        with btn3:
            if st.button("✕ Annuler", use_container_width=True, key=f"cancel_{proc_id}"):
                st.session_state.current_process = None
                st.rerun()

    def _valider_negatif(sch_id):
        proc = next((x for x in st.session_state.schedules if x["id"] == sch_id), None)
        if not proc:
            return
        smp = next((p for p in st.session_state.prelevements if p["id"] == proc["sample_id"]), None)
        proc["status"]    = "done"
        proc["colonies"]  = 0
        proc["date_read"] = str(datetime.today().date())
        save_schedules(st.session_state.schedules)
        if proc["when"] == "J7" and smp and not smp.get("archived"):
            smp["archived"] = True
            st.session_state.archived_samples.append(smp)
            save_archived_samples(st.session_state.archived_samples)
            save_prelevements(st.session_state.prelevements)
            _lc_neg = 1
            try:
                _lc_neg = int(smp.get("location_criticality", 1) or 1)
            except (ValueError, TypeError):
                pass
            st.session_state.surveillance.append({
                "date":                 str(datetime.today().date()),   # date de lecture (inchangé)
                "date_prelevement":     smp.get("date", ""),            # ← AJOUTER cette ligne
                "prelevement":          smp.get("label", "?"),
                "sample_id":            proc["sample_id"],
                "germ_match":           "Négatif",
                "germ_saisi":           "Négatif",
                "ufc":                  0,
                "ufc_total":            0,
                "germ_score":           0,
                "location_criticality": _lc_neg,
                "total_score":          0,
                "status":               "ok",
                "operateur":            smp.get("operateur", "?"),
                "remarque":             "",
                "readings":             "J7",
                "room_class":           smp.get("room_class", ""),
            })
            save_surveillance(st.session_state.surveillance)

    def _render_batch_negatif_section(pending_list, tab_key):
        if not pending_list:
            return
        batch_mode_key = f"batch_mode_{tab_key}"
        if batch_mode_key not in st.session_state:
            st.session_state[batch_mode_key] = False
        col_a, col_b = st.columns([2, 3])
        with col_a:
            btn_label = "✕ Fermer sélection" if st.session_state[batch_mode_key] else "☑️ Sélection multiple — tout négatif"
            if st.button(btn_label, key=f"toggle_batch_{tab_key}", use_container_width=True):
                st.session_state[batch_mode_key] = not st.session_state[batch_mode_key]
                for s in pending_list:
                    st.session_state.pop(f"batch_chk_{tab_key}_{s['id']}", None)
                st.session_state.pop(f"sel_all_{tab_key}", None)
                st.rerun()
        if not st.session_state[batch_mode_key]:
            return
        with col_b:
            if st.checkbox("✅ Tout sélectionner", key=f"sel_all_{tab_key}"):
                for s in pending_list:
                    st.session_state[f"batch_chk_{tab_key}_{s['id']}"] = True

    def _render_batch_confirm(pending_list, tab_key):
        if not st.session_state.get(f"batch_mode_{tab_key}"):
            return
        selected_ids = [
            s["id"] for s in pending_list
            if st.session_state.get(f"batch_chk_{tab_key}_{s['id']}", False)
        ]
        n_sel = len(selected_ids)
        st.markdown(
            f"<div style='font-size:.78rem;color:#475569;margin:6px 0'>"
            f"{n_sel} / {len(pending_list)} lecture(s) sélectionnée(s)</div>",
            unsafe_allow_html=True)
        if n_sel > 0:
            pl = "s" if n_sel > 1 else ""
            if st.button(f"✅ Valider {n_sel} lecture{pl} comme négative{pl}",
                         key=f"batch_confirm_{tab_key}", type="primary",
                         use_container_width=True):
                for sch_id in selected_ids:
                    _valider_negatif(sch_id)
                    st.session_state.pop(f"batch_chk_{tab_key}_{sch_id}", None)
                st.session_state[f"batch_mode_{tab_key}"] = False
                st.session_state.pop(f"sel_all_{tab_key}", None)
                st.success(f"✅ {n_sel} lecture{pl} validée{pl} comme négative{pl} !")
                st.rerun()

    # ══════════════════════════════════════════════════════════════════════════
    # ONGLET 2 — LECTURE J2
    # ══════════════════════════════════════════════════════════════════════════
    with tab_j2:
        st.markdown("#### 📖 Lectures J2 en attente")
        _active_sids = {p["id"] for p in st.session_state.prelevements if not p.get("archived")}
        pending_j2   = [s for s in st.session_state.schedules
                        if s["when"] == "J2" and s["status"] == "pending"
                        and s.get("sample_id") in _active_sids]
        overdue_j2   = [s for s in pending_j2
                        if datetime.fromisoformat(s["due_date"]).date() <= today]
        upcoming_j2  = [s for s in pending_j2
                        if datetime.fromisoformat(s["due_date"]).date() > today]

        if not pending_j2:
            st.success("✅ Aucune lecture J2 en attente — tout est à jour !")
        else:
            if overdue_j2:
                st.markdown(
                    f'<div style="background:#fef2f2;border:1.5px solid #fca5a5;border-radius:10px;'
                    f'padding:12px 16px;margin-bottom:12px">'
                    f'<span style="color:#dc2626;font-weight:700">'
                    f'🔔 {len(overdue_j2)} lecture(s) J2 en retard</span></div>',
                    unsafe_allow_html=True)
            if upcoming_j2:
                st.markdown(
                    f'<div style="background:#f0fdf4;border:1px solid #86efac;border-radius:10px;'
                    f'padding:10px 16px;margin-bottom:12px">'
                    f'<span style="color:#16a34a;font-size:.8rem">'
                    f'📆 {len(upcoming_j2)} lecture(s) J2 à venir</span></div>',
                    unsafe_allow_html=True)

            _render_batch_negatif_section(pending_j2, "j2")
            st.divider()

            batch_active_j2 = st.session_state.get("batch_mode_j2", False)
            sort_col_j2, filter_col_j2 = st.columns(2)
            with sort_col_j2:
                sort_j2 = st.selectbox(
                    "🔃 Trier par",
                    options=["echeance", "label", "operateur", "date_prelevement"],
                    format_func=lambda x: {
                        "echeance": "📅 Échéance", "label": "📍 Lieu (A→Z)",
                        "operateur": "👤 Opérateur (A→Z)", "date_prelevement": "🗓️ Date prélèvement",
                    }[x], key="sort_j2")
            with filter_col_j2:
                all_labels_j2 = sorted({s.get("label", "—") for s in overdue_j2 + upcoming_j2})
                filter_j2 = st.multiselect("🔍 Filtrer par point", options=all_labels_j2,
                                           default=[], key="filter_j2",
                                           placeholder="Tous les points…")

            filtered_j2 = [s for s in overdue_j2 + upcoming_j2
                           if not filter_j2 or s.get("label") in filter_j2]
            sorted_j2 = _sort_schedules(filtered_j2, sort_j2)

            for s in sorted_j2:
                if batch_active_j2:
                    chk_col, card_col = st.columns([0.35, 9.65])
                    with chk_col:
                        st.markdown("<div style='margin-top:28px'>", unsafe_allow_html=True)
                        st.checkbox("", key=f"batch_chk_j2_{s['id']}")
                        st.markdown("</div>", unsafe_allow_html=True)
                    with card_col:
                        _render_lecture_card(s, "j2_")
                else:
                    _render_lecture_card(s, "j2_")
                    if st.session_state.get("current_process") == s["id"]:
                        _cp = next((x for x in st.session_state.schedules
                                    if x["id"] == st.session_state.current_process), None)
                        if _cp and _cp.get("when") == "J2":
                            _render_traitement_lecture(st.session_state.current_process)

            _render_batch_confirm(pending_j2, "j2")

    # ══════════════════════════════════════════════════════════════════════════
    # ONGLET 3 — LECTURE J7
    # ══════════════════════════════════════════════════════════════════════════
    with tab_j7:
        st.markdown("#### 📗 Lectures J7 en attente")
        _active_sids_j7 = {p["id"] for p in st.session_state.prelevements if not p.get("archived")}

        def _j2_done_for(sample_id):
            j2 = next((x for x in st.session_state.schedules
                       if x["sample_id"] == sample_id and x["when"] == "J2"), None)
            return j2 is None or j2["status"] == "done"

        all_pending_j7 = [s for s in st.session_state.schedules
                          if s["when"] == "J7" and s["status"] == "pending"
                          and s.get("sample_id") in _active_sids_j7
                          and _j2_done_for(s["sample_id"])]

        overdue_j7  = [s for s in all_pending_j7
                       if datetime.fromisoformat(s["due_date"]).date() <= today
                       and _days_since_sample(s) >= 7]
        upcoming_j7 = [s for s in all_pending_j7 if s not in overdue_j7]
        pending_j7  = all_pending_j7

        if not all_pending_j7:
            st.success("✅ Aucune lecture J7 en attente — tout est à jour !")
        else:
            if overdue_j7:
                st.markdown(
                    f'<div style="background:#fef2f2;border:1.5px solid #fca5a5;border-radius:10px;'
                    f'padding:12px 16px;margin-bottom:12px">'
                    f'<span style="color:#dc2626;font-weight:700">'
                    f'🔔 {len(overdue_j7)} lecture(s) J7 en retard</span></div>',
                    unsafe_allow_html=True)
            if upcoming_j7:
                not_ready   = [s for s in upcoming_j7 if _days_since_sample(s) < 7]
                truly_ahead = [s for s in upcoming_j7 if _days_since_sample(s) >= 7]
                if not_ready:
                    st.markdown(
                        f'<div style="background:#fef9c3;border:1px solid #fde047;border-radius:10px;'
                        f'padding:10px 16px;margin-bottom:8px">'
                        f'<span style="color:#92400e;font-size:.8rem">'
                        f'⏳ {len(not_ready)} lecture(s) J7 pas encore disponible(s) (< 7 jours)'
                        f'</span></div>', unsafe_allow_html=True)
                if truly_ahead:
                    st.markdown(
                        f'<div style="background:#eff6ff;border:1px solid #93c5fd;border-radius:10px;'
                        f'padding:10px 16px;margin-bottom:12px">'
                        f'<span style="color:#1d4ed8;font-size:.8rem">'
                        f'📆 {len(truly_ahead)} lecture(s) J7 à venir</span></div>',
                        unsafe_allow_html=True)

            _render_batch_negatif_section(pending_j7, "j7")
            st.divider()

            batch_active_j7 = st.session_state.get("batch_mode_j7", False)
            sort_col_j7, filter_col_j7 = st.columns(2)
            with sort_col_j7:
                sort_j7 = st.selectbox(
                    "🔃 Trier par",
                    options=["echeance", "label", "operateur", "date_prelevement"],
                    format_func=lambda x: {
                        "echeance": "📅 Échéance", "label": "📍 Lieu (A→Z)",
                        "operateur": "👤 Opérateur (A→Z)", "date_prelevement": "🗓️ Date prélèvement",
                    }[x], key="sort_j7")
            with filter_col_j7:
                all_labels_j7 = sorted({s.get("label", "—") for s in all_pending_j7})
                filter_j7 = st.multiselect("🔍 Filtrer par point", options=all_labels_j7,
                                           default=[], key="filter_j7",
                                           placeholder="Tous les points…")

            filtered_j7 = [s for s in overdue_j7 + upcoming_j7
                           if not filter_j7 or s.get("label") in filter_j7]
            sorted_j7 = _sort_schedules(filtered_j7, sort_j7)

            for s in sorted_j7:
                if batch_active_j7:
                    chk_col, card_col = st.columns([0.35, 9.65])
                    with chk_col:
                        st.markdown("<div style='margin-top:28px'>", unsafe_allow_html=True)
                        st.checkbox("", key=f"batch_chk_j7_{s['id']}")
                        st.markdown("</div>", unsafe_allow_html=True)
                    with card_col:
                        _render_lecture_card(s, "j7_")
                else:
                    _render_lecture_card(s, "j7_")
                    if st.session_state.get("current_process") == s["id"]:
                        _cp = next((x for x in st.session_state.schedules
                                    if x["id"] == st.session_state.current_process), None)
                        if _cp and _cp.get("when") == "J7":
                            _render_traitement_lecture(st.session_state.current_process)

            _render_batch_confirm(pending_j7, "j7")

    # ══════════════════════════════════════════════════════════════════════════
    # ONGLET 4 — IDENTIFICATIONS EN ATTENTE
    # ══════════════════════════════════════════════════════════════════════════

    with tab_ident:
        from datetime import date

        st.markdown("#### 🔴 Identifications en attente")

        # ── Popup mesures correctives ─────────────────────────────
        if st.session_state.get("_show_mesures_popup"):
            _render_mesures_correctives(
                st.session_state["_show_mesures_popup"],
                entry_idx=None,
                popup_mode=True,
            )

        # ── sécurité germes ───────────────────────────────────────
        germ_names = sorted([g["name"] for g in st.session_state.get("germs", [])])

        # ── couleurs criticité ────────────────────────────────────
        LOC_CRIT_COLORS = {
            "1": "#22c55e",
            "2": "#0babf5",
            "3": "#ee811a",
            "4": "#f50b0b",
        }

        LOC_CRIT_LABELS = {
            "1": "Limité",
            "2": "Modéré",
            "3": "Important",
            "4": "Critique",
        }

        def _j7_done_or_absent(sample_id):
            j7 = next(
                (x for x in st.session_state.schedules
                if x["sample_id"] == sample_id and x["when"] == "J7"),
                None
            )
            return j7 is None or j7["status"] in ("done", "skipped")

        # ── filtrage ──────────────────────────────────────────────
        _all_pending = [
            p for p in st.session_state.pending_identifications
            if p.get("status") == "pending"
            and _j7_done_or_absent(p["sample_id"])
        ]

        grouped = {}

        for p in _all_pending:
            sid = p["sample_id"]

            if sid not in grouped:
                grouped[sid] = {
                    "sample_id": sid,
                    "label": p["label"],
                    "date": p["date"],
                    "entries": [],
                    "when_list": [],
                    "colonies": 0,
                }

            grouped[sid]["entries"].append(p)
            grouped[sid]["when_list"].append(p["when"])

            if p["when"] == "J7" or grouped[sid]["colonies"] == 0:
                grouped[sid]["colonies"] = p["colonies"]

        pending_ids_grouped = list(grouped.values())

        # ── affichage ─────────────────────────────────────────────
        if not pending_ids_grouped:
            st.success("✅ Aucune identification en attente.")

        else:
            for pg in pending_ids_grouped:

                _sid       = pg["sample_id"]
                _when_str  = " + ".join(sorted(set(pg["when_list"])))
                _ufc       = pg["colonies"]
                _label     = pg["label"]
                _date      = pg["date"]
                _entries   = pg["entries"]   # ✅ CORRECTION Bug 3 : défini ici

                # ── prélèvement SAFE ───────────────────────────────
                smp = next(
                    (p for p in st.session_state.prelevements if p["id"] == _sid),
                    None
                )

                pt_oper  = smp.get("operateur", "?") if smp else "?"
                pt_class = smp.get("room_class", "") if smp else ""

                _comment_prelev = (
                    smp.get("commentaire", "").strip()
                    if smp else ""
                )

                _date_prelev = (
                    date.fromisoformat(smp["date"])
                    if smp and smp.get("date")
                    else date.today()
                )

                _key          = _sid.replace("-", "_")
                germs_list_key = f"germs_list_{_key}"

                if germs_list_key not in st.session_state:
                    st.session_state[germs_list_key] = [
                        {"germ": "— Sélectionner un germe —", "ufc": 0}
                    ]

                # ── expander ───────────────────────────────────────
                with st.expander(
                    f"🔴 {_label} — {_when_str} — {_ufc} UFC — Prélevé le {_date_prelev}",
                    expanded=True
                ):

                    # ── COMMENTAIRE (uniquement si existe) ─────────
                    if _comment_prelev:
                        st.markdown(
                            f"""
                            <div style='background:#f0f9ff;border:1px solid #bae6fd;
                            border-radius:8px;padding:8px 12px;margin-bottom:8px;
                            font-size:.75rem;color:#0369a1'>
                            💬 <b>Commentaire prélèvement :</b><br>{_comment_prelev}
                            </div>
                            """,
                            unsafe_allow_html=True
                        )

                    # ── criticité ───────────────────────────────────
                    loc_crit = int(
                        smp.get("location_criticality", 1) if smp else 1
                    )

                    _lc_col = LOC_CRIT_COLORS.get(str(loc_crit), "#94a3b8")
                    _lc_lbl = LOC_CRIT_LABELS.get(str(loc_crit), "?")

                    st.markdown(
                        f"<div style='background:{_lc_col}11;border:1px solid {_lc_col}44;"
                        f"border-radius:8px;padding:8px 12px;margin-bottom:10px;"
                        f"font-size:.75rem;font-weight:700;color:{_lc_col}'>"
                        f"🏷️ Criticité : Niveau {loc_crit} — {_lc_lbl}</div>",
                        unsafe_allow_html=True
                    )

                    # ── germes ─────────────────────────────────────
                    current_germs  = st.session_state[germs_list_key]
                    germs_to_remove = []

                    for gi, g in enumerate(current_germs):

                        cols = st.columns([3, 1, 0.4])

                        with cols[0]:
                            selected = st.selectbox(
                                f"Germe {gi+1}",
                                ["— Sélectionner un germe —"] + germ_names,
                                index=(
                                    ["— Sélectionner un germe —"] + germ_names
                                ).index(g["germ"])
                                if g["germ"] in germ_names else 0,
                                key=f"germ_{_key}_{gi}"
                            )
                            current_germs[gi]["germ"] = selected

                        with cols[1]:
                            current_germs[gi]["ufc"] = st.number_input(
                                "UFC",
                                min_value=0,
                                value=int(g["ufc"]),
                                step=1,
                                key=f"ufc_{_key}_{gi}"
                            )

                        with cols[2]:
                            if gi > 0 and st.button("🗑️", key=f"del_{_key}_{gi}"):
                                germs_to_remove.append(gi)

                    for i in reversed(germs_to_remove):
                        st.session_state[germs_list_key].pop(i)
                        st.rerun()

                    if st.button("➕ Ajouter un germe", key=f"add_{_key}"):
                        st.session_state[germs_list_key].append(
                            {"germ": "— Sélectionner un germe —", "ufc": 0}
                        )
                        st.rerun()
                    # ── APERÇU SCORE (preview dynamique) ───────────────────
                    valid_germs_preview = [
                        g for g in current_germs
                        if g["germ"] and g["germ"] != "— Sélectionner un germe —"
                    ]

                    if valid_germs_preview:
                        scored_preview = []
                        for vg in valid_germs_preview:
                            gobj = next(
                                (g for g in st.session_state.germs if g["name"] == vg["germ"]),
                                None
                            )
                            if gobj:
                                gs = (
                                    int(gobj.get("pathogenicity", 1))
                                    * int(gobj.get("resistance", 1))
                                    * int(gobj.get("dissemination", 1))
                                )
                                scored_preview.append({
                                    "name": vg["germ"],
                                    "score": gs,
                                    "ufc": vg["ufc"],
                                })

                        if scored_preview:
                            worst_prev     = max(scored_preview, key=lambda x: x["score"])
                            ts_prev        = loc_crit * worst_prev["score"]
                            st_prev, _, sc_prev = _evaluate_score(ts_prev)
                            ufc_total_prev = sum(s["ufc"] for s in scored_preview)

                            preview_rows = "".join(
                                f"<tr>"
                                f"<td style='padding:2px 8px;color:#475569'>{s['name']}</td>"
                                f"<td style='padding:2px 8px;text-align:center;color:#475569'>{s['ufc']} UFC</td>"
                                f"<td style='padding:2px 8px;text-align:center;font-weight:700;"
                                f"color:{'#ef4444' if s['name'] == worst_prev['name'] else '#64748b'}'>"
                                f"{s['score']}{'  👑' if s['name'] == worst_prev['name'] else ''}</td>"
                                f"</tr>"
                                for s in scored_preview
                            )

                            st.markdown(
                                f"""<div style="background:{sc_prev}11;border:1.5px solid {sc_prev}44;
                                border-radius:8px;padding:10px 14px;margin-top:8px;margin-bottom:10px">
                                <div style="font-size:.6rem;color:#475569;text-transform:uppercase;
                                font-weight:700;margin-bottom:6px">
                                Aperçu score — germe le plus critique 👑</div>
                                <table style="width:100%;border-collapse:collapse;font-size:.72rem;margin-bottom:8px">
                                    <tr style="border-bottom:1px solid #e2e8f0">
                                        <th style="padding:2px 8px;text-align:left;color:#94a3b8">Germe</th>
                                        <th style="padding:2px 8px;text-align:center;color:#94a3b8">UFC</th>
                                        <th style="padding:2px 8px;text-align:center;color:#94a3b8">Score germe</th>
                                    </tr>
                                    {preview_rows}
                                    <tr style="border-top:2px solid #e2e8f0;background:#f0fdf4">
                                        <td style="padding:4px 8px;font-weight:800;color:#166534">Σ UFC TOTAL</td>
                                        <td style="padding:4px 8px;text-align:center;font-weight:900;
                                        color:#166534;font-size:.85rem">{ufc_total_prev}</td>
                                        <td style="padding:4px 8px;text-align:center;font-size:.65rem;
                                        color:#64748b">somme des germes</td>
                                    </tr>
                                </table>
                                <div style="display:flex;align-items:center;gap:12px">
                                    <div style="font-size:1.6rem;font-weight:900;color:{sc_prev}">{ts_prev}</div>
                                    <div style="font-size:.72rem;color:#475569">
                                        Lieu {loc_crit} × Germe le + critique {worst_prev['score']}<br>
                                        <span style="font-weight:700;color:{sc_prev}">
                                        {'🚨 ACTION' if st_prev == 'action' else '⚠️ ALERTE' if st_prev == 'alert' else '✅ Conforme'}
                                        </span>
                                    </div>
                                </div>
                                </div>""",
                                unsafe_allow_html=True
                            )
                    # ── REMARQUE ────────────────────────────────────
                    remarque = st.text_area(
                        "Remarque",
                        key=f"rem_{_key}",
                        height=60
                    )

                    _when_set = set(pg["when_list"])
                    _has_j7   = "J7" in _when_set

                    idc1, idc2, idc3, idc4 = st.columns([2, 1.5, 1.5, 0.6])

                    with idc1:
                        if st.button("🔍 Analyser & Enregistrer", use_container_width=True,
                                     key=f"submit_id_{_key}"):

                            # ✅ CORRECTION Bug 2 : real_indices calculé ici
                            real_indices = [
                                i for i, p in enumerate(st.session_state.pending_identifications)
                                if p["sample_id"] == _sid
                            ]

                            valid_entries = [
                                g for g in st.session_state[germs_list_key]
                                if g["germ"] and g["germ"] != "— Sélectionner un germe —"
                            ]

                            if not valid_entries:
                                st.error("Veuillez sélectionner au moins un germe.")
                            else:
                                scored_entries = []
                                for ve in valid_entries:
                                    match, score_fuzzy = find_germ_match(ve["germ"],
                                                                         st.session_state.germs)
                                    if match and score_fuzzy > 0.4:
                                        gs = _get_germ_score(match)
                                        scored_entries.append({
                                            "germ_saisi":  ve["germ"],
                                            "germ_match":  match["name"],
                                            "match_score": f"{int(score_fuzzy * 100)}%",
                                            "ufc":         ve["ufc"],
                                            "germ_score":  gs,
                                            "match_obj":   match,
                                        })

                                if not scored_entries:
                                    st.warning("⚠️ Aucune correspondance trouvée.")
                                else:
                                    worst_entry  = max(scored_entries, key=lambda x: x["germ_score"])
                                    total_sc     = loc_crit * worst_entry["germ_score"]
                                    status, status_lbl, status_col = _evaluate_score(total_sc)
                                    ufc_total    = sum(e["ufc"] for e in scored_entries)
                                    triggered_by = (
                                        f"lieu {loc_crit} × germe {worst_entry['germ_score']}"
                                        f" ({worst_entry['germ_match']})"
                                        if status in ("alert", "action") else None
                                    )
                                    germs_detail = [
                                        {
                                            "name":        e["germ_match"],
                                            "germ_saisi":  e["germ_saisi"],
                                            "match_score": e["match_score"],
                                            "ufc":         e["ufc"],
                                            "germ_score":  e["germ_score"],
                                            "is_worst":    e["germ_match"] == worst_entry["germ_match"],
                                        }
                                        for e in scored_entries
                                    ]

                                    st.session_state.surveillance.append({
                                        "date":                 str(_date_prelev),   # ✅ CORRECTION Bug 1
                                        "date_prelevement":     str(_date_prelev),
                                        "prelevement":          _label,
                                        "sample_id":            _sid,
                                        "germ_saisi":           worst_entry["germ_saisi"],
                                        "germ_match":           worst_entry["germ_match"],
                                        "match_score":          worst_entry["match_score"],
                                        "ufc":                  worst_entry["ufc"],
                                        "ufc_total":            ufc_total,
                                        "germ_score":           worst_entry["germ_score"],
                                        "germs_detail":         germs_detail,
                                        "multi_germ":           len(scored_entries) > 1,
                                        "location_criticality": loc_crit,
                                        "total_score":          total_sc,
                                        "risk":                 worst_entry["match_obj"].get("risk", worst_entry["germ_score"]),
                                        "room_class":           pt_class,
                                        "alert_threshold":      "Score ≥ 24",
                                        "action_threshold":     "Score > 36",
                                        "triggered_by":         triggered_by,
                                        "status":               status,
                                        "operateur":            pt_oper,
                                        "remarque":             remarque,
                                        "readings":             _when_str,
                                    })
                                    save_surveillance(st.session_state.surveillance)

                                    for _ri in real_indices:
                                        st.session_state.pending_identifications[_ri]["status"] = "done"
                                    save_pending_identifications(st.session_state.pending_identifications)

                                    if smp and not smp.get("archived"):
                                        smp["archived"] = True
                                        st.session_state.archived_samples.append(smp)
                                        save_archived_samples(st.session_state.archived_samples)
                                        save_prelevements(st.session_state.prelevements)

                                    st.session_state.pop(germs_list_key, None)

                                    if status in ("alert", "action"):
                                        st.session_state["_show_mesures_popup"] = {
                                            "status":               status,
                                            "germ":                 worst_entry["germ_match"],
                                            "germ_saisi":           worst_entry["germ_saisi"],
                                            "germ_match":           worst_entry["germ_match"],
                                            "ufc":                  worst_entry["ufc"],
                                            "risk":                 worst_entry["match_obj"].get("risk", worst_entry["germ_score"]),
                                            "label":                _label,
                                            "room_class":           pt_class,
                                            "triggered_by":         triggered_by,
                                            "germ_score":           worst_entry["germ_score"],
                                            "loc_criticality":      loc_crit,
                                            "location_criticality": loc_crit,
                                            "total_score":          total_sc,
                                            "germs_detail":         germs_detail,
                                            "date":                 str(_date_prelev),   # ✅ CORRECTION Bug 1
                                            "sample_id":            _sid,
                                        }
                                    else:
                                        germs_summary = ", ".join(
                                            f"{e['name']} ({e['ufc']} UFC)"
                                            for e in germs_detail)
                                        st.success(
                                            f"✅ {germs_summary} — **Conforme** (score {total_sc})")
                                    st.rerun()

                    with idc2:
                        _back_lbl = (
                            "↩️ Corriger J7" if _when_set == {"J7"}
                            else "↩️ Corriger J2" if _when_set == {"J2"}
                            else "↩️ Corriger lecture"
                        )
                        if st.button(_back_lbl, use_container_width=True,
                                     key=f"cancel_id_{_key}"):

                            # ✅ CORRECTION Bug 3 : _entries remplacé par pg["entries"]
                            for _e in pg["entries"]:
                                sch = next((x for x in st.session_state.schedules
                                            if x["sample_id"] == _sid
                                            and x["when"] == _e["when"]
                                            and x["status"] == "done"), None)
                                if sch:
                                    sch["status"] = "pending"
                            save_schedules(st.session_state.schedules)

                            real_indices = [
                                i for i, p in enumerate(st.session_state.pending_identifications)
                                if p["sample_id"] == _sid
                            ]
                            for _ri in sorted(real_indices, reverse=True):
                                st.session_state.pending_identifications.pop(_ri)
                            save_pending_identifications(st.session_state.pending_identifications)
                            if germs_list_key in st.session_state:
                                del st.session_state[germs_list_key]
                            st.rerun()

                    with idc3:
                        if _has_j7:
                            if st.button("↩️ Revenir à J2", use_container_width=True,
                                         key=f"back_j2_id_{_key}"):
                                _j2_back = next((x for x in st.session_state.schedules
                                                 if x["sample_id"] == _sid and x["when"] == "J2"), None)
                                if _j2_back:
                                    _j2_back["status"] = "pending"
                                _j7_back = next((x for x in st.session_state.schedules
                                                 if x["sample_id"] == _sid and x["when"] == "J7"), None)
                                if _j7_back:
                                    _j7_back["status"] = "pending"
                                save_schedules(st.session_state.schedules)
                                st.session_state.pending_identifications = [
                                    x for x in st.session_state.pending_identifications
                                    if x.get("sample_id") != _sid
                                ]
                                save_pending_identifications(st.session_state.pending_identifications)
                                if germs_list_key in st.session_state:
                                    del st.session_state[germs_list_key]
                                st.success("↩️ J2 et J7 remises en attente.")
                                st.rerun()

                    with idc4:
                        if st.button("🗑️", use_container_width=True, key=f"del_id_{_key}"):
                            real_indices = [
                                i for i, p in enumerate(st.session_state.pending_identifications)
                                if p["sample_id"] == _sid
                            ]
                            for _ri in sorted(real_indices, reverse=True):
                                st.session_state.pending_identifications.pop(_ri)
                            save_pending_identifications(st.session_state.pending_identifications)
                            if germs_list_key in st.session_state:
                                del st.session_state[germs_list_key]
                            st.rerun()

# ═══════════════════════════════════════════════════════════════════════════════
# TAB : PLANNING — Charge hebdo & Planning mensuel | Export Excel | Étiquettes
# Algorithme : max prélèvements/classe/semaine → répartition mensuelle
#              jamais 2× le même point le même jour (sauf freq > 1/jour)
# Auto-skip   : tout prélèvement non réalisé avant vendredi 23h59 de sa semaine
#               est automatiquement mis en « non-fait » et redistribué.
# ═══════════════════════════════════════════════════════════════════════════════
if active == "planning":
    st.markdown("### 📅 Planning des prélèvements & lectures")

    _today_dt     = datetime.today().date()
    MOIS_FR       = ["","Janvier","Février","Mars","Avril","Mai","Juin",
                     "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]
    JOURS_FR_LONG = ["Lundi","Mardi","Mercredi","Jeudi","Vendredi","Samedi","Dimanche"]

    # ── Initialisation planning_skips ─────────────────────────────────────────
    if "planning_skips" not in st.session_state:
        _raw_skips = _supa_get('planning_skips')
        st.session_state["planning_skips"] = json.loads(_raw_skips) if _raw_skips else {}

    # ── Initialisation planning_overrides ─────────────────────────────────────
    if "planning_overrides_loaded" not in st.session_state:
        for k, v in st.session_state.get("planning_overrides", {}).items():
            if k not in st.session_state:
                try:
                    st.session_state[k] = int(v)
                except Exception:
                    pass
        st.session_state["planning_overrides_loaded"] = True

    # ── Helpers fréquence ──────────────────────────────────────────────────────
    def _frc_default(rc):
        rc = (rc or '').strip().upper()
        if 'A' in rc: return 20
        if 'D' in rc: return 10
        return 2

    def _freq_en_semaine(pt, nb_jours_ouvres=5):
        fr = pt.get('frequency')
        u  = pt.get('frequency_unit', '/ semaine')
        try:
            fr = float(fr)
        except Exception:
            fr = 0.0
        if fr <= 0:
            return float(_frc_default((pt.get('room_class') or '').strip()))
        if '/ jour'  in u: return fr * nb_jours_ouvres
        if '/ mois'  in u: return fr / 4.33
        return fr

    def _semaines_du_mois(year, month):
        import calendar as _c
        _, n_days = _c.monthrange(year, month)
        first = date_type(year, month, 1)
        last  = date_type(year, month, n_days)
        cur   = first - timedelta(days=first.weekday())
        ms    = []
        while cur <= last:
            ms.append(cur)
            cur += timedelta(weeks=1)
        return ms

    def get_week_start(d):
        return d - timedelta(days=d.weekday())

    def fmt_week(ws):
        we = ws + timedelta(days=6)
        return ws.strftime('%d/%m') + ' – ' + we.strftime('%d/%m/%Y')

    def _make_qr_bytes(point_id) -> bytes:
        qr = qrcode.QRCode(
            version=None,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=4,
            border=2,
        )
        qr.add_data(str(point_id))
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buf = BytesIO()
        img.save(buf, format="PNG")
        return buf.getvalue()

    # ════════════════════════════════════════════════════════════════════════
    # AUTO-SKIP : vendredi 23h59 passé → non-fait automatique + Supabase
    # ════════════════════════════════════════════════════════════════════════
    def _auto_skip_past_weeks(raw_plan, planning_skips, prelevements, today):
        from datetime import datetime as _dt, time as _time
        nb_new = 0
        _now = _dt.now()

        for day in sorted(raw_plan.keys()):
            friday_of_week = day + timedelta(days=(4 - day.weekday()))
            # ✅ Cutoff : vendredi à 16h00 (et non plus 23h59)
            friday_cutoff = _dt.combine(friday_of_week, _time(16, 0))
            if _now < friday_cutoff:
                continue  # Semaine encore en cours ou future → on ne touche pas

            tasks = raw_plan.get(day, [])
            if not tasks:
                continue

            done_labels = {
                p["label"]
                for p in prelevements
                if p.get("date")
                and not p.get("archived", False)
                and datetime.fromisoformat(p["date"]).date() == day
            }
            dk             = day.isoformat()
            existing_skips = set(planning_skips.get(dk, []))

            for task in tasks:
                lbl = task["label"]
                if lbl not in done_labels and lbl not in existing_skips:
                    planning_skips.setdefault(dk, [])
                    planning_skips[dk].append(lbl)
                    nb_new += 1

        return planning_skips, nb_new

    def _compute_monthly_planning(year, month, holidays_set, prelevements=None):
        import calendar as _cm
        import random as _rnd

        _, n_days = _cm.monthrange(year, month)
        first = date_type(year, month, 1)
        last  = date_type(year, month, n_days)
        cur   = first - timedelta(days=first.weekday())
        mondays = []
        while cur <= last:
            mondays.append(cur)
            cur += timedelta(weeks=1)

        all_wd = [
            first + timedelta(days=i)
            for i in range(n_days)
            if (first + timedelta(days=i)).weekday() < 5
            and (first + timedelta(days=i)) not in holidays_set
        ]

        planning = {d: [] for d in all_wd}

        if prelevements is None:
            prelevements = []

        # ── Occurrences déjà réalisées ce mois ───────────────────────────
        _done_this_month = {}
        for _p in prelevements:
            if _p.get("archived", False) or not _p.get("date"):
                continue
            try:
                _d = datetime.fromisoformat(_p["date"]).date()
            except Exception:
                continue
            if _d.year != year or _d.month != month:
                continue
            _lbl = _p.get("label", "")
            _done_this_month.setdefault(_lbl, {})
            _done_this_month[_lbl][_d] = _done_this_month[_lbl].get(_d, 0) + 1

        for pt in st.session_state.points:
            rc        = (pt.get('room_class') or '').strip()
            freq_raw  = pt.get('frequency')
            freq_unit = pt.get('frequency_unit', '/ semaine')
            try:
                freq_val = float(freq_raw) if freq_raw else 0.0
            except (TypeError, ValueError):
                freq_val = 0.0

            nb_wd = len(all_wd)

            if freq_val <= 0:
                default_by_class  = {'A': 20, 'B': 10, 'C': 4, 'D': 2}
                total_occurrences = default_by_class.get(rc[:1] if rc else '', 2)
                max_per_day       = 1
                is_daily          = False
            elif '/ jour' in freq_unit:
                total_occurrences = int(freq_val) * nb_wd
                max_per_day       = int(freq_val)
                is_daily          = True
            elif '/ semaine' in freq_unit:
                nb_semaines       = len(mondays)
                total_occurrences = int(freq_val) * nb_semaines
                max_per_day       = 1
                is_daily          = False
            elif 'mois' in (freq_unit or '').lower():
                total_occurrences = int(freq_val)
                max_per_day       = 1
                is_daily          = False
            else:
                total_occurrences = int(freq_val)
                max_per_day       = 1
                is_daily          = False

            _already_done_count = sum(_done_this_month.get(pt['label'], {}).values())
            total_occurrences   = max(0, total_occurrences - _already_done_count)

            if total_occurrences <= 0:
                continue

            task_base = {
                'label':       pt['label'],
                'type':        pt.get('type', '—'),
                'risk':        int(pt.get('risk_level', 1)),
                'room_class':  rc,
                'max_per_day': max(1, max_per_day),
                '_freq_unit':  freq_unit,
            }

            _done_dates_for_pt = set(_done_this_month.get(pt['label'], {}).keys())

            if is_daily:
                for d in all_wd:
                    if d in _done_dates_for_pt:
                        continue
                    for _ in range(max_per_day):
                        planning[d].append(dict(task_base))
            else:
                rng          = _rnd.Random(year * 10000 + month * 100 + hash(pt['label']) % 100)
                available_wd = [d for d in all_wd if d not in _done_dates_for_pt]
                day_counts   = {d: 0 for d in available_wd}
                day_labels   = {d: {} for d in available_wd}
                remaining    = total_occurrences
                max_attempts = remaining * max(len(available_wd), 1) * 4 + 50
                attempts     = 0
                while remaining > 0 and attempts < max_attempts:
                    attempts += 1
                    candidates = [
                        d for d in available_wd
                        if day_labels[d].get(pt['label'], 0) < max_per_day
                    ]
                    if not candidates:
                        break
                    best = min(candidates, key=lambda d: (day_counts[d], rng.random()))
                    planning[best].append(dict(task_base))
                    day_counts[best] += 1
                    day_labels[best][pt['label']] = day_labels[best].get(pt['label'], 0) + 1
                    remaining -= 1

        for d in planning:
            planning[d].sort(key=lambda x: (-x['risk'], x['label']))

        return planning

    def _redistribute_skips(monthly_plan, planning_skips, holidays_set):
        import copy
        from datetime import datetime as _dt, time as _time
        plan            = copy.deepcopy(monthly_plan)
        today           = _today_dt
        _now            = _dt.now()
        all_days_sorted = sorted(plan.keys())
        future_days     = [d for d in all_days_sorted if d > today]

        # ── Gel semaine +1 dès vendredi 16h00 ────────────────────────────────
        # Calcul du vendredi de la semaine EN COURS
        _cur_monday  = today - timedelta(days=today.weekday())
        _cur_friday  = _cur_monday + timedelta(days=4)
        _freeze_cutoff = _dt.combine(_cur_friday, _time(16, 0))

        if _now >= _freeze_cutoff:
            # La semaine +1 est gelée : lundi et vendredi de la semaine prochaine
            _next_monday = _cur_monday + timedelta(weeks=1)
            _next_friday = _next_monday + timedelta(days=4)
            # On exclut tous les jours de la semaine +1 des cibles de redistribution
            future_days = [
                d for d in future_days
                if not (_next_monday <= d <= _next_friday)
            ]

        for day in all_days_sorted:
            skipped_labels = planning_skips.get(day.isoformat(), [])
            if not skipped_labels:
                continue

            tasks_to_move = [
                t for t in plan.get(day, [])
                if t["label"] in skipped_labels
                and "/ jour" not in t.get("_freq_unit", "")
            ]
            plan[day] = [
                t for t in plan.get(day, [])
                if t["label"] not in skipped_labels
            ]

            for task in tasks_to_move:
                candidates = [
                    d for d in future_days
                    if not any(t["label"] == task["label"] for t in plan.get(d, []))
                ]
                if not candidates:
                    candidates = future_days
                if not candidates:
                    continue
                best = min(candidates, key=lambda d: len(plan.get(d, [])))
                plan.setdefault(best, []).append(task)

        return plan

    def _generate_pdf_etiquettes(tasks, date_obj_or_list):
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units     import cm as rl_cm
        from reportlab.lib           import colors as rlc
        from reportlab.platypus      import (
            Table, TableStyle, Paragraph, HRFlowable,
            BaseDocTemplate, Frame, PageTemplate, Image as RLImage,
            Spacer,
        )
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums  import TA_RIGHT
        from io import BytesIO

        A4_W, A4_H = A4
        N_COLS      = 4
        W_ETQ       = 5.2  * rl_cm
        H_ETQ       = 2.95 * rl_cm

        buf = BytesIO()

        RISK_RL = {k: rlc.HexColor(v) for k, v in {
            "1": "#22c55e", "2": "#84cc16",
            "3": "#f59e0b", "4": "#f97316", "5": "#ef4444",
        }.items()}

        s_titre   = ParagraphStyle("et_t",  fontName="Helvetica-Bold",
                                   fontSize=7.5, leading=9, spaceAfter=2,
                                   textColor=rlc.HexColor("#0f172a"))
        s_lbl     = ParagraphStyle("et_l",  fontName="Helvetica",
                                   fontSize=5.5, leading=7,
                                   textColor=rlc.HexColor("#64748b"))
        s_date    = ParagraphStyle("et_d",  fontName="Helvetica-Bold",
                                   fontSize=9, leading=10,
                                   textColor=rlc.HexColor("#1e40af"))
        s_logo    = ParagraphStyle("et_lo", fontName="Helvetica",
                                   fontSize=5, leading=6,
                                   textColor=rlc.HexColor("#94a3b8"),
                                   alignment=TA_RIGHT)
        s_classea = ParagraphStyle("et_ca", fontName="Helvetica-Bold",
                                   fontSize=6, leading=7,
                                   textColor=rlc.HexColor("#854d0e"),
                                   spaceAfter=2)
        s_val     = ParagraphStyle("et_v",  fontName="Helvetica-Bold",
                                   fontSize=7.5, leading=9,
                                   textColor=rlc.HexColor("#0f172a"))
        s_day_sep = ParagraphStyle("et_ds", fontName="Helvetica-Bold",
                                   fontSize=11, leading=14,
                                   textColor=rlc.HexColor("#1a4e66"))

        if isinstance(date_obj_or_list, list):
            days_data = date_obj_or_list
        else:
            days_data = [(date_obj_or_list, tasks)]

        doc = BaseDocTemplate(
            buf, pagesize=A4,
            leftMargin=0, rightMargin=0, topMargin=0, bottomMargin=0,
        )
        frame = Frame(
            x1=0, y1=0, width=A4_W, height=A4_H,
            leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0,
        )
        doc.addPageTemplates([PageTemplate(id="full", frames=[frame])])

        def _build_cell(task, d_obj):
            rv      = str(task.get("risk", ""))
            rc_etiq = RISK_RL.get(rv, rlc.HexColor("#6366f1"))
            lv      = task.get("label", "—")
            dv      = d_obj.strftime("%d/%m/%Y")
            W_QR    = 2.0 * rl_cm
            W_INNER = W_ETQ - 0.55 * rl_cm
            W_TEXT  = W_INNER - W_QR

            _pt_data = next(
                (p for p in st.session_state.points if p.get("label") == lv), None
            )
            qr_flowable = None
            if _pt_data:
                try:
                    _qr_buf = BytesIO(_make_qr_bytes(_pt_data["id"]))
                    qr_flowable = RLImage(_qr_buf, width=2.0 * rl_cm, height=2.0 * rl_cm)
                except Exception:
                    qr_flowable = None

            classea_rows = []
            if task.get("room_class", "").strip().upper() == "A":
                iso_display = task.get("_isolateur")
                if not iso_display and _pt_data:
                    iso_display = _pt_data.get("num_isolateur", "—") or "—"
                if not iso_display:
                    iso_display = "—"
                pst       = (_pt_data.get("poste", "") or "") if _pt_data else ""
                label_iso = iso_display + (f" · {pst}" if pst else "")
                classea_rows = [[Paragraph(label_iso, s_classea)]]

            try:
                left_tbl = Table(
                    [
                        [Paragraph(lv, s_titre)],
                        [HRFlowable(width=W_TEXT, thickness=0.6, color=rc_etiq, spaceAfter=2)],
                        *classea_rows,
                        [Paragraph("📅 Date", s_lbl)],
                        [Paragraph(dv, s_date)],
                        [Paragraph("👤 Préleveur :", s_lbl)],
                        [Paragraph("", s_val)],
                        [Paragraph("URC — MicroSurveillance", s_logo)],
                    ],
                    colWidths=[W_TEXT],
                )
                left_tbl.setStyle(TableStyle([
                    ("LEFTPADDING",   (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING",  (0, 0), (-1, -1), 0),
                    ("TOPPADDING",    (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
                    ("TOPPADDING",    (0, -1), (0, -1), 4),
                ]))
            except Exception:
                left_tbl = Table([[Paragraph(lv, s_titre)]], colWidths=[W_TEXT])

            try:
                if qr_flowable:
                    inner = Table([[left_tbl, qr_flowable]], colWidths=[W_TEXT, W_QR])
                    inner.setStyle(TableStyle([
                        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
                        ("LEFTPADDING",   (0, 0), (-1, -1), 0),
                        ("RIGHTPADDING",  (0, 0), (-1, -1), 0),
                        ("TOPPADDING",    (0, 0), (-1, -1), 0),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                        ("ALIGN",         (1, 0), (1, 0),   "CENTER"),
                    ]))
                else:
                    inner = left_tbl
            except Exception:
                inner = left_tbl

            outer = Table([[inner]], colWidths=[W_ETQ], rowHeights=[H_ETQ])
            outer.setStyle(TableStyle([
                ("LINEBEFORE",    (0, 0), (0, 0), 1.2, rc_etiq),
                ("LINEABOVE",     (0, 0), (0, 0), 1.2, rc_etiq),
                ("LINEAFTER",     (0, 0), (0, 0), 5.5, rc_etiq),
                ("LEFTPADDING",   (0, 0), (0, 0), 11),
                ("RIGHTPADDING",  (0, 0), (0, 0), 11),
                ("TOPPADDING",    (0, 0), (0, 0), 11),
                ("BOTTOMPADDING", (0, 0), (0, 0), 11),
                ("VALIGN",        (0, 0), (0, 0), "TOP"),
                ("BACKGROUND",    (0, 0), (0, 0), rlc.white),
            ]))
            return outer

        def _build_sep_cell(label):
            sep_inner = Table(
                [[Paragraph(label, s_day_sep)]],
                colWidths=[W_ETQ - 0.55 * rl_cm],
            )
            sep_inner.setStyle(TableStyle([
                ("LEFTPADDING",   (0, 0), (-1, -1), 0),
                ("RIGHTPADDING",  (0, 0), (-1, -1), 0),
                ("TOPPADDING",    (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ]))
            outer = Table([[sep_inner]], colWidths=[W_ETQ], rowHeights=[H_ETQ])
            outer.setStyle(TableStyle([
                ("LINEBEFORE",    (0, 0), (0, 0), 1.2, rlc.HexColor("#1a4e66")),
                ("LINEABOVE",     (0, 0), (0, 0), 1.2, rlc.HexColor("#1a4e66")),
                ("LINEAFTER",     (0, 0), (0, 0), 5.5, rlc.HexColor("#1a4e66")),
                ("LEFTPADDING",   (0, 0), (0, 0), 11),
                ("RIGHTPADDING",  (0, 0), (0, 0), 11),
                ("TOPPADDING",    (0, 0), (0, 0), 11),
                ("BOTTOMPADDING", (0, 0), (0, 0), 11),
                ("VALIGN",        (0, 0), (0, 0), "MIDDLE"),
                ("BACKGROUND",    (0, 0), (0, 0), rlc.HexColor("#e0f2fe")),
            ]))
            return outer

        story           = []
        all_rows        = []
        all_row_heights = []

        for (d_obj, day_tasks) in days_data:
            n_prelevements = len(day_tasks)
            sep_label      = (
                f"{d_obj.strftime('%A %d/%m/%Y').capitalize()} "
                f"— {n_prelevements} prélèvement{'s' if n_prelevements > 1 else ''}"
            )
            cells_day = [_build_sep_cell(sep_label)] + [_build_cell(t, d_obj) for t in day_tasks]
            for i in range(0, len(cells_day), N_COLS):
                chunk = cells_day[i:i + N_COLS]
                while len(chunk) < N_COLS:
                    chunk.append("")
                all_rows.append(chunk)
                all_row_heights.append(H_ETQ)

        if all_rows:
            full_tbl = Table(
                all_rows,
                colWidths=[W_ETQ] * N_COLS,
                rowHeights=all_row_heights,
            )
            full_tbl.setStyle(TableStyle([
                ("LEFTPADDING",   (0, 0), (-1, -1), 0),
                ("RIGHTPADDING",  (0, 0), (-1, -1), 0),
                ("TOPPADDING",    (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ("LINEBELOW",     (0, -1), (-1, -1), 1.2, rlc.HexColor("#94a3b8")),
            ]))
            story.append(full_tbl)

        doc.build(story)
        buf.seek(0)
        return buf.getvalue()

    # ════════════════════════════════════════════════════════════════════════
    # ONGLETS — déclaration unique
    # ════════════════════════════════════════════════════════════════════════
    plan_tab_charge, plan_tab_export = st.tabs([
        "📊 Charge hebdo & Planning mensuel",
        "📥 Export Excel",
    ])

    # ════════════════════════════════════════════════════════════════════════
    # ONGLET 1 — Charge hebdo & Planning mensuel
    # ════════════════════════════════════════════════════════════════════════
    with plan_tab_charge:
        if not st.session_state.get("points"):
            _raw_points = _supa_get("points")
            if _raw_points:
                try:
                    st.session_state.points = json.loads(_raw_points)
                except Exception:
                    st.session_state.points = []

        # ── Sélecteurs Année / Mois ───────────────────────────────────────
        col_y, col_m = st.columns(2)
        with col_y:
            _ch_year = st.number_input(
                "Année", min_value=2020, max_value=2030,
                value=st.session_state.cal_year, step=1,
                key="cal_year_sel",
            )
            st.session_state.cal_year = _ch_year
        with col_m:
            _ch_month = st.selectbox(
                "Mois", range(1, 13),
                format_func=lambda m: MOIS_FR[m],
                index=st.session_state.cal_month - 1,
                key="cal_month_sel",
            )
            st.session_state.cal_month = _ch_month

        _ch_holidays = get_holidays_cached(_ch_year)

        # ── Calcul des lundis du mois ─────────────────────────────────────
        import calendar as _cal_pm
        _, _pm_ndays = _cal_pm.monthrange(_ch_year, _ch_month)
        _pm_start    = date_type(_ch_year, _ch_month, 1)
        _pm_end      = date_type(_ch_year, _ch_month, _pm_ndays)
        cur_pm       = _pm_start - timedelta(days=_pm_start.weekday())
        pm_mondays   = []
        while cur_pm <= _pm_end:
            pm_mondays.append(cur_pm)
            cur_pm += timedelta(weeks=1)

        # ── Normalisation des clés des points ─────────────────────────────
        for _pt in st.session_state.points:
            if not _pt.get("room_class"):
                _pt["room_class"] = _pt.get("class", "")
            if "frequency" not in _pt or _pt.get("frequency") is None:
                _pt["frequency"] = _pt.get("freq", None)
            if "frequency_unit" not in _pt or not _pt.get("frequency_unit"):
                _pt["frequency_unit"] = _pt.get("unit", "/ semaine")

        st.divider()

        # ── Planning mensuel ──────────────────────────────────────────────
        st.markdown("#### 📅 Planning mensuel automatique")
        st.caption(
            "Répartition basée sur la fréquence de chaque point. "
            "Un point n'apparaît jamais 2× le même jour sauf fréquence ≥ 1/jour. "
            "Les tâches non réalisées avant vendredi 23h59 sont automatiquement "
            "marquées non-faites et redistribuées sur les jours ouvrés futurs du mois."
        )

        # ── Calcul du planning brut ───────────────────────────────────────
        monthly_plan_raw = _compute_monthly_planning(
            _ch_year, _ch_month, _ch_holidays,
            prelevements=st.session_state.get("prelevements", []),
        )
        monthly_plan_raw = {
            (k.date() if hasattr(k, 'date') else k): v
            for k, v in monthly_plan_raw.items()
        }

        # ── AUTO-SKIP : semaines entièrement passées (vendredi < aujourd'hui) ──
        _skips_before = sum(len(v) for v in st.session_state["planning_skips"].values())
        st.session_state["planning_skips"], _nb_auto = _auto_skip_past_weeks(
            monthly_plan_raw,
            st.session_state["planning_skips"],
            st.session_state.get("prelevements", []),
            _today_dt,
        )
        if _nb_auto > 0:
            # Sauvegarde Supabase uniquement si de nouveaux skips ont été ajoutés
            _supa_upsert('planning_skips', json.dumps(st.session_state["planning_skips"]))
            st.info(
                f"🔄 **{_nb_auto} prélèvement(s)** non réalisé(s) sur des semaines "
                f"passées ont été automatiquement marqués non-faits et redistribués.",
                icon="⏭️",
            )

        # ── Redistribution des skips (manuels + auto) ────────────────────
        monthly_plan = _redistribute_skips(
            monthly_plan_raw,
            st.session_state["planning_skips"],
            _ch_holidays,
        )

        if "pm_selected_day" not in st.session_state:
            st.session_state["pm_selected_day"] = None

        # ── Rendu calendrier semaine par semaine ──────────────────────────
        for week_monday in pm_mondays:
            wd_week = [
                week_monday + timedelta(days=i)
                for i in range(5)
                if (week_monday + timedelta(days=i)) not in _ch_holidays
                and _pm_start <= (week_monday + timedelta(days=i)) <= _pm_end
            ]
            if not wd_week:
                continue

            _total_sem    = sum(len(monthly_plan.get(wd, [])) for wd in wd_week)
            _friday_sem   = week_monday + timedelta(days=4)
            _week_passed  = _friday_sem < _today_dt
            _week_current = week_monday <= _today_dt <= _friday_sem

            # ── En-tête semaine avec indicateur de statut ─────────────────
            c1, c2 = st.columns([4, 2])
            with c1:
                if _week_passed:
                    _sem_badge = "<span style='color:#94a3b8;font-size:.72rem'>✅ Semaine passée</span>"
                elif _week_current:
                    _sem_badge = "<span style='color:#22c55e;font-size:.72rem'>▶ Semaine en cours</span>"
                else:
                    _sem_badge = ""
                st.markdown(
                    f"<div style='background:#1e293b;padding:10px;border-radius:10px;"
                    f"display:flex;justify-content:space-between;align-items:center'>"
                    f"<div><b style='color:white'>Semaine {week_monday.isocalendar()[1]}</b>"
                    f"&nbsp;&nbsp;{_sem_badge}</div>"
                    f"<span style='color:#93c5fd'>{_total_sem} prélèv.</span>"
                    f"</div>",
                    unsafe_allow_html=True,
                )
            with c2:
                if st.button("🖨️ Imprimer", key=f"print_{week_monday}"):
                    week_days = [(wd, monthly_plan.get(wd, [])) for wd in wd_week]

                    def filt(tasks, cls, iso=None):
                        out = []
                        for t in tasks:
                            if (t.get("room_class") or "").upper() == cls:
                                t2 = dict(t)
                                if iso:
                                    t2["_isolateur"] = iso
                                out.append(t2)
                        return out

                    days_D  = [(d, filt(t, "D"))                 for d, t in week_days if filt(t, "D")]
                    days_14 = [(d, filt(t, "A", "Iso 14/07169")) for d, t in week_days if filt(t, "A")]
                    days_16 = [(d, filt(t, "A", "Iso 16/0724"))  for d, t in week_days if filt(t, "A")]

                    if days_D:
                        st.session_state[f"pdf_D_{week_monday}"]  = _generate_pdf_etiquettes(days_D, days_D)
                    if days_14:
                        st.session_state[f"pdf_14_{week_monday}"] = _generate_pdf_etiquettes(days_14, days_14)
                    if days_16:
                        st.session_state[f"pdf_16_{week_monday}"] = _generate_pdf_etiquettes(days_16, days_16)
                    st.rerun()

            for suffix, label in [("D", "Classe D"), ("14", "Iso 14"), ("16", "Iso 16")]:
                key_pdf = f"pdf_{suffix}_{week_monday}"
                if key_pdf in st.session_state:
                    st.download_button(
                        f"⬇️ {label}",
                        data=st.session_state[key_pdf],
                        file_name=f"etiquettes_{suffix}_{week_monday}.pdf",
                        mime="application/pdf",
                        key=f"dl_{suffix}_{week_monday}",
                    )

            # ── Grille des jours ──────────────────────────────────────────
            cols = st.columns(len(wd_week))
            for i, wd in enumerate(wd_week):
                tasks       = monthly_plan.get(wd, [])
                done_labels = {
                    p["label"]
                    for p in st.session_state.prelevements
                    if p.get("date")
                    and datetime.fromisoformat(p["date"]).date() == wd
                }
                skipped   = set(st.session_state["planning_skips"].get(wd.isoformat(), []))
                non_faits = [
                    t for t in tasks
                    if t["label"] not in done_labels
                    and t["label"] not in skipped
                ]

                # Tâches auto-skippées (semaine passée)
                auto_skipped_this_day = (
                    set(st.session_state["planning_skips"].get(wd.isoformat(), []))
                    - done_labels
                ) if _friday_sem < _today_dt else set()

                with cols[i]:
                    st.markdown(f"**{JOURS_FR_LONG[wd.weekday()][:3]} {wd.strftime('%d/%m')}**")

                    if st.button("🔍", key=f"detail_{wd}"):
                        st.session_state["pm_selected_day"] = wd
                        st.rerun()

                    for t in tasks:
                        _dn     = t["label"] in done_labels
                        _sk     = t["label"] in skipped
                        _auto   = t["label"] in auto_skipped_this_day
                        _ic     = "💨" if t.get("type") == "Air" else "🧴"
                        _col    = "#22c55e" if _dn else "#dc2626" if _auto else "#94a3b8" if _sk else "#1e40af"
                        _strike = "line-through" if (_sk or _auto) else "none"
                        _suffix = " ⏭️" if _auto else ""
                        st.markdown(
                            f"<div style='font-size:.72rem;color:{_col};"
                            f"text-decoration:{_strike};padding:2px 0'>"
                            f"{_ic} {t['label'][:25]}{_suffix}</div>",
                            unsafe_allow_html=True,
                        )

                    # Bouton "Non faits" uniquement pour les jours non passés (semaine en cours ou future)
                    if non_faits and not _week_passed:
                        with st.popover(f"⬜ {len(non_faits)} Non faits"):
                            selections = []
                            for ti, t in enumerate(non_faits):
                                key = (
                                    f"skip_{wd.isoformat()}_{ti}_"
                                    f"{t['label'][:20].replace(' ', '_').replace('/', '_')}"
                                )
                                checked = st.checkbox(t["label"], key=key)
                                if checked:
                                    selections.append(t["label"])

                            if st.button("💾 Enregistrer", key=f"save_skips_{wd}"):
                                _skips = st.session_state["planning_skips"]
                                dk     = wd.isoformat()
                                _skips.setdefault(dk, [])
                                for label in selections:
                                    if label not in _skips[dk]:
                                        _skips[dk].append(label)
                                st.session_state["planning_skips"] = _skips
                                _supa_upsert('planning_skips', json.dumps(_skips))
                                st.rerun()
                    elif _week_passed and auto_skipped_this_day:
                        # Affiche un résumé compact pour les semaines passées
                        st.markdown(
                            f"<div style='font-size:.65rem;color:#dc2626;margin-top:2px'>"
                            f"⏭️ {len(auto_skipped_this_day)} non-fait(s) auto</div>",
                            unsafe_allow_html=True,
                        )
                    elif not non_faits:
                        st.button("✅", disabled=True, key=f"done_{wd}")

            st.divider()

        # ── Panel détail jour ─────────────────────────────────────────────
        sel = st.session_state.get("pm_selected_day")

        if sel:
            st.markdown("---")
            _day_lbl = JOURS_FR_LONG[sel.weekday()] + " " + sel.strftime("%d/%m/%Y")
            st.markdown(f"## 📋 {_day_lbl}")

            if st.button("✖️ Fermer", key="close_detail"):
                st.session_state["pm_selected_day"] = None
                st.rerun()

            tasks = monthly_plan.get(sel, [])

            if not tasks:
                st.info("Aucun prélèvement planifié ce jour.")
            else:
                done_labels = {
                    p["label"]
                    for p in st.session_state.prelevements
                    if p.get("date")
                    and datetime.fromisoformat(p["date"]).date() == sel
                }
                skipped_labels = set(
                    st.session_state["planning_skips"].get(sel.isoformat(), [])
                )

                rcp_fix = {
                    "1": "#22c55e", "2": "#84cc16", "3": "#f59e0b",
                    "4": "#f97316", "5": "#ef4444",
                }

                st.markdown(
                    "<div style='background:#eff6ff;border:1px solid #bfdbfe;"
                    "border-radius:8px;padding:8px 14px;margin-bottom:8px'>"
                    "<span style='font-size:.88rem;font-weight:800;color:#1e40af'>"
                    "🏷️ Générer les étiquettes pour ce jour</span></div>",
                    unsafe_allow_html=True,
                )

                col_sa, col_sn, col_sk, _ = st.columns([1, 1, 1, 3])
                with col_sa:
                    if st.button("☑️ Tout sélectionner",
                                 key=f"etiq_all_{sel.isoformat()}",
                                 use_container_width=True):
                        for _ti, _t in enumerate(tasks):
                            st.session_state[f"etiq_chk_{sel.isoformat()}_{_ti}"] = True
                        st.rerun()
                with col_sn:
                    if st.button("⬜ Tout désélectionner",
                                 key=f"etiq_none_{sel.isoformat()}",
                                 use_container_width=True):
                        for _ti, _t in enumerate(tasks):
                            st.session_state[f"etiq_chk_{sel.isoformat()}_{_ti}"] = False
                        st.rerun()
                with col_sk:
                    if st.button("🚫 Tout reporter",
                                 key=f"all_skip_{sel.isoformat()}",
                                 use_container_width=True):
                        dk     = sel.isoformat()
                        _skips = st.session_state["planning_skips"]
                        _skips[dk] = list(set(
                            _skips.get(dk, []) + [t["label"] for t in tasks]
                        ))
                        st.session_state["planning_skips"] = _skips
                        _supa_upsert('planning_skips', json.dumps(_skips))
                        st.success("Tous les prélèvements reportés.")
                        st.rerun()

                _chk_cols       = st.columns(3)
                _selected_tasks = []

                for _ti, _t in enumerate(tasks):
                    _chk_key = f"etiq_chk_{sel.isoformat()}_{_ti}"
                    if _chk_key not in st.session_state:
                        st.session_state[_chk_key] = True
                    _rc = rcp_fix.get(str(_t.get("risk", "1")), "#94a3b8")
                    _ic = "💨" if _t.get("type") == "Air" else "🧴"
                    _dn = _t["label"] in done_labels
                    _sk = _t["label"] in skipped_labels
                    bg  = "#f0fdf4" if _dn else "#fff8f1" if _sk else "#fff"

                    with _chk_cols[_ti % 3]:
                        st.markdown(
                            f"<div style='background:{bg};"
                            f"border:1px solid {_rc}44;border-left:3px solid {_rc};"
                            f"border-radius:7px;padding:4px 8px;margin-bottom:4px'>"
                            f"<span style='font-size:.72rem;font-weight:700;color:#0f172a'>"
                            f"{_ic} {'✅ ' if _dn else '⏭️ ' if _sk else ''}"
                            f"{_t['label'][:28]}</span><br>"
                            f"<span style='font-size:.6rem;color:#64748b'>"
                            f"Cl.{_t.get('room_class') or '—'} · "
                            f"Nv.{_t.get('risk', '—')}</span></div>",
                            unsafe_allow_html=True,
                        )
                        if st.checkbox(
                            "Inclure",
                            value=st.session_state[_chk_key],
                            key=_chk_key,
                            label_visibility="collapsed",
                        ):
                            _selected_tasks.append(_t)

                _n_sel_etiq = len(_selected_tasks)
                st.markdown(
                    f"<div style='font-size:.8rem;color:#475569;margin:6px 0'>"
                    f"{_n_sel_etiq} / {len(tasks)} point(s) sélectionné(s)</div>",
                    unsafe_allow_html=True,
                )

                if _n_sel_etiq > 0:
                    if st.button(
                        f"📄 Générer {_n_sel_etiq} "
                        f"étiquette{'s' if _n_sel_etiq > 1 else ''} — {_day_lbl}",
                        use_container_width=True,
                        key=f"etiq_gen_{sel.isoformat()}",
                        type="primary",
                    ):
                        try:
                            _pdf_bytes = _generate_pdf_etiquettes(_selected_tasks, sel)
                            fname_etiq = f"etiquettes_{sel.strftime('%Y%m%d')}.pdf"
                            st.session_state[f"pdf_quick_{sel.isoformat()}"] = {
                                "data":  _pdf_bytes,
                                "fname": fname_etiq,
                            }
                            st.rerun()
                        except ImportError:
                            st.error("❌ **ReportLab** non installé.")
                        except Exception as _e:
                            st.error(f"Erreur génération PDF : {_e}")
                            import traceback
                            st.code(traceback.format_exc())

                    _pdf_cache = st.session_state.get(f"pdf_quick_{sel.isoformat()}")
                    if _pdf_cache:
                        st.download_button(
                            label=f"⬇️ Télécharger {_pdf_cache['fname']}",
                            data=_pdf_cache["data"],
                            file_name=_pdf_cache["fname"],
                            mime="application/pdf",
                            use_container_width=True,
                            key=f"etiq_dl_{sel.isoformat()}",
                        )
                        st.success(
                            f"✅ {_n_sel_etiq} étiquette"
                            f"{'s' if _n_sel_etiq > 1 else ''} prête"
                            f"{'s' if _n_sel_etiq > 1 else ''} — {_day_lbl}"
                        )

    # ════════════════════════════════════════════════════════════════════════
    # ONGLET 2 — Export Excel
    # ════════════════════════════════════════════════════════════════════════
    with plan_tab_export:
        st.markdown("#### 📥 Exporter le planning en Excel")

        try:
            import openpyxl
            _openpyxl_ok = True
        except ImportError:
            _openpyxl_ok = False

        if not _openpyxl_ok:
            st.error(
                "❌ **openpyxl** n'est pas installé.\n\n"
                "Ajoutez `openpyxl` dans votre fichier **requirements.txt** "
                "puis redémarrez l'application."
            )
            st.stop()

        exp_scope = st.selectbox(
            "Période",
            ["Semaine en cours", "Semaine choisie", "Mois en cours",
             "4 semaines à venir", "Tout le planning"],
            key="exp_scope",
        )

        if exp_scope == "Semaine choisie":
            _today_ref   = _today_dt
            _cur_monday  = _today_ref - timedelta(days=_today_ref.weekday())
            _week_opts   = [_cur_monday + timedelta(weeks=i) for i in range(-4, 13)]
            _week_labels = []
            for _wm in _week_opts:
                _we     = _wm + timedelta(days=4)
                _marker = " ← semaine en cours" if _wm == _cur_monday else ""
                _week_labels.append(
                    f"Sem. {_wm.isocalendar()[1]} — "
                    f"{_wm.strftime('%d/%m/%Y')} → {_we.strftime('%d/%m/%Y')}{_marker}"
                )
            _sel_week_idx = st.selectbox(
                "Choisir la semaine",
                range(len(_week_labels)),
                format_func=lambda i: _week_labels[i],
                index=4,
                key="exp_chosen_week",
            )
            _chosen_monday = _week_opts[_sel_week_idx]
        else:
            _chosen_monday = None

        only_working = st.checkbox("Inclure uniquement les jours ouvrés", value=True)

        # Option pour inclure les non-faits redistribués dans l'export
        include_nonfaits = st.checkbox(
            "Inclure les prélèvements redistribués (non-faits des semaines passées)",
            value=True,
            help="Affiche dans l'Excel les prélèvements reportés automatiquement "
                 "depuis des semaines passées non réalisées.",
        )

        if st.button("📊 Générer Excel", use_container_width=True, key="gen_xlsx"):
            import io as _io
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils  import get_column_letter

            wb = openpyxl.Workbook()

            # ── Palette couleurs ──────────────────────────────────────────
            C_BLUE     = "1E40AF"; C_BLUE2  = "2563EB"; C_BLUE_L = "DBEAFE"
            C_WHITE    = "FFFFFF"; C_TEXT   = "0F172A"; C_GREY_L  = "F1F5F9"
            C_GREY_HDR = "334155"; C_GREEN  = "16A34A"; C_RED    = "DC2626"

            thin   = Side(style="thin",   color="CBD5E1")
            medium = Side(style="medium", color="94A3B8")

            def fill(h):
                return PatternFill("solid", fgColor=h)

            def font(size=9, bold=False, color=C_TEXT):
                return Font(name="Arial", size=size, bold=bold, color=color)

            def al_c(wrap=False):
                return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

            def al_l(wrap=False):
                return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

            def border_all():
                return Border(left=thin, right=thin, top=thin, bottom=thin)

            def border_medium():
                return Border(left=medium, right=medium, top=medium, bottom=medium)

            # ── Plage de dates ────────────────────────────────────────────
            exp_today = _today_dt

            if exp_scope == "Semaine en cours":
                _monday   = exp_today - timedelta(days=exp_today.weekday())
                exp_dates = [_monday + timedelta(days=i) for i in range(7)]

            elif exp_scope == "Semaine choisie":
                exp_dates = [_chosen_monday + timedelta(days=i) for i in range(7)]

            elif exp_scope == "Mois en cours":
                import calendar as cal_module
                first     = exp_today.replace(day=1)
                last      = exp_today.replace(
                    day=cal_module.monthrange(exp_today.year, exp_today.month)[1])
                exp_dates = [
                    first + timedelta(days=i)
                    for i in range((last - first).days + 1)
                ]

            elif exp_scope == "4 semaines à venir":
                ws_e      = exp_today - timedelta(days=exp_today.weekday())
                exp_dates = [ws_e + timedelta(days=i) for i in range(28)]

            else:
                all_d = []
                for p in st.session_state.prelevements:
                    try:
                        all_d.append(datetime.fromisoformat(p["date"]).date())
                    except Exception:
                        pass
                for s in st.session_state.schedules:
                    try:
                        all_d.append(datetime.fromisoformat(s["due_date"]).date())
                    except Exception:
                        pass
                exp_dates = (
                    [min(all_d) + timedelta(days=i)
                     for i in range((max(all_d) - min(all_d)).days + 1)]
                    if all_d
                    else [exp_today + timedelta(days=i) for i in range(7)]
                )

            if only_working:
                exp_dates = [d for d in exp_dates if is_working_day(d)]

            # ── Plan Excel : on utilise monthly_plan déjà redistribué ─────
            # Si la période exporte un autre mois, recalculer
            _xl_months = set((d.year, d.month) for d in exp_dates)
            if len(_xl_months) == 1 and list(_xl_months)[0] == (_ch_year, _ch_month):
                xl_plan = monthly_plan  # déjà calculé + redistribué
            else:
                xl_plan = {}
                for (yr, mo) in _xl_months:
                    _hol = get_holidays_cached(yr)
                    _raw = _compute_monthly_planning(
                        yr, mo, _hol,
                        prelevements=st.session_state.get("prelevements", []),
                    )
                    _raw = {(k.date() if hasattr(k, 'date') else k): v for k, v in _raw.items()}
                    # Auto-skip + redistribution pour ce mois
                    _sk_tmp, _ = _auto_skip_past_weeks(
                        _raw, dict(st.session_state["planning_skips"]),
                        st.session_state.get("prelevements", []), exp_today,
                    )
                    _raw = _redistribute_skips(_raw, _sk_tmp, _hol)
                    xl_plan.update(_raw)

            # ── Feuille 1 — Vue matricielle ───────────────────────────────
            ws_matrix = wb.active
            ws_matrix.title = "Planning Semaine"
            ws_matrix.sheet_view.showGridLines = False

            from collections import OrderedDict
            weeks_map = OrderedDict()
            for d in exp_dates:
                ws_key = d - timedelta(days=d.weekday())
                if ws_key not in weeks_map:
                    weeks_map[ws_key] = []
                weeks_map[ws_key].append(d)

            JOURS_XL      = ["Lundi","Mardi","Mercredi","Jeudi","Vendredi","Samedi","Dimanche"]
            pts_all       = st.session_state.points
            n_weeks       = len(weeks_map)
            FIXED_COLS    = 3
            DAYS_PER_WEEK = 5
            total_cols    = FIXED_COLS + DAYS_PER_WEEK * n_weeks

            # Ligne 1 — titre
            ws_matrix.merge_cells(start_row=1, start_column=1,
                                   end_row=1,   end_column=total_cols)
            ws_matrix.cell(1, 1).value     = "PLANNING MICROBIOLOGIQUE — MicroSurveillance URC"
            ws_matrix.cell(1, 1).font      = Font(name="Arial", size=13, bold=True, color=C_WHITE)
            ws_matrix.cell(1, 1).fill      = fill(C_BLUE)
            ws_matrix.cell(1, 1).alignment = al_c()
            ws_matrix.row_dimensions[1].height = 28

            # Ligne 2 — légende
            ws_matrix.merge_cells(start_row=2, start_column=1,
                                   end_row=2,   end_column=total_cols)
            ws_matrix.cell(2, 1).value = (
                f"Généré le {exp_today.strftime('%d/%m/%Y')} — "
                f"{'Jours ouvrés uniquement' if only_working else 'Tous les jours'} — "
                f"X = prélèvement prévu · ⏭ = reporté (non-fait) · ✓ = réalisé"
            )
            ws_matrix.cell(2, 1).font      = Font(name="Arial", size=8, color="475569")
            ws_matrix.cell(2, 1).fill      = fill(C_BLUE_L)
            ws_matrix.cell(2, 1).alignment = al_c()
            ws_matrix.row_dimensions[2].height = 16

            # Ligne 3 — en-têtes semaines
            ws_matrix.row_dimensions[3].height = 20
            for wi, (ws_key, ws_days) in enumerate(weeks_map.items()):
                col_start = FIXED_COLS + 1 + wi * DAYS_PER_WEEK
                col_end   = col_start + DAYS_PER_WEEK - 1
                ws_matrix.merge_cells(start_row=3, start_column=col_start,
                                       end_row=3,   end_column=col_end)
                we_key        = ws_key + timedelta(days=4)
                _is_past_week = we_key < exp_today
                _is_cur_week  = ws_key <= exp_today <= we_key
                _wk_bg        = "475569" if _is_past_week else "1E40AF" if _is_cur_week else C_GREY_HDR
                c = ws_matrix.cell(3, col_start)
                c.value     = (
                    f"{'✅ ' if _is_past_week else '▶ ' if _is_cur_week else ''}"
                    f"Sem. {ws_key.isocalendar()[1]}  "
                    f"{ws_key.strftime('%d/%m')} → {we_key.strftime('%d/%m/%Y')}"
                )
                c.font      = Font(name="Arial", size=9, bold=True, color=C_WHITE)
                c.fill      = fill(_wk_bg)
                c.alignment = al_c()
                c.border    = border_medium()

            # Ligne 4 — en-têtes colonnes fixes
            ws_matrix.row_dimensions[4].height = 34
            fixed_headers = ["Point de prélèvement", "Lieu", "Type"]
            fixed_widths  = [32, 9, 10]
            for ci, (h, w) in enumerate(zip(fixed_headers, fixed_widths), start=1):
                c = ws_matrix.cell(4, ci)
                c.value     = h
                c.font      = Font(name="Arial", size=9, bold=True, color=C_WHITE)
                c.fill      = fill(C_BLUE2)
                c.alignment = al_c(wrap=True)
                c.border    = border_all()
                ws_matrix.column_dimensions[get_column_letter(ci)].width = w

            # Ligne 4 — en-têtes jours
            for wi, (ws_key, ws_days) in enumerate(weeks_map.items()):
                day_date_map = {d.weekday(): d for d in ws_days}
                for di in range(DAYS_PER_WEEK):
                    col       = FIXED_COLS + 1 + wi * DAYS_PER_WEEK + di
                    d_for_col = day_date_map.get(di)
                    if d_for_col:
                        is_today_col = (d_for_col == exp_today)
                        label  = f"{JOURS_XL[di][:3]}\n{d_for_col.strftime('%d/%m')}"
                        bg_col = "DBEAFE" if is_today_col else "EFF6FF"
                        fc_col = "1E40AF"
                    else:
                        label  = f"{JOURS_XL[di][:3]}\n—"
                        bg_col = "F1F5F9"
                        fc_col = "94A3B8"
                    c = ws_matrix.cell(4, col)
                    c.value     = label
                    c.font      = Font(name="Arial", size=8, bold=True, color=fc_col)
                    c.fill      = fill(bg_col)
                    c.alignment = al_c(wrap=True)
                    c.border    = border_all()
                    ws_matrix.column_dimensions[get_column_letter(col)].width = 8

            ws_matrix.freeze_panes = "D5"

            RISK_BG = {
                "1": "DCFCE7", "2": "D9F99D",
                "3": "FEF9C3", "4": "FFEDD5", "5": "FEE2E2",
            }
            RISK_FC = {
                "1": "166534", "2": "365314",
                "3": "713F12", "4": "7C2D12", "5": "7F1D1D",
            }

            # ── Lignes données ────────────────────────────────────────────
            data_row = 5
            for pt_idx, pt in enumerate(pts_all):
                rc         = (pt.get('room_class') or '').strip()
                rv         = str(pt.get('risk_level', ''))
                type_lbl   = pt.get('type', '—')
                poste_type = pt.get('poste_type', 'non_applicable')
                is_class_a = rc.strip().upper() == "A"
                row_bg     = "FFFFFF" if pt_idx % 2 == 0 else "F8FAFC"

                ws_matrix.row_dimensions[data_row].height = 18

                c = ws_matrix.cell(data_row, 1)
                c.value     = ("💨 " if type_lbl == "Air" else "🧴 ") + pt['label']
                c.font      = Font(name="Arial", size=9, bold=True, color=C_TEXT)
                c.fill      = fill(RISK_BG.get(rv, row_bg))
                c.alignment = al_l()
                c.border    = border_all()

                rc_upper = rc.strip().upper()
                if rc_upper == "A":
                    lieu_lbl = "Isolateur"
                elif rc_upper in ("B", "C", "D"):
                    lieu_lbl = "Salle"
                else:
                    lieu_lbl = rc or "—"

                c = ws_matrix.cell(data_row, 2)
                c.value     = lieu_lbl
                c.font      = Font(name="Arial", size=9, bold=True,
                                   color=RISK_FC.get(rv, C_TEXT))
                c.fill      = fill(RISK_BG.get(rv, row_bg))
                c.alignment = al_c()
                c.border    = border_all()

                c = ws_matrix.cell(data_row, 3)
                c.value     = type_lbl
                c.font      = font(9, color=C_TEXT)
                c.fill      = fill(row_bg)
                c.alignment = al_c()
                c.border    = border_all()

                _freq_raw_xl  = pt.get('frequency')
                _freq_unit_xl = pt.get('frequency_unit', '/ semaine')
                try:
                    _freq_val_xl = float(_freq_raw_xl) if _freq_raw_xl else 0.0
                except (TypeError, ValueError):
                    _freq_val_xl = 0.0
                _is_multi_day = ('/ jour' in _freq_unit_xl and _freq_val_xl > 1)

                poste_counter = 0

                for wi, (ws_key, ws_days) in enumerate(weeks_map.items()):
                    day_date_map  = {d.weekday(): d for d in ws_days}
                    _fri_this_wk  = ws_key + timedelta(days=4)
                    _wk_is_past   = _fri_this_wk < exp_today

                    for di in range(DAYS_PER_WEEK):
                        col       = FIXED_COLS + 1 + wi * DAYS_PER_WEEK + di
                        d_for_col = day_date_map.get(di)
                        c         = ws_matrix.cell(data_row, col)

                        if not d_for_col:
                            c.value     = ""
                            c.fill      = fill("F1F5F9")
                            c.alignment = al_c()
                            c.border    = border_all()
                            continue

                        tasks_day  = xl_plan.get(d_for_col, [])
                        is_planned = any(t.get("label") == pt["label"] for t in tasks_day)
                        is_done    = any(
                            p.get("label") == pt["label"]
                            and not p.get("archived", False)
                            and p.get("date")
                            and datetime.fromisoformat(p["date"]).date() == d_for_col
                            for p in st.session_state.prelevements
                        )
                        # Non-fait auto : planifié dans le plan brut mais skippé
                        _raw_day_tasks = monthly_plan_raw.get(d_for_col, [])
                        _was_planned_raw = any(t.get("label") == pt["label"] for t in _raw_day_tasks)
                        _is_skipped = pt["label"] in set(
                            st.session_state["planning_skips"].get(d_for_col.isoformat(), [])
                        )
                        is_auto_nonfait = (
                            _wk_is_past
                            and _was_planned_raw
                            and not is_done
                            and _is_skipped
                            and include_nonfaits
                        )

                        if is_done:
                            c.value = "✓"
                            c.font  = Font(name="Arial", size=11, bold=True, color=C_GREEN)
                            c.fill  = fill("DCFCE7")

                        elif is_auto_nonfait:
                            # Affiché dans la colonne d'origine (semaine passée)
                            c.value = "⏭"
                            c.font  = Font(name="Arial", size=10, bold=True, color=C_RED)
                            c.fill  = fill("FEE2E2")

                        elif is_planned:
                            if _is_multi_day:
                                n_passages = max(
                                    sum(1 for t in tasks_day if t.get('label') == pt['label']),
                                    int(_freq_val_xl),
                                )
                                c.value = "  ".join(f"X {i+1}" for i in range(n_passages))
                                c.font  = Font(name="Arial", size=10, bold=True, color=C_BLUE2)
                                c.fill  = fill("DBEAFE")

                            elif is_class_a and poste_type == "specifique":
                                poste_num     = (poste_counter % 2) + 1
                                poste_counter += 1
                                c.value = f"X {poste_num}"
                                c.font  = Font(name="Arial", size=10, bold=True, color=C_BLUE2)
                                c.fill  = fill("DBEAFE")

                            else:
                                c.value = "X"
                                c.font  = Font(name="Arial", size=10, bold=True, color=C_BLUE2)
                                c.fill  = fill("DBEAFE")

                        else:
                            c.value = ""
                            c.fill  = fill("FFFFFF" if pt_idx % 2 == 0 else "F8FAFC")

                        c.alignment = al_c()
                        c.border    = border_all()

                data_row += 1

            # ── Ligne totaux par jour ─────────────────────────────────────
            ws_matrix.row_dimensions[data_row].height = 20
            ws_matrix.merge_cells(start_row=data_row, start_column=1,
                                   end_row=data_row,   end_column=FIXED_COLS)
            c = ws_matrix.cell(data_row, 1)
            c.value     = "TOTAL prélèvements planifiés"
            c.font      = Font(name="Arial", size=9, bold=True, color=C_WHITE)
            c.fill      = fill(C_GREY_HDR)
            c.alignment = al_c()
            c.border    = border_all()

            for wi, (ws_key, ws_days) in enumerate(weeks_map.items()):
                day_date_map = {d.weekday(): d for d in ws_days}
                for di in range(DAYS_PER_WEEK):
                    col       = FIXED_COLS + 1 + wi * DAYS_PER_WEEK + di
                    d_for_col = day_date_map.get(di)
                    c         = ws_matrix.cell(data_row, col)
                    if d_for_col:
                        total_day = len(xl_plan.get(d_for_col, []))
                        c.value   = total_day if total_day > 0 else "—"
                        c.font    = Font(name="Arial", size=9, bold=True,
                                        color=C_WHITE if total_day > 0 else "94A3B8")
                        c.fill    = fill(C_BLUE2 if total_day > 0 else C_GREY_HDR)
                    else:
                        c.value = ""
                        c.fill  = fill("334155")
                    c.alignment = al_c()
                    c.border    = border_all()

            # ── Feuille 2 — Non-faits récapitulatif ──────────────────────
            ws_nf = wb.create_sheet("Non-faits")
            ws_nf.sheet_view.showGridLines = False

            ws_nf.merge_cells("A1:F1")
            ws_nf.cell(1, 1).value     = "RÉCAPITULATIF DES PRÉLÈVEMENTS NON RÉALISÉS"
            ws_nf.cell(1, 1).font      = Font(name="Arial", size=12, bold=True, color=C_WHITE)
            ws_nf.cell(1, 1).fill      = fill("DC2626")
            ws_nf.cell(1, 1).alignment = al_c()
            ws_nf.row_dimensions[1].height = 24

            nf_headers = ["Date prévue", "Semaine", "Point", "Classe", "Type", "Statut"]
            nf_widths   = [14, 10, 38, 9, 9, 18]
            for ci, (h, w) in enumerate(zip(nf_headers, nf_widths), start=1):
                c = ws_nf.cell(2, ci)
                c.value     = h
                c.font      = Font(name="Arial", size=9, bold=True, color=C_WHITE)
                c.fill      = fill("7F1D1D")
                c.alignment = al_c(wrap=True)
                c.border    = border_all()
                ws_nf.column_dimensions[get_column_letter(ci)].width = w
            ws_nf.row_dimensions[2].height = 20

            nf_row = 3
            for d in sorted(st.session_state["planning_skips"].keys()):
                try:
                    d_obj = date_type.fromisoformat(d)
                except Exception:
                    continue
                if d_obj not in [ed for ed in exp_dates]:
                    continue
                skipped_lbls = st.session_state["planning_skips"][d]
                if not skipped_lbls:
                    continue
                for lbl in skipped_lbls:
                    done = any(
                        p.get("label") == lbl and p.get("date")
                        and datetime.fromisoformat(p["date"]).date() == d_obj
                        for p in st.session_state.prelevements
                    )
                    if done:
                        continue
                    _pt_info = next((p for p in pts_all if p.get("label") == lbl), {})
                    statut   = "⏭ Auto (sem. passée)" if (
                        d_obj + timedelta(days=(4 - d_obj.weekday()))) < exp_today else "⬜ Manuel"
                    ws_nf.row_dimensions[nf_row].height = 16
                    row_bg_nf = "FFF1F2" if nf_row % 2 == 0 else "FFFFFF"
                    vals = [
                        d_obj.strftime("%d/%m/%Y"),
                        f"Sem. {d_obj.isocalendar()[1]}",
                        lbl,
                        _pt_info.get("room_class", "—"),
                        _pt_info.get("type", "—"),
                        statut,
                    ]
                    for ci, val in enumerate(vals, start=1):
                        c = ws_nf.cell(nf_row, ci)
                        c.value     = val
                        c.font      = Font(name="Arial", size=9,
                                          color="DC2626" if "Auto" in statut else C_TEXT)
                        c.fill      = fill(row_bg_nf)
                        c.alignment = al_l() if ci == 3 else al_c()
                        c.border    = border_all()
                    nf_row += 1

            if nf_row == 3:
                ws_nf.merge_cells("A3:F3")
                ws_nf.cell(3, 1).value     = "✅ Aucun prélèvement non-réalisé sur cette période"
                ws_nf.cell(3, 1).font      = Font(name="Arial", size=9, color="16A34A")
                ws_nf.cell(3, 1).fill      = fill("F0FDF4")
                ws_nf.cell(3, 1).alignment = al_c()

            # ── Export ────────────────────────────────────────────────────
            buf = _io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            fname = f"planning_URC_{exp_today.strftime('%Y%m%d')}.xlsx"
            st.download_button(
                "⬇️ Télécharger le planning Excel",
                data=buf.getvalue(),
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            st.success(
                f"✅ Fichier **{fname}** généré — "
                f"Onglets : **Planning Semaine** (matriciel) · **Non-faits** (récapitulatif)"
            )
# ═══════════════════════════════════════════════════════════════════════════════
# TAB : ANALYSE
# ═══════════════════════════════════════════════════════════════════════════════
if active == "analyse":
    st.markdown("### 📋 Analyse des prélèvements")
    surv  = st.session_state.surveillance
    total = len(surv)

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _get_criticite(germ_name):
        for g in st.session_state.get("germs", []):
            if g.get("name", "") == germ_name:
                return int(g.get("criticite", 0) or 0)
        return 0

    def _crit_label(c):
        return {5:"Critique",4:"Majeur",3:"Important",2:"Modéré",1:"Limité"}.get(c,"—")

    def _crit_color(c):
        return {5:"#7c3aed",4:"#ef4444",3:"#f97316",2:"#f59e0b",1:"#22c55e"}.get(c,"#94a3b8")

    def _parse_qr_to_point_id(raw: str):
        import json
        if not raw:
            return None, None
        raw = raw.strip()
        if raw.startswith("{"):
            try:
                data = json.loads(raw)
                pid = str(data.get("id", "")).strip()
                if pid:
                    return pid, None
                return None, "QR JSON valide mais sans champ 'id'."
            except json.JSONDecodeError:
                return None, "QR JSON invalide."
        return raw, None

    def _parse_date(date_str):
        try:
            return datetime.strptime(date_str, "%Y-%m-%d").date()
        except Exception:
            return None

    def _render_liste_entries(entries, surv_ref):
        for _li, r in enumerate(entries):
            _real_idx = next(
                (
                    i for i, s in enumerate(st.session_state.surveillance)
                    if s.get("date") == r.get("date")
                    and s.get("prelevement") == r.get("prelevement")
                    and s.get("operateur", s.get("preleveur","")) == r.get("operateur", r.get("preleveur",""))
                    and s.get("ufc") == r.get("ufc")
                ),
                None,
            )
            if _real_idx is None:
                continue

            status_r  = r.get("status", "ok")
            mc_statut = r.get("mc_statut", "")
            mc_detail = r.get("mc_detail", "")
            mc_date   = r.get("mc_date", "")
            germ_r    = r.get("germ_saisi", "") or r.get("germ_match","") or ""

            if status_r == "action":
                _sc = "#dc2626"; _sb = "#fef2f2"; _sl = "🚨 ACTION"
            elif status_r == "alert":
                _sc = "#d97706"; _sb = "#fffbeb"; _sl = "⚠️ ALERTE"
            else:
                _sc = "#16a34a"; _sb = "#f0fdf4"; _sl = "✅ CONFORME"

            _mc_badge = ""
            if status_r in ("alert", "action"):
                if mc_statut == "fait":
                    _mc_badge = (
                        "<span style='background:#22c55e;color:#fff;border-radius:5px;"
                        "padding:1px 8px;font-size:.65rem;font-weight:700;margin-left:6px'>"
                        "🔧 MC FAITES</span>"
                    )
                else:
                    _mc_badge = (
                        "<span style='background:#f59e0b;color:#fff;border-radius:5px;"
                        "padding:1px 8px;font-size:.65rem;font-weight:700;margin-left:6px'>"
                        "🔧 MC EN ATTENTE</span>"
                    )

            st.markdown(
                f"<div style='background:{_sb};border:1px solid {_sc}44;"
                f"border-left:4px solid {_sc};border-radius:8px;"
                f"padding:10px 14px;margin-bottom:6px'>"
                f"<div style='display:flex;justify-content:space-between;"
                f"align-items:center;margin-bottom:4px'>"
                f"<span style='font-size:.82rem;font-weight:700;color:#0f172a'>"
                f"📍 {r.get('prelevement','—')} &nbsp;·&nbsp; "
                f"🦠 {germ_r or '—'} &nbsp;·&nbsp; "
                f"{r.get('ufc', 0)} UFC</span>"
                f"<div>"
                f"<span style='background:{_sc};color:#fff;border-radius:5px;"
                f"padding:1px 8px;font-size:.7rem;font-weight:700'>{_sl}</span>"
                f"{_mc_badge}</div>"
                f"</div>"
                f"<div style='font-size:.72rem;color:#475569'>"
                f"📅 {r.get('date_prelevement', r.get('date','—'))} &nbsp;·&nbsp; "
                f"👤 {r.get('operateur', r.get('preleveur','—'))} &nbsp;·&nbsp; "
                f"🏥 {r.get('room_class','—')} &nbsp;·&nbsp; "
                f"Score : {r.get('total_score','—')}"
                f"</div>"
                f"</div>",
                unsafe_allow_html=True,
            )

            _remarque  = r.get("remarque", "")
            _triggered = r.get("triggered_by", "")
            if _remarque or _triggered:
                st.markdown(
                    f"<div style='background:#f8fafc;border-left:3px solid #cbd5e1;"
                    f"border-radius:0 6px 6px 0;padding:6px 12px;margin-top:-4px;"
                    f"margin-bottom:6px;font-size:.72rem;color:#475569'>"
                    + (f"💬 {_remarque}" if _remarque else "")
                    + (" &nbsp;·&nbsp; " if _remarque and _triggered else "")
                    + "</div>",
                    unsafe_allow_html=True,
                )

            _germs_detail = r.get("germs_detail", [])
            if len(_germs_detail) > 1:
                _gd_html = " &nbsp;·&nbsp; ".join(
                    f"<b>{'👑 ' if g.get('is_worst') else ''}{g['name']}</b> "
                    f"({g.get('ufc',0)} UFC · score {g.get('germ_score','?')})"
                    for g in _germs_detail
                )
                st.markdown(
                    f"<div style='background:#eff6ff;border-left:3px solid #93c5fd;"
                    f"border-radius:0 6px 6px 0;padding:6px 12px;margin-top:-4px;"
                    f"margin-bottom:6px;font-size:.72rem;color:#1e40af'>"
                    f"🦠 {_gd_html}</div>",
                    unsafe_allow_html=True,
                )

            _tab_mc, _tab_edit, _tab_del = st.tabs([
                "🔧 Mesures correctives" if status_r in ("alert","action") else "ℹ️ Détails",
                "✏️ Modifier",
                "🗑️ Supprimer",
            ])

            with _tab_mc:
                if status_r in ("alert", "action"):
                    _render_mesures_correctives(st.session_state.surveillance[_real_idx], _real_idx)
                else:
                    st.markdown(
                        "<div style='font-size:.82rem;color:#16a34a;padding:8px 0'>"
                        "✅ Résultat conforme — aucune mesure corrective requise.</div>",
                        unsafe_allow_html=True,
                    )

            with _tab_edit:
                st.markdown("**✏️ Modifier cette entrée**")
                germ_names_edit  = sorted([g['name'] for g in st.session_state.germs])
                _germ_opts_edit  = ["Négatif"] + germ_names_edit
                _germs_det_edit  = r.get("germs_detail", [])

                # ── Champs communs ─────────────────────────────────────────────
                _ec1, _ec2 = st.columns(2)
                with _ec1:
                    _new_prelev = st.text_input(
                        "📍 Point de prélèvement",
                        value=r.get("prelevement", ""),
                        key=f"edit_prelev_{_li}",
                    )
                    _new_date = st.text_input(
                        "📅 Date (YYYY-MM-DD)",
                        value=r.get("date_prelevement", r.get("date", "")),
                        key=f"edit_date_{_li}",
                    )
                    _new_oper = st.text_input(
                        "👤 Opérateur",
                        value=r.get("operateur", r.get("preleveur", "")),
                        key=f"edit_oper_{_li}",
                    )
                with _ec2:
                    if len(_germs_det_edit) > 1:
                        st.markdown(
                            f"<div style='background:#eff6ff;border:1px solid #93c5fd;"
                            f"border-radius:10px;padding:12px;margin-top:4px;"
                            f"font-size:.8rem;color:#1e40af;font-weight:700'>"
                            f"🦠 {len(_germs_det_edit)} germes saisis<br>"
                            f"<span style='font-weight:400;font-size:.75rem;color:#475569'>"
                            f"Éditez chaque germe dans ses onglets ci-dessous.</span></div>",
                            unsafe_allow_html=True,
                        )
                    else:
                        _cur_germ_e = (
                            (_germs_det_edit[0].get("name","") if _germs_det_edit else "")
                            or r.get("germ_saisi","") or "Négatif"
                        )
                        _germ_idx_e = (
                            _germ_opts_edit.index(_cur_germ_e)
                            if _cur_germ_e in _germ_opts_edit else 0
                        )
                        _new_germ = st.selectbox(
                            "🦠 Germe identifié",
                            _germ_opts_edit,
                            index=_germ_idx_e,
                            key=f"edit_germ_{_li}",
                        )
                        _new_ufc = st.number_input(
                            "UFC",
                            min_value=0,
                            value=int(
                                (_germs_det_edit[0].get("ufc",0) if _germs_det_edit
                                else r.get("ufc", 0)) or 0
                            ),
                            step=1,
                            key=f"edit_ufc_{_li}",
                        )

                _new_remarque = st.text_area(
                    "💬 Remarque",
                    value=r.get("remarque", ""),
                    height=70,
                    key=f"edit_rem_{_li}",
                )

                # ══════════════════════════════════════════════════════════════
                # CAS MULTI-GERMES : onglets par germe
                # ══════════════════════════════════════════════════════════════
                if len(_germs_det_edit) > 1:
                    st.markdown(
                        "<div style='font-size:.78rem;font-weight:700;color:#1e40af;"
                        "margin:10px 0 4px 0'>🦠 Modifier chaque germe</div>",
                        unsafe_allow_html=True,
                    )
                    _mg_tab_labels = [
                        f"{'👑 ' if gde.get('is_worst') else ''}Germe {gi+1} · "
                        f"{(gde.get('name') or '?')[:22]}"
                        for gi, gde in enumerate(_germs_det_edit)
                    ]
                    _mg_tabs = st.tabs(_mg_tab_labels)

                    for gi, (mgtab, gde) in enumerate(zip(_mg_tabs, _germs_det_edit)):
                        with mgtab:
                            _mg_c1, _mg_c2 = st.columns(2)
                            _cur_mg_name = gde.get("name","") or "Négatif"
                            _mg_idx = (
                                _germ_opts_edit.index(_cur_mg_name)
                                if _cur_mg_name in _germ_opts_edit else 0
                            )
                            with _mg_c1:
                                _ng_name = st.selectbox(
                                    "🦠 Germe",
                                    _germ_opts_edit,
                                    index=_mg_idx,
                                    key=f"edit_mg_name_{_li}_{gi}",
                                )
                            with _mg_c2:
                                _ng_ufc = st.number_input(
                                    "UFC",
                                    min_value=0,
                                    value=int(gde.get("ufc",0) or 0),
                                    step=1,
                                    key=f"edit_mg_ufc_{_li}_{gi}",
                                )
                            # Aperçu score pour ce germe
                            if _ng_name != "Négatif":
                                _gobj_mg = next(
                                    (g for g in st.session_state.germs if g['name'] == _ng_name), None
                                )
                                if _gobj_mg:
                                    _mg_gscore = (
                                        int(_gobj_mg.get('pathogenicity',1))
                                        * int(_gobj_mg.get('resistance',1))
                                        * int(_gobj_mg.get('dissemination',1))
                                    )
                                    _loc_c_mg  = int(r.get("location_criticality",1) or 1)
                                    _mg_total  = _loc_c_mg * _mg_gscore
                                    _mg_st, _mg_lbl, _mg_col = _evaluate_score(_mg_total)
                                    st.markdown(
                                        f"<div style='background:{_mg_col}11;border:1px solid {_mg_col}44;"
                                        f"border-radius:8px;padding:6px 10px;font-size:.75rem;"
                                        f"font-weight:700;color:{_mg_col}'>"
                                        f"Score : {_mg_total} (lieu {_loc_c_mg} × germe {_mg_gscore}) → {_mg_lbl}"
                                        f"</div>",
                                        unsafe_allow_html=True,
                                    )

                    if st.button(
                        "💾 Sauvegarder les modifications",
                        key=f"edit_save_{_li}",
                        type="primary",
                        use_container_width=True,
                    ):
                        # ── Reconstruire germs_detail depuis les widgets ───────
                        _loc_c_sv = int(r.get("location_criticality",1) or 1)
                        _new_gd   = []
                        for gi, gde in enumerate(_germs_det_edit):
                            _sv_name = st.session_state.get(f"edit_mg_name_{_li}_{gi}", gde.get("name","Négatif"))
                            _sv_ufc  = st.session_state.get(f"edit_mg_ufc_{_li}_{gi}",  int(gde.get("ufc",0) or 0))
                            _sv_gscore = 0
                            if _sv_name != "Négatif":
                                _go2 = next((g for g in st.session_state.germs if g['name'] == _sv_name), None)
                                if _go2:
                                    _sv_gscore = (
                                        int(_go2.get('pathogenicity',1))
                                        * int(_go2.get('resistance',1))
                                        * int(_go2.get('dissemination',1))
                                    )
                            _new_gd.append({
                                "name":       _sv_name,
                                "ufc":        _sv_ufc,
                                "germ_score": _sv_gscore,
                                "is_worst":   False,
                            })

                        # ── Trouver le germe le plus critique (worst) ─────────
                        _best_ts    = -1
                        _worst_name = None
                        _worst_gs   = 0
                        _total_ufc_sv = 0
                        for ngd in _new_gd:
                            if ngd["name"] != "Négatif" and ngd["ufc"] > 0:
                                _ts_ngd = _loc_c_sv * ngd["germ_score"]
                                _total_ufc_sv += ngd["ufc"]
                                if _ts_ngd > _best_ts:
                                    _best_ts    = _ts_ngd
                                    _worst_name = ngd["name"]
                                    _worst_gs   = ngd["germ_score"]

                        for ngd in _new_gd:
                            ngd["is_worst"] = (ngd["name"] == _worst_name and _worst_name is not None)

                        if _worst_name:
                            _final_ts  = _loc_c_sv * _worst_gs
                            _final_st, _, _ = _evaluate_score(_final_ts)
                        else:
                            _final_ts = 0; _final_st = "ok"; _worst_gs = 0

                        st.session_state.surveillance[_real_idx].update({
                            "prelevement":  _new_prelev.strip(),
                            "date":         _new_date.strip(),
                            "operateur":    _new_oper.strip(),
                            "germ_saisi":   _worst_name or "Négatif",
                            "ufc":          _total_ufc_sv,
                            "ufc_total":    _total_ufc_sv,
                            "germ_score":   _worst_gs,
                            "total_score":  _final_ts,
                            "status":       _final_st,
                            "remarque":     _new_remarque.strip(),
                            "germs_detail": _new_gd,
                        })
                        save_surveillance(st.session_state.surveillance)
                        st.success("✅ Entrée mise à jour — germes et stats recalculés.")
                        st.rerun()

                # ══════════════════════════════════════════════════════════════
                # CAS MONO-GERME : logique d'origine
                # ══════════════════════════════════════════════════════════════
                else:
                    if _new_germ != "Négatif":
                        _gobj_s = next(
                            (g for g in st.session_state.germs if g['name'] == _new_germ), None
                        )
                        if _gobj_s:
                            _new_germ_score = (
                                int(_gobj_s.get('pathogenicity',1))
                                * int(_gobj_s.get('resistance',1))
                                * int(_gobj_s.get('dissemination',1))
                            )
                            _loc_c_edit = int(r.get("location_criticality",1) or 1)
                            _new_total  = _loc_c_edit * _new_germ_score
                            _new_status, _new_lbl, _new_col = _evaluate_score(_new_total)
                            st.markdown(
                                f"<div style='background:{_new_col}11;border:1px solid {_new_col}44;"
                                f"border-radius:8px;padding:8px 12px;font-size:.78rem;"
                                f"font-weight:700;color:{_new_col}'>"
                                f"Score recalculé : {_new_total} "
                                f"(lieu {_loc_c_edit} × germe {_new_germ_score}) → {_new_lbl}"
                                f"</div>",
                                unsafe_allow_html=True,
                            )
                        else:
                            _new_germ_score = int(r.get("germ_score",0) or 0)
                            _new_total      = int(r.get("total_score",0) or 0)
                            _new_status     = r.get("status","ok")
                    else:
                        _new_germ_score = 0
                        _new_total      = 0
                        _new_status     = "ok"

                    if st.button(
                        "💾 Sauvegarder les modifications",
                        key=f"edit_save_{_li}",
                        type="primary",
                        use_container_width=True,
                    ):
                        # Met à jour germs_detail si existant, sinon champ plat
                        _upd_gd = (
                            [{"name": _new_germ, "ufc": _new_ufc,
                            "germ_score": _new_germ_score, "is_worst": True}]
                            if _new_germ != "Négatif" else []
                        )
                        st.session_state.surveillance[_real_idx].update({
                            "prelevement":  _new_prelev.strip(),
                            "date":         _new_date.strip(),
                            "operateur":    _new_oper.strip(),
                            "germ_saisi":   _new_germ,
                            "ufc":          _new_ufc,
                            "ufc_total":    _new_ufc,
                            "germ_score":   _new_germ_score,
                            "total_score":  _new_total,
                            "status":       _new_status,
                            "remarque":     _new_remarque.strip(),
                            "germs_detail": _upd_gd,
                        })
                        save_surveillance(st.session_state.surveillance)
                        st.success("✅ Entrée mise à jour — stats recalculées automatiquement.")
                        st.rerun()

            with _tab_del:
                st.markdown(
                    f"<div style='background:#fef2f2;border:1.5px solid #fca5a5;"
                    f"border-radius:10px;padding:14px 16px;margin-bottom:12px'>"
                    f"<div style='font-weight:700;color:#991b1b;margin-bottom:6px'>"
                    f"🗑️ Supprimer cette entrée ?</div>"
                    f"<div style='font-size:.8rem;color:#7f1d1d'>"
                    f"📍 {r.get('prelevement','—')} &nbsp;·&nbsp; "
                    f"🦠 {germ_r or '—'} &nbsp;·&nbsp; "
                    f"📅 {r.get('date_prelevement', r.get('date','—'))} &nbsp;·&nbsp; "
                    f"Cette action est <strong>irréversible</strong> et mettra à jour "
                    f"toutes les statistiques.</div>"
                    f"</div>",
                    unsafe_allow_html=True,
                )
                _confirm_key = f"confirm_del_{_li}"
                if not st.session_state.get(_confirm_key, False):
                    if st.button(
                        "🗑️ Supprimer cette entrée",
                        key=f"del_btn_{_li}",
                        use_container_width=True,
                    ):
                        st.session_state[_confirm_key] = True
                        st.rerun()
                else:
                    st.error("⚠️ Confirmer la suppression définitive ?")
                    _dc1, _dc2 = st.columns(2)
                    with _dc1:
                        if st.button(
                            "✅ OUI — Supprimer définitivement",
                            key=f"del_confirm_{_li}",
                            type="primary",
                            use_container_width=True,
                        ):
                            st.session_state.surveillance.pop(_real_idx)
                            save_surveillance(st.session_state.surveillance)
                            st.session_state[_confirm_key] = False
                            st.success("🗑️ Entrée supprimée — stats recalculées.")
                            st.rerun()
                    with _dc2:
                        if st.button(
                            "❌ Annuler",
                            key=f"del_cancel_{_li}",
                            use_container_width=True,
                        ):
                            st.session_state[_confirm_key] = False
                            st.rerun()

    # ── fin helpers ───────────────────────────────────────────────────────────

    if surv:
        with st.expander("⚙️ Export, filtres & métriques", expanded=True):

            # ── Export / Vider ────────────────────────────────────────────────
            c_dl, c_cl = st.columns(2)
            with c_dl:
                csv_str  = io.StringIO()
                all_keys = list(dict.fromkeys(k for r in surv for k in r.keys()))
                writer   = csv.DictWriter(csv_str, fieldnames=all_keys, extrasaction="ignore")
                writer.writeheader()
                writer.writerows(surv)
                st.download_button(
                    "⬇️ Télécharger CSV", csv_str.getvalue(),
                    "surveillance.csv", "text/csv",
                    use_container_width=True)
            with c_cl:
                if st.button("🗑️ Vider l'historique", use_container_width=True):
                    st.session_state.surveillance = []
                    save_surveillance([])
                    try:
                        supa = get_supabase_client()
                        if supa:
                            supa.table("app_data").upsert({
                                "key": "surveillance",
                                "value": json.dumps([], ensure_ascii=False)
                            }).execute()
                    except Exception:
                        pass
                    if os.path.exists(CSV_FILE):
                        os.remove(CSV_FILE)
                    st.rerun()

            from datetime import date as dt_date

            # ── Bornes ────────────────────────────────────────────────────────
            all_dates_ok = [d for d in (
                _parse_date(r.get("date_prelevement", r.get("date", ""))) for r in surv
            ) if d]
            d_min = min(all_dates_ok) if all_dates_ok else dt_date.today()
            d_max = max(all_dates_ok) if all_dates_ok else dt_date.today()

            if "hist_date_debut_val" not in st.session_state:
                st.session_state["hist_date_debut_val"] = d_min
            if "hist_date_fin_val" not in st.session_state:
                st.session_state["hist_date_fin_val"] = d_max

            st.session_state["hist_date_debut_val"] = max(
                d_min, min(st.session_state["hist_date_debut_val"], d_max))
            st.session_state["hist_date_fin_val"] = max(
                d_min, min(st.session_state["hist_date_fin_val"], d_max))
            if st.session_state["hist_date_debut_val"] > st.session_state["hist_date_fin_val"]:
                st.session_state["hist_date_fin_val"] = st.session_state["hist_date_debut_val"]

            st.markdown(
                "<div style='background:#f0f9ff;border:1px solid #bae6fd;border-radius:10px;"
                "padding:12px 16px;margin:10px 0 6px 0'>"
                "<div style='font-size:.8rem;font-weight:700;color:#0369a1;margin-bottom:8px'>"
                "🗓️ Filtrer par période</div>",
                unsafe_allow_html=True)
            cf1, cf2, cf3 = st.columns([2, 2, 1])
            with cf1:
                date_debut = st.date_input(
                    "Du", value=st.session_state["hist_date_debut_val"],
                    min_value=d_min, max_value=d_max, key="hist_date_debut_input")
                st.session_state["hist_date_debut_val"] = date_debut
            with cf2:
                date_fin = st.date_input(
                    "Au", value=st.session_state["hist_date_fin_val"],
                    min_value=d_min, max_value=d_max, key="hist_date_fin_input")
                st.session_state["hist_date_fin_val"] = date_fin
            with cf3:
                st.markdown("<div style='height:22px'></div>", unsafe_allow_html=True)
                if st.button("↺ Reset", use_container_width=True):
                    st.session_state["hist_date_debut_val"] = d_min
                    st.session_state["hist_date_fin_val"]   = d_max
                    st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)

            # ── Métriques ─────────────────────────────────────────────────────
            surv_f = [r for r in surv
                if _parse_date(r.get("date_prelevement", r.get("date", ""))) is not None
                and date_debut <= _parse_date(r.get("date_prelevement", r.get("date", ""))) <= date_fin]
            total_f = len(surv_f)
            if total_f < total:
                st.caption(f"🔍 {total_f} résultat(s) sur {total} — "
                           f"{date_debut.strftime('%d/%m/%Y')} → {date_fin.strftime('%d/%m/%Y')}")

            alerts  = sum(1 for r in surv_f if r.get("status") == "alert")
            actions = sum(1 for r in surv_f if r.get("status") == "action")
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Total",        total_f)
            c2.metric("✅ Conformes", total_f - alerts - actions)
            c3.metric("⚠️ Alertes",  alerts)
            c4.metric("🚨 Actions",   actions)

        # ── Tabs (en dehors de l'expander) ────────────────────────────────────
        st.divider()
        hist_tab_pts, hist_tab_germs, hist_tab_prev, hist_tab_liste = st.tabs([
            "📍 Stats par point",
            "🦠 Stats par germe",
            "👤 Répartition par préleveur",
            "📋 Liste des entrées",
        ])

        # ══════════════════════════════════════════════════════════════════════
        # ONGLET 1 : STATS PAR POINT
        # ══════════════════════════════════════════════════════════════════════
        with hist_tab_pts:
            from collections import defaultdict
            import json as _json_pts

            pts_stats = defaultdict(lambda: {
                "total":0,"positives":0,"negatives":0,
                "alertes":0,"actions":0,
                "germes":defaultdict(int),
                "ufc_j2_list":[],"ufc_j7_list":[],
            })
            for r in surv_f:
                pt     = r.get("prelevement","—")
                germ   = r.get("germ_saisi","") or r.get("germ_match","") or ""
                st_r   = r.get("status","ok")
                ufc_j2 = int(r.get("ufc_48h", r.get("ufc",0)) or 0)
                ufc_j7 = int(r.get("ufc_5j",  r.get("ufc",0)) or 0)

                j2_pos = ufc_j2 > 0 and germ not in ("Négatif","—","")
                j7_pos = ufc_j7 > 0 and germ not in ("Négatif","—","")

                if j2_pos:
                    pts_stats[pt]["total"]     += 1
                    pts_stats[pt]["positives"] += 1
                    pts_stats[pt]["germes"][germ] += 1
                    if st_r == "alert":    pts_stats[pt]["alertes"] += 1
                    elif st_r == "action": pts_stats[pt]["actions"] += 1
                    pts_stats[pt]["ufc_j2_list"].append(ufc_j2)
                    if ufc_j7 > 0:
                        pts_stats[pt]["ufc_j7_list"].append(ufc_j7)

                elif not j2_pos and not j7_pos:
                    pts_stats[pt]["total"]     += 1
                    pts_stats[pt]["negatives"] += 1
                    if st_r == "alert":    pts_stats[pt]["alertes"] += 1
                    elif st_r == "action": pts_stats[pt]["actions"] += 1

                else:
                    pts_stats[pt]["total"]     += 1
                    pts_stats[pt]["negatives"] += 1
                    pts_stats[pt]["total"]     += 1
                    pts_stats[pt]["positives"] += 1
                    pts_stats[pt]["germes"][germ] += 1
                    if st_r == "alert":    pts_stats[pt]["alertes"] += 1
                    elif st_r == "action": pts_stats[pt]["actions"] += 1
                    pts_stats[pt]["ufc_j7_list"].append(ufc_j7)

            sorted_pts   = sorted(pts_stats.items(), key=lambda x: -x[1]["positives"])
            chart_labels = [p[:22]+"…" if len(p)>22 else p for p,_ in sorted_pts]
            chart_data   = {
                "labels": chart_labels,
                "neg": [d["negatives"] for _,d in sorted_pts],
                "pos": [d["positives"] for _,d in sorted_pts],
            }
            chart_html = f"""
            <script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.0/chart.umd.min.js"></script>
            <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:12px;
                        padding:16px;margin-bottom:18px">
              <div style="font-size:.8rem;font-weight:700;color:#1e40af;margin-bottom:10px">
                📊 Résultats par point de prélèvement
              </div>
              <div style="width:100%;height:420px"><canvas id="ptChart"></canvas></div>
            </div>
            <script>
            (function(){{
              const d = {_json_pts.dumps(chart_data)};
              new Chart(document.getElementById('ptChart'), {{
                type:'bar',
                data:{{
                  labels:d.labels,
                  datasets:[
                    {{label:'✅ Négatifs',data:d.neg,backgroundColor:'#22c55e88',borderColor:'#22c55e',borderWidth:1}},
                    {{label:'🦠 Positifs',data:d.pos,backgroundColor:'#ef444488',borderColor:'#ef4444',borderWidth:1}}
                  ]
                }},
                options:{{
                  responsive:true,maintainAspectRatio:false,
                  plugins:{{legend:{{position:'top',labels:{{font:{{size:11}}}}}}}},
                  scales:{{
                    x:{{stacked:true,ticks:{{font:{{size:10}},maxRotation:45}}}},
                    y:{{stacked:true,beginAtZero:true,ticks:{{stepSize:1,font:{{size:10}}}}}}
                  }}
                }}
              }});
            }})();
            </script>"""
            st.components.v1.html(chart_html, height=500)

            with st.expander("📍 Détail par point de prélèvement", expanded=False):
                st.markdown(
                    "<div style='display:grid;"
                    "grid-template-columns:2fr 0.55fr 0.55fr 0.55fr 0.7fr 0.7fr 0.7fr 1.8fr;"
                    "gap:4px;background:#1e40af;border-radius:10px 10px 0 0;padding:10px 14px'>"
                    "<div style='font-size:.72rem;font-weight:800;color:#fff'>Point</div>"
                    "<div style='font-size:.72rem;font-weight:800;color:#fff;text-align:center'>Total</div>"
                    "<div style='font-size:.72rem;font-weight:800;color:#fff;text-align:center'>✅</div>"
                    "<div style='font-size:.72rem;font-weight:800;color:#fff;text-align:center'>🦠</div>"
                    "<div style='font-size:.72rem;font-weight:800;color:#fff;text-align:center'>Taux+</div>"
                    "<div style='font-size:.72rem;font-weight:800;color:#7dd3fc;text-align:center'>Moy J2</div>"
                    "<div style='font-size:.72rem;font-weight:800;color:#c4b5fd;text-align:center'>Moy J7</div>"
                    "<div style='font-size:.72rem;font-weight:800;color:#fff'>Germes détectés</div>"
                    "</div>",
                    unsafe_allow_html=True)
                for ri,(pt_name,pt_data) in enumerate(sorted_pts):
                    t     = pt_data["total"]
                    pos   = pt_data["positives"]
                    taux  = pos/t*100 if t>0 else 0
                    tc    = "#ef4444" if taux>=50 else "#f59e0b" if taux>0 else "#22c55e"
                    germes_str = ", ".join(
                        g+"("+str(n)+"x)"
                        for g,n in sorted(pt_data["germes"].items(),key=lambda x:-x[1])[:3]
                    ) or "—"
                    j2_list = pt_data["ufc_j2_list"]
                    j7_list = pt_data["ufc_j7_list"]
                    moy_j2  = str(round(sum(j2_list)/len(j2_list))) if j2_list else "—"
                    moy_j7  = str(round(sum(j7_list)/len(j7_list))) if j7_list else "—"
                    row_bg  = "#f8fafc" if ri%2==0 else "#ffffff"
                    st.markdown(
                        "<div style='display:grid;"
                        "grid-template-columns:2fr 0.55fr 0.55fr 0.55fr 0.7fr 0.7fr 0.7fr 1.8fr;"
                        "gap:4px;background:"+row_bg+";border:1px solid #e2e8f0;border-top:none;"
                        "padding:9px 14px;align-items:center'>"
                        "<div style='font-size:.88rem;font-weight:700;color:#0f172a'>📍 "+pt_name+"</div>"
                        "<div style='font-size:1rem;font-weight:800;color:#1e40af;text-align:center'>"+str(t)+"</div>"
                        "<div style='font-size:1rem;font-weight:800;color:#22c55e;text-align:center'>"+str(pt_data["negatives"])+"</div>"
                        "<div style='text-align:center'><span style='background:"+tc+"22;color:"+tc+";"
                        "border:1px solid "+tc+"55;border-radius:6px;padding:2px 7px;"
                        "font-size:.8rem;font-weight:700'>"+str(pos)+"</span></div>"
                        "<div style='font-size:.85rem;font-weight:700;color:"+tc+";text-align:center'>"+str(round(taux))+"%</div>"
                        "<div style='font-size:.82rem;font-weight:700;color:#0369a1;text-align:center'>"+moy_j2+"</div>"
                        "<div style='font-size:.82rem;font-weight:700;color:#7c3aed;text-align:center'>"+moy_j7+"</div>"
                        "<div style='font-size:.72rem;color:#475569;font-style:italic'>"+germes_str+"</div>"
                        "</div>",
                        unsafe_allow_html=True)
                st.markdown(
                    "<div style='background:#1e293b;border-radius:0 0 10px 10px;padding:8px 14px'>"
                    "<div style='font-size:.78rem;color:#94a3b8'>"
                    +str(len(pts_stats))+" point(s) — "+str(total_f)+" résultats"
                    +" &nbsp;|&nbsp; <span style='color:#7dd3fc'>J2 = 48h</span>"
                    +" &nbsp;|&nbsp; <span style='color:#c4b5fd'>J7 = 5 jours</span>"
                    "</div></div>",
                    unsafe_allow_html=True)

            st.divider()
            st.markdown(
                "<div style='font-size:.85rem;font-weight:700;color:#1e40af;margin-bottom:8px'>"
                "📈 Évolution UFC dans le temps</div>",
                unsafe_allow_html=True)
            pt_choices = [p for p,_ in sorted_pts]
            if pt_choices:
                selected_pt = st.selectbox("Point", options=pt_choices,
                    key="hist_pt_evol", label_visibility="collapsed")
                pt_records = sorted(
                    [r for r in surv_f if r.get("prelevement")==selected_pt],
                    key=lambda x: _parse_date(x.get("date_prelevement", x.get("date",""))) or dt_date.min)
                evol_dates=[]; evol_j2=[]; evol_j7=[]
                seuil_alerte=None; seuil_action=None
                for r in pt_records:
                    d = _parse_date(r.get("date_prelevement", r.get("date","")))
                    if not d: continue
                    evol_dates.append(d.strftime("%d/%m/%y"))
                    evol_j2.append(int(r.get("ufc_48h",r.get("ufc",0)) or 0))
                    evol_j7.append(int(r.get("ufc_5j", r.get("ufc",0)) or 0))
                    if seuil_alerte is None:
                        try: seuil_alerte=int(float(r.get("alert_threshold",50) or 50))
                        except: seuil_alerte=50
                        try: seuil_action=int(float(r.get("action_threshold",100) or 100))
                        except: seuil_action=100
                if evol_dates:
                    evol_data={"dates":evol_dates,"j2":evol_j2,"j7":evol_j7,
                               "alerte":seuil_alerte,"action":seuil_action}
                    evol_html = f"""
                    <script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.0/chart.umd.min.js"></script>
                    <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:12px;
                                padding:16px;margin-bottom:8px">
                      <div style="font-size:.78rem;font-weight:700;color:#334155;margin-bottom:10px">
                        📍 {selected_pt} — UFC/m³ au fil du temps
                      </div>
                      <div style="width:100%;height:220px"><canvas id="evolChart"></canvas></div>
                    </div>
                    <script>
                    (function(){{
                      const d={_json_pts.dumps(evol_data)};
                      const thr={{id:'thr',afterDraw(chart){{
                        const {{ctx,chartArea:{{left,right}},scales:{{y}}}}=chart;
                        if(!y)return;
                        function dl(val,color,lbl){{
                          const yp=y.getPixelForValue(val);
                          ctx.save();ctx.beginPath();ctx.setLineDash([6,4]);
                          ctx.strokeStyle=color;ctx.lineWidth=1.5;
                          ctx.moveTo(left,yp);ctx.lineTo(right,yp);ctx.stroke();
                          ctx.setLineDash([]);ctx.fillStyle=color;
                          ctx.font='bold 10px sans-serif';
                          ctx.fillText(lbl,right-56,yp-4);ctx.restore();
                        }}
                        dl(d.alerte,'#f59e0b','⚠ Alerte');
                        dl(d.action,'#ef4444','🚨 Action');
                      }}}};
                      new Chart(document.getElementById('evolChart'),{{
                        type:'line',plugins:[thr],
                        data:{{labels:d.dates,datasets:[
                            {{
                                label:'🔵 J2',data:d.j2,borderColor:'#0ea5e9',
                                backgroundColor:'#0ea5e922',borderWidth:2,tension:0.3,fill:false,
                                pointRadius:d.j2.map((v,i)=>(v===0&&d.j7[i]>0)?6:4),
                                pointStyle:d.j2.map((v,i)=>(v===0&&d.j7[i]>0)?'crossRot':'circle'),
                                pointBorderColor:'#0ea5e9',pointBorderWidth:d.j2.map((v,i)=>(v===0&&d.j7[i]>0)?2.5:1.5),
                            }},
                            {{
                                label:'🟣 J7',data:d.j7,borderColor:'#8b5cf6',
                                backgroundColor:'#8b5cf622',borderWidth:2,tension:0.3,fill:false,
                                pointRadius:d.j7.map((v,i)=>(v>0&&d.j2[i]===0)?7:4),
                                pointStyle:d.j7.map((v,i)=>(v>0&&d.j2[i]===0)?'star':'circle'),
                                pointBorderColor:'#8b5cf6',pointBorderWidth:1.5,
                            }}
                        ]}},
                        options:{{
                            responsive:true,maintainAspectRatio:false,
                            interaction:{{mode:'index',intersect:false}},
                            plugins:{{
                                legend:{{position:'top',labels:{{font:{{size:11}},boxWidth:14}}}},
                                tooltip:{{callbacks:{{
                                    label:function(item){{
                                        const label=item.dataset.label;
                                        const val=item.parsed.y;
                                        const note=(val===0&&label.includes('J2'))?' (négatif J2)':(val===0?'':' UFC/m³');
                                        return label+': '+val+note;
                                    }},
                                    footer:items=>{{
                                        const v=Math.max(...items.map(i=>i.parsed.y));
                                        return v>=d.action?'🚨 ACTION':v>=d.alerte?'⚠️ ALERTE':'✅ Conforme';
                                    }}
                                }}}}
                            }},
                            scales:{{
                                x:{{ticks:{{font:{{size:10}},maxRotation:45}}}},
                                y:{{beginAtZero:true,title:{{display:true,text:'UFC/m³',font:{{size:10}}}},ticks:{{font:{{size:10}}}}}}
                            }}
                        }}
                      }});
                    }})();
                    </script>"""
                    st.components.v1.html(evol_html, height=290)
                    st.caption(f"⚠️ Alerte ≥ {seuil_alerte} · 🚨 Action ≥ {seuil_action} · {len(evol_dates)} mesure(s)")
                else:
                    st.info("Aucune donnée datée pour ce point.")

            st.divider()

        # ══════════════════════════════════════════════════════════════════════
        # ONGLET 2 : STATS PAR GERME
        # ══════════════════════════════════════════════════════════════════════
        with hist_tab_germs:
            from collections import defaultdict
            import json as _json_germs

            germs_stats=defaultdict(lambda:{"count":0,"ufc_sum":0,"points":set(),"criticite":0})
            total_pos=0
            for r in surv_f:
                _gd_list = r.get("germs_detail", [])

                if _gd_list:
                    # ── Multi-germes : chaque germe compté individuellement ──
                    for gde in _gd_list:
                        germ    = gde.get("name","")
                        ufc_val = int(gde.get("ufc",0) or 0)
                        if ufc_val > 0 and germ not in ("Négatif","—",""):
                            total_pos += 1
                            germs_stats[germ]["count"]   += 1
                            germs_stats[germ]["ufc_sum"] += ufc_val
                            germs_stats[germ]["points"].add(r.get("prelevement","—"))
                            if germs_stats[germ]["criticite"] == 0:
                                germs_stats[germ]["criticite"] = _get_criticite(germ)
                else:
                    # ── Mono-germe : logique d'origine ───────────────────────
                    germ   = r.get("germ_saisi","") or r.get("germ_match","") or ""
                    ufc_j2 = int(r.get("ufc_48h", r.get("ufc",0)) or 0)
                    ufc_j7 = int(r.get("ufc_5j",  r.get("ufc",0)) or 0)
                    j2_pos = ufc_j2 > 0 and germ not in ("Négatif","—","")
                    j7_pos = ufc_j7 > 0 and germ not in ("Négatif","—","")
                    if j2_pos:
                        total_pos += 1
                        germs_stats[germ]["count"]   += 1
                        germs_stats[germ]["ufc_sum"] += ufc_j2
                        germs_stats[germ]["points"].add(r.get("prelevement","—"))
                        if germs_stats[germ]["criticite"] == 0:
                            germs_stats[germ]["criticite"] = _get_criticite(germ)
                    elif not j2_pos and j7_pos:
                        total_pos += 1
                        germs_stats[germ]["count"]   += 1
                        germs_stats[germ]["ufc_sum"] += ufc_j7
                        germs_stats[germ]["points"].add(r.get("prelevement","—"))
                        if germs_stats[germ]["criticite"] == 0:
                            germs_stats[germ]["criticite"] = _get_criticite(germ)

            if not germs_stats:
                st.info("Aucun germe positif dans l'historique.")
            else:
                sorted_germs=sorted(germs_stats.items(),key=lambda x:-x[1]["count"])
                palette=["#ef4444","#f97316","#f59e0b","#84cc16","#22c55e",
                         "#06b6d4","#6366f1","#a855f7","#ec4899","#14b8a6"]
                g_labels=[g[:28] for g,_ in sorted_germs]
                g_counts=[d["count"] for _,d in sorted_germs]
                g_colors=[palette[i%len(palette)] for i in range(len(g_labels))]
                gchart_html=f"""
                <script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.0/chart.umd.min.js"></script>
                <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:12px;
                            padding:16px;margin-bottom:18px">
                  <div style="font-size:.8rem;font-weight:700;color:#1e40af;margin-bottom:10px">
                    📊 Distribution des germes positifs
                  </div>
                  <div style="width:100%;max-width:400px;height:280px;margin:0 auto">
                    <canvas id="germDoughnut"></canvas>
                  </div>
                </div>
                <script>
                (function(){{
                  const d={_json_germs.dumps({"labels":g_labels,"counts":g_counts,"colors":g_colors})};
                  new Chart(document.getElementById('germDoughnut'),{{
                    type:'doughnut',
                    data:{{labels:d.labels,datasets:[{{data:d.counts,backgroundColor:d.colors,borderWidth:2}}]}},
                    options:{{responsive:true,maintainAspectRatio:false,
                      plugins:{{legend:{{position:'bottom',labels:{{font:{{size:11}},boxWidth:14,padding:10}}}}}}}}
                  }});
                }})();
                </script>"""
                st.components.v1.html(gchart_html, height=360)
                st.markdown(
                    "<div style='display:grid;"
                    "grid-template-columns:2fr 0.6fr 1fr 0.9fr 0.9fr 1.5fr;"
                    "gap:4px;background:#1e40af;border-radius:10px 10px 0 0;padding:10px 14px'>"
                    "<div style='font-size:.72rem;font-weight:800;color:#fff'>Germe</div>"
                    "<div style='font-size:.72rem;font-weight:800;color:#fff;text-align:center'>Cas</div>"
                    "<div style='font-size:.72rem;font-weight:800;color:#fff;text-align:center'>% positifs</div>"
                    "<div style='font-size:.72rem;font-weight:800;color:#fff;text-align:center'>Moy. UFC</div>"
                    "<div style='font-size:.72rem;font-weight:800;color:#fff;text-align:center'>Criticité</div>"
                    "<div style='font-size:.72rem;font-weight:800;color:#fff'>Points touchés</div>"
                    "</div>",
                    unsafe_allow_html=True)
                for gi,(gname,gdata) in enumerate(sorted_germs):
                    pct=gdata["count"]/total_pos*100 if total_pos>0 else 0
                    avg_ufc=gdata["ufc_sum"]/gdata["count"] if gdata["count"]>0 else 0
                    pts_str=", ".join(list(gdata["points"])[:3])
                    crit=gdata["criticite"]; cc=_crit_color(crit)
                    row_bg="#f8fafc" if gi%2==0 else "#ffffff"
                    crit_cell=(
                        "<span style='background:"+cc+"22;color:"+cc+";border:1px solid "+cc+"55;"
                        "border-radius:5px;padding:1px 7px;font-size:.72rem;font-weight:700'>"
                        +str(crit)+" – "+_crit_label(crit)+"</span>"
                        if crit>0 else "<span style='color:#94a3b8'>—</span>")
                    st.markdown(
                        "<div style='display:grid;"
                        "grid-template-columns:2fr 0.6fr 1fr 0.9fr 0.9fr 1.5fr;"
                        "gap:4px;background:"+row_bg+";border:1px solid #e2e8f0;border-top:none;"
                        "padding:9px 14px;align-items:center'>"
                        "<div style='font-size:.88rem;font-weight:700;color:#0f172a'>🦠 "+gname+"</div>"
                        "<div style='font-size:1rem;font-weight:800;color:#1e40af;text-align:center'>"+str(gdata["count"])+"</div>"
                        "<div style='text-align:center'>"
                        "<div style='background:#e2e8f0;border-radius:4px;height:8px;margin-bottom:2px'>"
                        "<div style='background:#ef4444;border-radius:4px;height:8px;width:"+str(int(pct))+"%'></div></div>"
                        "<span style='font-size:.75rem;font-weight:700;color:#ef4444'>"+str(round(pct,1))+"%</span></div>"
                        "<div style='font-size:.85rem;font-weight:700;color:#475569;text-align:center'>"+str(round(avg_ufc))+"</div>"
                        "<div style='text-align:center'>"+crit_cell+"</div>"
                        "<div style='font-size:.72rem;color:#475569;font-style:italic'>"+pts_str+"</div>"
                        "</div>",
                        unsafe_allow_html=True)
                st.divider()
                st.markdown(
                    "<div style='font-size:.85rem;font-weight:700;color:#1e40af;margin-bottom:8px'>"
                    "🔧 Suivi des mesures correctives par germe</div>",
                    unsafe_allow_html=True,
                )

                from collections import defaultdict as _dfd
                mc_by_germ = _dfd(lambda: {"en_attente": 0, "fait": 0, "details": []})

                for r in surv_f:
                    germ   = r.get("germ_saisi", "") or r.get("germ_match","") or ""
                    status = r.get("status", "ok")
                    mc     = r.get("mc_statut", "")
                    if status not in ("alert", "action") or germ in ("Négatif", "—", ""):
                        continue
                    if mc == "fait":
                        mc_by_germ[germ]["fait"] += 1
                        if r.get("mc_detail", "").strip():
                            mc_by_germ[germ]["details"].append({
                                "date":    r.get("date", "—"),
                                "point":   r.get("prelevement", "—"),
                                "detail":  r.get("mc_detail", ""),
                                "mc_date": r.get("mc_date", ""),
                            })
                    else:
                        mc_by_germ[germ]["en_attente"] += 1

                if not mc_by_germ:
                    st.info("Aucun dépassement de seuil sur la période.")
                else:
                    for germ_mc, gd in sorted(mc_by_germ.items(),
                                              key=lambda x: -(x[1]["en_attente"] + x[1]["fait"])):
                        total_mc = gd["en_attente"] + gd["fait"]
                        pct_fait = gd["fait"] / total_mc * 100 if total_mc > 0 else 0
                        with st.expander(
                            f"🦠 {germ_mc}  —  "
                            f"{'✅' if gd['en_attente'] == 0 else '⚠️'} "
                            f"{gd['fait']}/{total_mc} mesures validées "
                            f"({round(pct_fait)}%)",
                            expanded=(gd["en_attente"] > 0),
                        ):
                            ca, cb = st.columns(2)
                            ca.metric("✅ Mesures faites", gd["fait"])
                            cb.metric("⏳ En attente",     gd["en_attente"])
                            if gd["details"]:
                                st.markdown(
                                    "<div style='font-size:.78rem;font-weight:700;"
                                    "color:#0369a1;margin:10px 0 6px'>📝 Actions réalisées</div>",
                                    unsafe_allow_html=True,
                                )
                                for det in gd["details"]:
                                    st.markdown(
                                        f"<div style='background:#f0f9ff;border-left:3px solid #0ea5e9;"
                                        f"border-radius:6px;padding:8px 12px;margin-bottom:6px;"
                                        f"font-size:.78rem'>"
                                        f"<span style='color:#0369a1;font-weight:700'>"
                                        f"📍 {det['point']} · {det['date']}</span>"
                                        f"<span style='color:#64748b;font-size:.7rem'>"
                                        f" — validé le {det['mc_date']}</span><br>"
                                        f"<span style='color:#0f172a'>{det['detail']}</span>"
                                        f"</div>",
                                        unsafe_allow_html=True,
                                    )

                st.markdown(
                    "<div style='background:#1e293b;border-radius:0 0 10px 10px;padding:8px 14px'>"
                    "<div style='font-size:.78rem;color:#94a3b8'>"
                    +str(len(germs_stats))+" germe(s) — "+str(total_pos)+" positifs"
                    "</div></div>",
                    unsafe_allow_html=True)

        # ══════════════════════════════════════════════════════════════════════
        # ONGLET 3 : RÉPARTITION PAR PRÉLEVEUR
        # ══════════════════════════════════════════════════════════════════════
        with hist_tab_prev:
            from collections import defaultdict

            prev_stats=defaultdict(lambda:{
                "total":0,"positives":0,"negatives":0,
                "alertes":0,"actions":0,"germes":defaultdict(int)
            })
            for r in surv_f:
                op     = (r.get("operateur","") or "Non renseigné").strip() or "Non renseigné"
                germ   = (r.get("germ_saisi","") or r.get("germ_match","") or "").strip()
                st_r   = r.get("status","ok")
                ufc_j2 = int(r.get("ufc_48h", r.get("ufc",0)) or 0)
                ufc_j7 = int(r.get("ufc_5j",  r.get("ufc",0)) or 0)

                is_positive = (
                    (ufc_j2 > 0 or ufc_j7 > 0)
                    and germ not in ("Négatif","—","")
                )

                prev_stats[op]["total"] += 1
                if is_positive:
                    prev_stats[op]["positives"]    += 1
                    prev_stats[op]["germes"][germ] += 1
                else:
                    prev_stats[op]["negatives"]    += 1

                if st_r == "alert":    prev_stats[op]["alertes"] += 1
                elif st_r == "action": prev_stats[op]["actions"] += 1

            op_list=sorted(prev_stats.items(),key=lambda x:-x[1]["total"])
            card_cols=st.columns(min(len(op_list),4))
            for ci,(op_name,op_data) in enumerate(op_list):
                t=op_data["total"]; pos=op_data["positives"]
                taux_pos=pos/t*100 if t>0 else 0
                tc="#ef4444" if taux_pos>=30 else "#f59e0b" if taux_pos>0 else "#22c55e"
                ini=op_name[0].upper() if op_name!="Non renseigné" else "?"
                with card_cols[ci%len(card_cols)]:
                    st.markdown(
                        "<div style='background:#fff;border:1.5px solid #e2e8f0;border-radius:14px;"
                        "padding:18px 14px;text-align:center;margin-bottom:12px'>"
                        "<div style='background:#2563eb;color:#fff;border-radius:50%;width:48px;height:48px;"
                        "display:flex;align-items:center;justify-content:center;font-weight:800;"
                        "font-size:1.2rem;margin:0 auto 10px auto'>"+ini+"</div>"
                        "<div style='font-size:.92rem;font-weight:700;color:#0f172a;margin-bottom:6px'>"+op_name+"</div>"
                        "<div style='font-size:2rem;font-weight:900;color:#1e40af'>"+str(t)+"</div>"
                        "<div style='font-size:.68rem;color:#64748b;margin-bottom:10px'>prélèvement(s)</div>"
                        "<div style='display:grid;grid-template-columns:1fr 1fr;gap:6px'>"
                        "<div style='background:#f0fdf4;border-radius:8px;padding:6px'>"
                        "<div style='font-size:1rem;font-weight:800;color:#22c55e'>"+str(op_data["negatives"])+"</div>"
                        "<div style='font-size:.6rem;color:#166534'>✅ Nég.</div></div>"
                        "<div style='background:#fef2f2;border-radius:8px;padding:6px'>"
                        "<div style='font-size:1rem;font-weight:800;color:#ef4444'>"+str(pos)+"</div>"
                        "<div style='font-size:.6rem;color:#991b1b'>🦠 Pos.</div></div>"
                        "<div style='background:#fffbeb;border-radius:8px;padding:6px'>"
                        "<div style='font-size:1rem;font-weight:800;color:#f59e0b'>"+str(op_data["alertes"])+"</div>"
                        "<div style='font-size:.6rem;color:#92400e'>⚠️ Alerte</div></div>"
                        "<div style='background:#fef2f2;border-radius:8px;padding:6px'>"
                        "<div style='font-size:1rem;font-weight:800;color:#dc2626'>"+str(op_data["actions"])+"</div>"
                        "<div style='font-size:.6rem;color:#991b1b'>🚨 Action</div></div></div>"
                        "<div style='margin-top:10px;background:"+tc+"22;border:1px solid "+tc+"55;"
                        "border-radius:8px;padding:5px'>"
                        "<div style='font-size:.8rem;font-weight:800;color:"+tc+"'>"
                        +str(round(taux_pos))+"% positifs</div></div></div>",
                        unsafe_allow_html=True)
            st.divider()
            st.markdown(
                "<div style='display:grid;grid-template-columns:2fr 0.7fr 0.7fr 0.7fr 0.7fr 0.7fr 2fr;"
                "gap:4px;background:#1e40af;border-radius:10px 10px 0 0;padding:10px 14px'>"
                "<div style='font-size:.72rem;font-weight:800;color:#fff'>Préleveur</div>"
                "<div style='font-size:.72rem;font-weight:800;color:#fff;text-align:center'>Total</div>"
                "<div style='font-size:.72rem;font-weight:800;color:#fff;text-align:center'>✅</div>"
                "<div style='font-size:.72rem;font-weight:800;color:#fff;text-align:center'>🦠</div>"
                "<div style='font-size:.72rem;font-weight:800;color:#fff;text-align:center'>⚠️</div>"
                "<div style='font-size:.72rem;font-weight:800;color:#fff;text-align:center'>🚨</div>"
                "<div style='font-size:.72rem;font-weight:800;color:#fff'>Germes fréquents</div>"
                "</div>",
                unsafe_allow_html=True)
            for ri,(op_name,op_data) in enumerate(op_list):
                t=op_data["total"]; pos=op_data["positives"]
                taux_pos=pos/t*100 if t>0 else 0
                tc="#ef4444" if taux_pos>=30 else "#f59e0b" if taux_pos>0 else "#22c55e"
                top_g=", ".join(
                    g+"("+str(n)+"x)"
                    for g,n in sorted(op_data["germes"].items(),key=lambda x:-x[1])[:3]
                ) or "—"
                row_bg="#f8fafc" if ri%2==0 else "#ffffff"
                st.markdown(
                    "<div style='display:grid;grid-template-columns:2fr 0.7fr 0.7fr 0.7fr 0.7fr 0.7fr 2fr;"
                    "gap:4px;background:"+row_bg+";border:1px solid #e2e8f0;border-top:none;"
                    "padding:9px 14px;align-items:center'>"
                    "<div style='font-size:.88rem;font-weight:700;color:#0f172a'>👤 "+op_name+"</div>"
                    "<div style='font-size:1rem;font-weight:800;color:#1e40af;text-align:center'>"+str(t)+"</div>"
                    "<div style='font-size:1rem;font-weight:800;color:#22c55e;text-align:center'>"+str(op_data["negatives"])+"</div>"
                    "<div style='text-align:center'><span style='background:"+tc+"22;color:"+tc+";"
                    "border-radius:6px;padding:2px 8px;font-size:.8rem;font-weight:700'>"+str(pos)+"</span></div>"
                    "<div style='font-size:1rem;font-weight:800;color:#f59e0b;text-align:center'>"+str(op_data["alertes"])+"</div>"
                    "<div style='font-size:1rem;font-weight:800;color:#ef4444;text-align:center'>"+str(op_data["actions"])+"</div>"
                    "<div style='font-size:.72rem;color:#475569;font-style:italic'>"+top_g+"</div>"
                    "</div>",
                    unsafe_allow_html=True)
            st.markdown(
                "<div style='background:#1e293b;border-radius:0 0 10px 10px;padding:8px 14px'>"
                "<div style='font-size:.78rem;color:#94a3b8'>"
                +str(len(op_list))+" préleveur(s)</div></div>",
                unsafe_allow_html=True)

        # ══════════════════════════════════════════════════════════════════════
        # ONGLET 4 : LISTE DES ENTRÉES
        # ══════════════════════════════════════════════════════════════════════
        with hist_tab_liste:
            # ── Filtre par point ──────────────────────────────────────────────
            all_points = sorted(set(r.get("prelevement","—") for r in surv_f))
            selected_points = st.multiselect(
                "📍 Filtrer par point de prélèvement",
                options=all_points,
                default=[],
                placeholder="Tous les points",
                key="liste_filter_points",
            )
            surv_f_liste = [r for r in surv_f if r.get("prelevement","—") in selected_points] \
                           if selected_points else list(surv_f)

            # ── Tri ───────────────────────────────────────────────────────────
            col_tri1, col_tri2 = st.columns([2,1])
            with col_tri1:
                tri_choix = st.selectbox(
                    "Trier par",
                    ["Date (récent → ancien)", "Date (ancien → récent)",
                     "Point A→Z", "Point Z→A", "UFC décroissant"],
                    key="liste_tri",
                    label_visibility="collapsed",
                )
            with col_tri2:
                st.caption(f"{len(surv_f_liste)} entrée(s) affichée(s)")

            if tri_choix == "Date (récent → ancien)":
                surv_f_liste = sorted(surv_f_liste,
                    key=lambda x: _parse_date(x.get("date_prelevement", x.get("date",""))) or dt_date.min, reverse=True)
            elif tri_choix == "Date (ancien → récent)":
                surv_f_liste = sorted(surv_f_liste,
                    key=lambda x: _parse_date(x.get("date_prelevement", x.get("date",""))) or dt_date.min)
            elif tri_choix == "Point A→Z":
                surv_f_liste = sorted(surv_f_liste, key=lambda x: x.get("prelevement",""))
            elif tri_choix == "Point Z→A":
                surv_f_liste = sorted(surv_f_liste, key=lambda x: x.get("prelevement",""), reverse=True)
            elif tri_choix == "UFC décroissant":
                surv_f_liste = sorted(surv_f_liste,
                    key=lambda x: int(x.get("ufc",0) or 0), reverse=True)

            # ── Affichage groupé ou plat ──────────────────────────────────────
            if selected_points:
                grouped = {}
                for r in surv_f_liste:
                    pt = r.get("prelevement","—")
                    grouped.setdefault(pt, []).append(r)

                for pt_name, pt_entries in grouped.items():
                    pt_pos = sum(
                        1 for r in pt_entries
                        if r.get("status") in ("alert","action")
                        or (int(r.get("ufc",0) or 0) > 0
                            and (r.get("germ_saisi","") or r.get("germ_match",""))
                            not in ("Négatif","—",""))
                    )
                    with st.expander(
                        f"📍 {pt_name}  —  {len(pt_entries)} prélèvement(s)  ·  "
                        f"{'🦠 '+str(pt_pos)+' positif(s)' if pt_pos else '✅ Tous conformes'}",
                        expanded=True,
                    ):
                        _render_liste_entries(pt_entries, surv)
            else:
                _render_liste_entries(surv_f_liste, surv)

    else:
        st.info("Aucun prélèvement enregistré.")
# ═══════════════════════════════════════════════════════════════════════════════
# TAB : PARAMÈTRES — COMPLET 
# ═══════════════════════════════════════════════════════════════════════════════

if active == "parametres":
    can_edit = check_access_protege("Paramètres & Seuils")
    if not can_edit:
        st.info("👁️ Mode lecture seule — connectez-vous pour modifier les paramètres.")
    st.markdown("### ⚙️ Paramètres")

    (subtab_mesures,
    subtab_points,
    subtab_seuils,
    subtab_operateurs,
    subtab_backup,
    subtab_supabase,
    subtab_faq) = st.tabs([
        "📋 Mesures correctives",
        "📍 Points de prélèvement",
        "⚖️ Seuils d'alerte",
        "👤 Opérateurs",
        "💾 Sauvegarde",
        "☁️ Base de données",
        "❓ FAQ"
    ])



    # ── Constantes Points ──────────────────────────────────────────────────────
    LOC_CRIT_OPTS = [
        "1 — Limité",
        "2 — Modéré",
        "3 — Important",
        "4 — Critique",
    ]
    LOC_CRIT_COLORS = {"1": "#22c55e", "2": "#0babf5", "3": "#ee811a", "4": "#f50b0b"}
    LOC_CRIT_LABELS = {"1": "Limité", "2": "Modéré", "3": "Important", "4": "Critique"}
    PT_FREQ_UNIT_OPTS = ["/ jour", "/ semaine", "/ mois"]

    def _freq_en_semaine(pt: dict, jours_par_semaine: int = 5) -> float:
        """Convertit la fréquence d'un point de prélèvement en nombre de fois par semaine."""
        freq  = float(pt.get("frequence", 1) or 1)
        unite = (pt.get("frequence_unit") or "/ semaine").strip()

        if unite == "/ jour":
            return freq * jours_par_semaine
        elif unite == "/ semaine":
            return freq
        elif unite == "/ mois":
            return freq / 4.33          # ≈ semaines par mois
        else:
            return freq                 # fallback : on suppose hebdomadaire
# MESURES CORRECTIVES #
    with subtab_mesures:
        om = st.session_state.origin_measures

        # ── Mappings ───────────────────────────────────────────────────────────────
        scope_labels = {
            "all":                          "🌐 Toutes",
            "Air":                          "💨 Air",
            "Humidité":                     "💧 Humidité",
            "Flore fécale":                 "🦠 Flore fécale",
            "Oropharynx / Gouttelettes":   "😷 Oropharynx",
            "Peau / Muqueuse":             "🖐️ Peau / Muqueuse",
            "Sol / Carton / Surface sèche":"📦 Sol / Surface sèche",
        }
        scope_r_map = {v: k for k, v in scope_labels.items()}

        type_labels = {"alert": "⚠️ Alerte", "action": "🚨 Action", "both": "⚠️🚨 Alerte & Action"}
        type_colors = {"alert": "#f59e0b", "action": "#ef4444", "both": "#818cf8"}
        type_r_map  = {v: k for k, v in type_labels.items()}

        risk_opts_map = {
            "all": "🌐 Toutes", "1": "🟢 1", "2": "🟢 2", "3": "🟡 3",
            "4": "🟠 4", "5": "🔴 5", "[3,4,5]": "3-4-5", "[4,5]": "4-5", "[1,2,3]": "1-2-3",
        }
        risk_opts_rev = {v: k for k, v in risk_opts_map.items()}

        risk_display = {
            "all": ("🌐", "#64748b"), "1": ("🟢 1", "#16a34a"), "2": ("🟢 2", "#16a34a"),
            "3": ("🟡 3", "#ca8a04"), "4": ("🟠 4", "#ea580c"), "5": ("🔴 5", "#dc2626"),
        }

        # ── germ_type extrait dynamiquement depuis st.session_state.germs ─────────
        # APRÈS
        _familles_raw = sorted({
            g["path"][1]
            for g in st.session_state.get("germs", [])
            if len(g.get("path", [])) > 1
        }) or ["Bactéries", "Champignons"]  # fallback si germs pas encore chargé
        _fam_key_map   = {"Bactéries": "bacteria", "Champignons": "fungi"}
        _fam_emoji_map = {"Bactéries": "🦠",       "Champignons": "🍄"}
        _fam_color_map = {"bacteria": "#0284c7",   "fungi": "#7c3aed"}

        germ_type_labels = {
            "all":      "🌐 Tous germes",
            "bacteria": "🦠 Bactéries",
            "fungi":    "🍄 Champignons",
            "both":     "🦠🍄 Bactéries & Champignons",
        }
        germ_type_r_map  = {v: k for k, v in germ_type_labels.items()}
        germ_type_colors = {
            "all":      "#64748b",
            "bacteria": "#0284c7",
            "fungi":    "#7c3aed",
            "both":     "#0f766e",
        }

        # ── Filtres ────────────────────────────────────────────────────────────────
        col_f1, col_f2, col_f3, col_f4, col_f5 = st.columns([2, 1.5, 1.5, 1.5, 1])
        with col_f1:
            filter_scope = st.selectbox(
                "Origine", ["Tout afficher"] + list(scope_labels.values()),
                label_visibility="collapsed", key="filter_scope")
        with col_f2:
            filter_risk_lbl = st.selectbox(
                "Criticité", ["🌐 Tout", "🟢 1", "🟢 2", "🟡 3", "🟠 4", "🔴 5"],
                label_visibility="collapsed", key="filter_risk")
        with col_f3:
            filter_type = st.selectbox(
                "Type", ["Tout", "⚠️ Alerte", "🚨 Action"],
                label_visibility="collapsed", key="filter_type")
        with col_f4:
            filter_germ_type = st.selectbox(
                "Germe", list(germ_type_labels.values()),
                label_visibility="collapsed", key="filter_germ_type")
        with col_f5:
            if can_edit:
                if st.button("➕ Nouvelle", use_container_width=True, key="btn_new_mesure"):
                    st.session_state.show_new_measure    = True
                    st.session_state["_edit_mesure_idx"] = None
                    st.rerun()

        active_scope     = scope_r_map.get(filter_scope) if filter_scope != "Tout afficher" else None
        active_risk      = filter_risk_lbl.split()[-1]   if filter_risk_lbl != "🌐 Tout"    else None
        active_type      = ("alert" if "Alerte" in filter_type else "action") if filter_type != "Tout" else None
        active_germ_type = germ_type_r_map.get(filter_germ_type, "all")

        # ── Formulaire nouvelle mesure ─────────────────────────────────────────────
        if can_edit and st.session_state.get("show_new_measure", False):
            st.markdown(
                "<div style='background:#f0fdf4;border:1.5px solid #86efac;"
                "border-radius:10px;padding:16px;margin-bottom:12px'>",
                unsafe_allow_html=True)
            st.markdown("#### ➕ Nouvelle mesure")

            # Ligne 1 : Texte + Origine + Criticité + Type
            nmc1, nmc2, nmc3, nmc4 = st.columns([3, 2, 1.5, 1.5])
            with nmc1:
                nm_text = st.text_input("Texte *", key="nm_text")
            with nmc2:
                nm_scope_label = st.selectbox("Origine", list(scope_labels.values()), key="nm_scope")
                nm_scope = scope_r_map.get(nm_scope_label, "all")
            with nmc3:
                nm_risk_lbl = st.selectbox("Criticité", list(risk_opts_map.values()), key="nm_risk")
                nm_risk_key = risk_opts_rev.get(nm_risk_lbl, "all")
                nm_risk = ("all" if nm_risk_key == "all"
                        else json.loads(nm_risk_key) if nm_risk_key.startswith("[")
                        else int(nm_risk_key))
            with nmc4:
                nm_type_label = st.selectbox("Type", list(type_labels.values()), key="nm_type")
                nm_type = type_r_map.get(nm_type_label, "alert")

            # Ligne 2 : Germe seul (large)
            ng1, ng2 = st.columns([3, 6])
            with ng1:
                nm_germ_lbl  = st.selectbox("Type de germe", list(germ_type_labels.values()), key="nm_germ_type")
                nm_germ_type = germ_type_r_map.get(nm_germ_lbl, "all")

            nb1, nb2 = st.columns(2)
            with nb1:
                if st.button("✅ Ajouter", use_container_width=True, key="nm_submit"):
                    if nm_text.strip():
                        om.append({
                            "id":        f"m{len(om)+1:03d}_custom",
                            "text":      nm_text.strip(),
                            "scope":     nm_scope,
                            "risk":      nm_risk,
                            "type":      nm_type,
                            "germ_type": nm_germ_type,
                        })
                        save_origin_measures(om, supa=False)
                        st.session_state.origin_measures  = om
                        st.session_state.show_new_measure = False
                        st.rerun()
                    else:
                        st.error("Le texte est obligatoire.")
            with nb2:
                if st.button("Annuler", use_container_width=True, key="nm_cancel"):
                    st.session_state.show_new_measure = False
                    st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)

        # ── Fonction filtre ────────────────────────────────────────────────────────
        def _passes_filter(m):
            if active_scope and m.get("scope") != active_scope:
                return False
            if active_type:
                mt = m.get("type", "alert")
                if mt != "both" and mt != active_type:
                    return False
            if active_risk:
                mr = m.get("risk", "all")
                if mr != "all":
                    if isinstance(mr, list):
                        if int(active_risk) not in mr:
                            return False
                    else:
                        if str(mr) != active_risk:
                            return False
            if active_germ_type != "all":
                mgt = m.get("germ_type", "all")
                if mgt != "all" and mgt != "both" and mgt != active_germ_type:
                    return False
            return True

        # ── Bandeau modifications non sauvegardées ─────────────────────────────────
        if st.session_state.get("_mesures_modifiees"):
            st.markdown(
                "<div style='background:#fffbeb;border:1.5px solid #fcd34d;border-radius:8px;"
                "padding:8px 14px;margin-bottom:10px;font-size:.78rem;color:#92400e'>"
                "⚠️ Modifications non sauvegardées — cliquez sur <strong>💾 Sauvegarder</strong>."
                "</div>", unsafe_allow_html=True)

        # ── Liste des mesures ──────────────────────────────────────────────────────
        for real_idx, m in enumerate(om):
            if not _passes_filter(m):
                continue

            tcol  = type_colors.get(m.get("type", "alert"), "#0f172a")
            tlbl  = type_labels.get(m.get("type", "alert"), m.get("type", ""))
            gtkey = m.get("germ_type", "all")
            gtlbl = germ_type_labels.get(gtkey, "🌐 Tous germes")
            gtcol = germ_type_colors.get(gtkey, "#64748b")

            # Criticité badge
            mr = m.get("risk", "all")
            if mr == "all":
                risk_lbl, risk_col = "🌐", "#64748b"
            elif isinstance(mr, list):
                risk_lbl, risk_col = "-".join(str(x) for x in mr), "#94a3b8"
            else:
                _rd = risk_display.get(str(mr), ("?", "#94a3b8"))
                risk_lbl, risk_col = _rd

            # ── Mode édition ───────────────────────────────────────────────────────
            if st.session_state.get("_edit_mesure_idx") == real_idx:
                st.markdown(
                    "<div style='background:#eff6ff;border:1.5px solid #93c5fd;"
                    "border-radius:10px;padding:14px;margin-bottom:8px'>",
                    unsafe_allow_html=True)
                st.markdown("**✏️ Modifier la mesure**")

                ec1, ec2, ec3, ec4, ec5 = st.columns([3, 2, 1.5, 1.5, 1.5])
                with ec1:
                    new_text = st.text_input("Texte *", value=m.get("text", ""), key=f"em_text_{real_idx}")
                with ec2:
                    cur_scope_lbl = scope_labels.get(m.get("scope", "all"), "🌐 Toutes")
                    scope_opts    = list(scope_labels.values())
                    new_scope_lbl = st.selectbox(
                        "Origine", scope_opts,
                        index=scope_opts.index(cur_scope_lbl) if cur_scope_lbl in scope_opts else 0,
                        key=f"em_scope_{real_idx}")
                    new_scope = scope_r_map.get(new_scope_lbl, "all")
                with ec3:
                    cur_risk     = m.get("risk", "all")
                    cur_risk_key = (str(cur_risk) if not isinstance(cur_risk, list)
                                    else json.dumps(cur_risk, separators=(',', ':')))
                    cur_risk_lbl  = risk_opts_map.get(cur_risk_key, "🌐 Toutes")
                    risk_opts_list = list(risk_opts_map.values())
                    new_risk_lbl  = st.selectbox(
                        "Criticité", risk_opts_list,
                        index=risk_opts_list.index(cur_risk_lbl) if cur_risk_lbl in risk_opts_list else 0,
                        key=f"em_risk_{real_idx}")
                    new_risk_key = risk_opts_rev.get(new_risk_lbl, "all")
                    new_risk     = ("all"                   if new_risk_key == "all"
                                    else json.loads(new_risk_key) if new_risk_key.startswith("[")
                                    else int(new_risk_key))
                with ec4:
                    cur_type_lbl = type_labels.get(m.get("type", "alert"), "⚠️ Alerte")
                    type_opts    = list(type_labels.values())
                    new_type_lbl = st.selectbox(
                        "Type", type_opts,
                        index=type_opts.index(cur_type_lbl) if cur_type_lbl in type_opts else 0,
                        key=f"em_type_{real_idx}")
                    new_type = type_r_map.get(new_type_lbl, "alert")
                with ec5:
                    cur_gt_lbl = germ_type_labels.get(m.get("germ_type", "all"), "🌐 Tous germes")
                    gt_opts    = list(germ_type_labels.values())
                    new_gt_lbl = st.selectbox(
                        "Germe", gt_opts,
                        index=gt_opts.index(cur_gt_lbl) if cur_gt_lbl in gt_opts else 0,
                        key=f"em_germ_type_{real_idx}")
                    new_germ_type = germ_type_r_map.get(new_gt_lbl, "all")

                sb1, sb2 = st.columns(2)
                with sb1:
                    if st.button("✔️ Valider", key=f"em_save_{real_idx}",
                                use_container_width=True, type="primary"):
                        if new_text.strip():
                            om[real_idx].update({
                                "text":      new_text.strip(),
                                "scope":     new_scope,
                                "risk":      new_risk,
                                "type":      new_type,
                                "germ_type": new_germ_type,
                            })
                            save_origin_measures(om, supa=False)
                            st.session_state.origin_measures       = om
                            st.session_state["_edit_mesure_idx"]   = None
                            st.session_state["_mesures_modifiees"] = True
                            st.rerun()
                        else:
                            st.error("Le texte est obligatoire.")
                with sb2:
                    if st.button("✕ Annuler", key=f"em_cancel_{real_idx}", use_container_width=True):
                        st.session_state["_edit_mesure_idx"] = None
                        st.rerun()
                st.markdown("</div>", unsafe_allow_html=True)

            # ── Mode lecture ───────────────────────────────────────────────────────
            else:
                rc1, rc2, rc3, rc4, rc5, rc6 = st.columns([4, 1.2, 1.2, 1.2, 0.8, 0.8])
                with rc1:
                    st.markdown(
                        f'<div style="padding:6px 0;font-size:.8rem;color:#1e293b">• {m.get("text","")}</div>',
                        unsafe_allow_html=True)
                with rc2:
                    st.markdown(
                        f'<div style="padding:6px 0;font-size:.65rem;color:{gtcol};'
                        f'font-weight:600;text-align:center">{gtlbl}</div>',
                        unsafe_allow_html=True)
                with rc3:
                    st.markdown(
                        f'<div style="padding:6px 0;font-size:.65rem;color:{tcol};'
                        f'font-weight:600;text-align:center">{tlbl}</div>',
                        unsafe_allow_html=True)
                with rc4:
                    st.markdown(
                        f'<div style="padding:6px 0;font-size:.65rem;color:{risk_col};'
                        f'font-weight:600;text-align:center">{risk_lbl}</div>',
                        unsafe_allow_html=True)
                with rc5:
                    if can_edit:
                        if st.button("✏️", key=f"edit_btn_{real_idx}"):
                            st.session_state["_edit_mesure_idx"] = real_idx
                            st.session_state["show_new_measure"] = False
                            st.rerun()
                with rc6:
                    if can_edit:
                        if st.button("🗑️", key=f"del_m_{real_idx}"):
                            om.pop(real_idx)
                            save_origin_measures(om, supa=False)
                            st.session_state.origin_measures       = om
                            st.session_state["_mesures_modifiees"] = True
                            st.rerun()

        # ── Boutons sauvegarde / réinitialisation ──────────────────────────────────
        col_sr, col_def = st.columns(2)
        with col_sr:
            if can_edit:
                if st.button("💾 Sauvegarder", use_container_width=True, key="save_mesures"):
                    save_origin_measures(om, supa=True)
                    st.session_state["_mesures_modifiees"] = False
                    st.success("✅ Mesures sauvegardées et synchronisées !")
        with col_def:
            if can_edit:
                if st.button("↩️ Réinitialiser", use_container_width=True, key="reinit_mesures"):
                    st.session_state.origin_measures = [dict(m) for m in DEFAULT_ORIGIN_MEASURES]
                    save_origin_measures(st.session_state.origin_measures, supa=True)
                    st.session_state["_mesures_modifiees"] = False
                    st.rerun()
    # ══════════════════════════════════════════════════════════════════════════
    # POINTS DE PRÉLÈVEMENT
    # ══════════════════════════════════════════════════════════════════════════
    with subtab_points:
        st.markdown("""
        <div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:10px;
        padding:12px 16px;margin-bottom:16px;font-size:.82rem;color:#1e40af">
        ℹ️ Le <strong>niveau de criticité du lieu</strong> (1–3) est automatiquement repris
        lors de l'identification microbiologique.<br>
        Score total = criticité lieu × score germe · ⚠️ Alerte : 16–24 · 🚨 Action : &gt; 24
        </div>""", unsafe_allow_html=True)

        if not st.session_state.points:
            st.info("Aucun point défini.")
        else:
            st.markdown("""
            <div style="display:grid;
            grid-template-columns:2.2fr 0.7fr 0.7fr 1.3fr 0.9fr 1.1fr 0.5fr 0.5fr;
            gap:4px;background:#1e40af;border-radius:10px 10px 0 0;padding:10px 14px">
              <div style="font-size:.72rem;font-weight:800;color:#fff">Point</div>
              <div style="font-size:.72rem;font-weight:800;color:#fff;text-align:center">Type</div>
              <div style="font-size:.72rem;font-weight:800;color:#fff;text-align:center">Classe</div>
              <div style="font-size:.72rem;font-weight:800;color:#fff;text-align:center">Criticité lieu</div>
              <div style="font-size:.72rem;font-weight:800;color:#fff;text-align:center">Gélose</div>
              <div style="font-size:.72rem;font-weight:800;color:#fff;text-align:center">Fréquence</div>
              <div></div><div></div>
            </div>""", unsafe_allow_html=True)

            for i, pt in enumerate(list(st.session_state.points)):
                pt_type    = pt.get('type', '—')
                type_icon  = "💨" if pt_type == "Air" else "🧴"
                loc_crit   = str(pt.get('location_criticality', 1))
                lc_color   = LOC_CRIT_COLORS.get(loc_crit, "#94a3b8")
                lc_label   = LOC_CRIT_LABELS.get(loc_crit, "—")
                room_cl    = pt.get('room_class', '—') or '—'
                freq       = pt.get('frequency', 1)
                freq_unit  = pt.get('frequency_unit', '/ semaine')
                freq_short = (str(freq) + "x/" +
                              ("j" if "jour" in freq_unit else
                               "sem" if "sem" in freq_unit else "mois"))
                row_bg = "#f8fafc" if i % 2 == 0 else "#ffffff"

                c1, c2 = st.columns([8, 1])
                with c1:
                    st.markdown(
                        f"<div style='display:grid;"
                        f"grid-template-columns:2.2fr 0.7fr 0.7fr 1.3fr 0.9fr 1.1fr;"
                        f"gap:4px;background:{row_bg};border:1px solid #e2e8f0;"
                        f"border-top:none;padding:9px 14px;align-items:center'>"
                        f"<div style='font-size:.88rem;font-weight:700;color:#0f172a'>"
                        f"{type_icon} {pt['label']}</div>"
                        f"<div style='font-size:.75rem;color:#475569;text-align:center'>{pt_type}</div>"
                        f"<div style='text-align:center'>"
                        f"<span style='background:#dbeafe;color:#1e40af;"
                        f"border:1px solid #93c5fd;border-radius:6px;"
                        f"padding:2px 8px;font-size:.78rem;font-weight:800'>{room_cl}</span></div>"
                        f"<div style='text-align:center'>"
                        f"<span style='background:{lc_color}22;color:{lc_color};"
                        f"border:1px solid {lc_color}55;border-radius:6px;"
                        f"padding:3px 8px;font-size:.68rem;font-weight:700'>"
                        f"Nv.{loc_crit} — {lc_label}</span></div>"
                        f"<div style='font-size:.72rem;color:#1d4ed8;text-align:center'>"
                        f"🧫 {pt.get('gelose', '—')[:12]}</div>"
                        f"<div style='text-align:center'>"
                        f"<span style='background:#eff6ff;color:#1e40af;"
                        f"border:1px solid #bfdbfe;border-radius:6px;"
                        f"padding:2px 8px;font-size:.75rem;font-weight:700'>"
                        f"🔁 {freq_short}</span></div>"
                        f"</div>", unsafe_allow_html=True)
                with c2:
                    be, bd = st.columns(2)
                    with be:
                        if can_edit:
                            if st.button("✏️", key=f"edit_pt_{i}"):
                                st.session_state._edit_point = i
                                st.rerun()
                    with bd:
                        if can_edit:
                            if st.button("🗑️", key=f"del_pt_{i}"):
                                st.session_state.points.pop(i)
                                save_points(st.session_state.points, supa=True)
                                st.rerun()

            st.markdown(
                f"<div style='background:#1e293b;border-radius:0 0 10px 10px;"
                f"padding:8px 14px;margin-bottom:16px'>"
                f"<div style='font-size:.78rem;font-weight:700;color:#94a3b8'>"
                f"{len(st.session_state.points)} point(s)</div></div>",
                unsafe_allow_html=True)

        st.divider()

        # ── Formulaire édition ────────────────────────────────────────────────
        if st.session_state.get('_edit_point') is not None:
            idx = st.session_state._edit_point
            pt  = st.session_state.points[idx]
            st.markdown(f"### ✏️ Modifier — {pt['label']}")

            er1, er2, er3, er_room = st.columns([3, 1.5, 1.5, 1.5])
            with er1:
                new_label = st.text_input("Nom", value=pt['label'], key="pt_edit_label")
            with er2:
                new_type = st.selectbox(
                    "Type", ["Air", "Surface"],
                    index=["Air", "Surface"].index(pt.get('type', 'Air'))
                          if pt.get('type', 'Air') in ["Air", "Surface"] else 0,
                    key="pt_edit_type")
            with er3:
                # ── CORRECTION : selectbox criticité lieu (édition) ──
                cur_lc_val = str(pt.get('location_criticality', 1))
                lc_edit_idx = next(
                    (i for i, o in enumerate(LOC_CRIT_OPTS) if o.startswith(cur_lc_val)), 0)
                new_lc_label = st.selectbox(
                    "Criticité lieu",
                    LOC_CRIT_OPTS,
                    index=lc_edit_idx,
                    key="pt_edit_loc_crit")
                new_lc = int(new_lc_label[0])          # extrait "1", "2" ou "3"
            with er_room:
                new_room = st.text_input(
                    "Classe ISO / GMP",
                    value=pt.get('room_class', ''),
                    placeholder="Ex: A, B, C, D…",
                    key="pt_edit_room")

            # ── Poste type si Classe A (édition) ─────────────────────────────
            if new_room and new_room.strip().upper() == "A":
                st.markdown(
                    "<div style='background:#fef9c3;border:1px solid #fde047;"
                    "border-radius:8px;padding:10px 14px;margin:6px 0'>"
                    "<div style='font-size:.7rem;font-weight:700;color:#854d0e;margin-bottom:6px'>"
                    "🔬 Configuration poste — Classe A</div>",
                    unsafe_allow_html=True)
                _cur_ptype = pt.get("poste_type", "commun")
                _cur_index = 0 if _cur_ptype == "commun" else 1
                new_poste_type = st.radio(
                    "Type de poste *",
                    ["commun", "specifique"],
                    format_func=lambda x: (
                        "🔵 Poste commun"
                        if x == "commun"
                        else "🔀 Poste spécifique (alternance Poste 1 / Poste 2)"
                    ),
                    index=_cur_index,
                    key="pt_edit_poste_type")
                st.markdown("</div>", unsafe_allow_html=True)
            else:
                new_poste_type = "non_applicable"

            er4, er5, er6 = st.columns([2, 1, 2])
            with er4:
                g_opts = (["Gélose de sédimentation", "Gélose TSA", "Gélose Columbia", "Autre"]
                          if new_type == "Air"
                          else ["Gélose contact TSA", "Ecouvillonnage", "Autre"])
                cur_g  = pt.get('gelose', g_opts[0])
                g_idx  = g_opts.index(cur_g) if cur_g in g_opts else 0
                new_gel = st.selectbox("Gélose", g_opts, index=g_idx, key="pt_edit_gelose")
            with er5:
                new_freq = st.number_input(
                    "🔁 Fréquence", min_value=1, max_value=31,
                    value=int(pt.get('frequency', 1)), step=1, key="pt_edit_freq")
            with er6:
                cur_unit = pt.get('frequency_unit', '/ semaine')
                unit_idx = PT_FREQ_UNIT_OPTS.index(cur_unit) if cur_unit in PT_FREQ_UNIT_OPTS else 1
                new_fu   = st.selectbox("Unité", PT_FREQ_UNIT_OPTS, index=unit_idx, key="pt_edit_freq_unit")

            # Aperçu grille seuils (utilise new_lc défini ci-dessus)
            st.markdown(f"""
            <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;
            padding:10px 14px;margin-top:6px">
              <div style="font-size:.65rem;color:#475569;text-transform:uppercase;
              font-weight:700;margin-bottom:8px">
                Grille d'alerte — criticité lieu {new_lc} (score = lieu × germe)
              </div>
              <div style="display:flex;gap:8px">
                <div style="flex:1;background:#f0fdf4;border-radius:6px;padding:8px;
                text-align:center;border:1px solid #86efac">
                  <div style="font-size:.6rem;color:#166534;font-weight:700">✅ Conforme</div>
                  <div style="font-size:.78rem;color:#166534;font-weight:800;margin-top:2px">Score &lt; 16</div>
                  <div style="font-size:.58rem;color:#94a3b8;margin-top:2px">
                    Germe ≤ {int(15/new_lc)}</div>
                </div>
                <div style="flex:1;background:#fffbeb;border-radius:6px;padding:8px;
                text-align:center;border:1px solid #fcd34d">
                  <div style="font-size:.6rem;color:#92400e;font-weight:700">⚠️ Alerte</div>
                  <div style="font-size:.78rem;color:#92400e;font-weight:800;margin-top:2px">Score 16–24</div>
                  <div style="font-size:.58rem;color:#94a3b8;margin-top:2px">
                    Germe {round(16/new_lc,1)}–{round(24/new_lc,1)}</div>
                </div>
                <div style="flex:1;background:#fef2f2;border-radius:6px;padding:8px;
                text-align:center;border:1px solid #fca5a5">
                  <div style="font-size:.6rem;color:#991b1b;font-weight:700">🚨 Action</div>
                  <div style="font-size:.78rem;color:#dc2626;font-weight:800;margin-top:2px">Score &gt; 24</div>
                  <div style="font-size:.58rem;color:#94a3b8;margin-top:2px">
                    Germe &gt; {round(24/new_lc,1)}</div>
                </div>
              </div>
            </div>""", unsafe_allow_html=True)

            eb1, eb2 = st.columns(2)
            with eb1:
                if st.button("✅ Enregistrer", key="pt_save_edit"):
                    _edit_pt_poste = (
                        new_poste_type
                        if new_room and new_room.strip().upper() == "A"
                        else "non_applicable"
                    )
                    st.session_state.points[idx] = {
                        "id":                   pt.get('id', f"p{idx+1}"),
                        "label":                new_label,
                        "type":                 new_type,
                        "gelose":               new_gel,
                        "location_criticality": new_lc,
                        "frequency":            new_freq,
                        "frequency_unit":       new_fu,
                        "room_class":           new_room.strip(),
                        "poste_type":           _edit_pt_poste,
                    }
                    save_points(st.session_state.points, supa=True)
                    st.session_state._edit_point = None
                    st.success("✅ Point mis à jour")
                    st.rerun()
            with eb2:
                if st.button("Annuler", key="pt_cancel_edit"):
                    st.session_state._edit_point = None
                    st.rerun()

        # ── Formulaire ajout ──────────────────────────────────────────────────
        elif can_edit:
            st.markdown("### ➕ Ajouter un point de prélèvement")

            np1, np2, np3, np_room_col = st.columns([3, 1.5, 1.5, 1.5])
            with np1:
                np_label = st.text_input(
                    "Nom *", placeholder="Ex: Salle 3 — Poste A", key="np_label")
            with np2:
                np_type = st.selectbox("Type", ["Air", "Surface"], key="np_type")
            with np3:
                # ── CORRECTION : selectbox criticité lieu (ajout) ──
                np_lc_label = st.selectbox(
                    "Criticité lieu",
                    LOC_CRIT_OPTS,
                    index=0,
                    key="np_loc_crit")
                np_lc = int(np_lc_label[0])             # extrait "1", "2" ou "3"
            with np_room_col:
                np_room = st.text_input(
                    "Classe ISO / GMP", placeholder="Ex: A, B, C, D…", key="np_room")

            # ── Poste type si Classe A ────────────────────────────────────────
            if np_room and np_room.strip().upper() == "A":
                st.markdown(
                    "<div style='background:#fef9c3;border:1px solid #fde047;"
                    "border-radius:8px;padding:10px 14px;margin:6px 0'>"
                    "<div style='font-size:.7rem;font-weight:700;color:#854d0e;margin-bottom:6px'>"
                    "🔬 Configuration poste — Classe A</div>",
                    unsafe_allow_html=True)
                np_poste_type = st.radio(
                    "Type de poste *",
                    ["commun", "specifique"],
                    format_func=lambda x: (
                        "🔵 Poste commun (un seul poste, identique chaque jour)"
                        if x == "commun"
                        else "🔀 Poste spécifique (alternance Poste 1 / Poste 2 chaque jour)"
                    ),
                    key="np_poste_type")
                st.markdown("</div>", unsafe_allow_html=True)
            else:
                np_poste_type = "non_applicable"

            np4, np5, np6 = st.columns([2, 1, 2])
            with np4:
                g_opts_new = (["Gélose de sédimentation", "Gélose TSA", "Gélose Columbia", "Autre"]
                              if np_type == "Air"
                              else ["Gélose contact TSA", "Ecouvillonnage", "Autre"])
                np_gel = st.selectbox("Gélose", g_opts_new, key="np_gelose")
            with np5:
                np_freq = st.number_input(
                    "🔁 Fréquence", min_value=1, max_value=31, value=1, step=1, key="np_freq")
            with np6:
                np_fu = st.selectbox("Unité", PT_FREQ_UNIT_OPTS, index=0, key="np_freq_unit")
            

            # Aperçu grille (utilise np_lc défini ci-dessus)
            st.markdown(f"""
            <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;
            padding:10px 14px;margin-top:4px;margin-bottom:10px">
              <div style="font-size:.65rem;color:#475569;text-transform:uppercase;
              font-weight:700;margin-bottom:8px">
                Aperçu grille (criticité lieu {np_lc} × score germe)
              </div>
              <div style="display:flex;gap:8px">
                <div style="flex:1;background:#f0fdf4;border-radius:6px;padding:7px;
                text-align:center;border:1px solid #86efac">
                  <div style="font-size:.6rem;color:#166534;font-weight:700">✅ Conforme</div>
                  <div style="font-size:.72rem;color:#166534;font-weight:800">Score &lt; 16</div>
                </div>
                <div style="flex:1;background:#fffbeb;border-radius:6px;padding:7px;
                text-align:center;border:1px solid #fcd34d">
                  <div style="font-size:.6rem;color:#92400e;font-weight:700">⚠️ Alerte</div>
                  <div style="font-size:.72rem;color:#92400e;font-weight:800">Score 16–24</div>
                </div>
                <div style="flex:1;background:#fef2f2;border-radius:6px;padding:7px;
                text-align:center;border:1px solid #fca5a5">
                  <div style="font-size:.6rem;color:#991b1b;font-weight:700">🚨 Action</div>
                  <div style="font-size:.72rem;color:#dc2626;font-weight:800">Score &gt; 24</div>
                </div>
              </div>
            </div>""", unsafe_allow_html=True)

            if st.button("➕ Ajouter", key="np_add"):
                if not np_label.strip():
                    st.error("Le nom est requis")
                else:
                    nid = f"p{len(st.session_state.points)+1}_{int(datetime.now().timestamp())}"
                    _save_pt = (
                        np_poste_type
                        if np_room and np_room.strip().upper() == "A"
                        else "non_applicable"
                    )
                    st.session_state.points.append({
                        "id":                   nid,
                        "label":                np_label.strip(),
                        "type":                 np_type,
                        "gelose":               np_gel,
                        "location_criticality": np_lc,
                        "frequency":            np_freq,
                        "frequency_unit":       np_fu,
                        "room_class":           np_room.strip(),
                        "poste_type":           _save_pt,
                    })
                    save_points(st.session_state.points, supa=True)
                    st.success(f"✅ Point **{np_label}** ajouté")
                    st.rerun()


   # ══════════════════════════════════════════════════════════════════════════
    # SEUILS D'ALERTE ET D'ACTION
    # ══════════════════════════════════════════════════════════════════════════
    with subtab_seuils:
        st.markdown("### ⚖️ Seuils d'alerte et d'action")

        st.markdown("""
        <div style="background:linear-gradient(135deg,#eff6ff,#dbeafe);border:1.5px solid #93c5fd;
        border-radius:14px;padding:20px 24px;margin-bottom:20px">
        <div style="font-size:1rem;font-weight:800;color:#1e40af;margin-bottom:14px">
            🧮 Comment est calculé le score de criticité ?
        </div>
        <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px;margin-bottom:16px">
            <div style="background:#fff;border-radius:10px;padding:14px;border:1px solid #bfdbfe;text-align:center">
            <div style="font-size:1.4rem;margin-bottom:4px">🧬</div>
            <div style="font-weight:800;color:#1e40af;font-size:.88rem">Pathogénicité</div>
            <div style="font-size:.72rem;color:#475569;margin-top:6px;line-height:1.6">
                <b>1</b> — Non pathogène<br>
                <b>2</b> — Pathogène opportuniste<br>
                <b>3</b> — Pathogène MR / primaire
            </div>
            </div>
            <div style="background:#fff;border-radius:10px;padding:14px;border:1px solid #bfdbfe;text-align:center">
            <div style="font-size:1.4rem;margin-bottom:4px">🧴</div>
            <div style="font-weight:800;color:#1e40af;font-size:.88rem">Résistance désinfectants</div>
            <div style="font-size:.72rem;color:#475569;margin-top:6px;line-height:1.6">
                <b>1</b> — Sensible<br>
                <b>2</b> — Résistant Surfa'Safe<br>
                <b>3</b> — Résistant Surfa'Safe + APA
            </div>
            </div>
            <div style="background:#fff;border-radius:10px;padding:14px;border:1px solid #bfdbfe;text-align:center">
            <div style="font-size:1.4rem;margin-bottom:4px">💨</div>
            <div style="font-weight:800;color:#1e40af;font-size:.88rem">Dissémination</div>
            <div style="font-size:.72rem;color:#475569;margin-top:6px;line-height:1.6">
                <b>1</b> — Environnemental<br>
                <b>2</b> — Manuporté<br>
                <b>3</b> — Aéroporté
            </div>
            </div>
        </div>
        <div style="background:#1e293b;border-radius:10px;padding:14px;text-align:center;margin-bottom:14px">
            <div style="font-size:.7rem;color:#94a3b8;text-transform:uppercase;letter-spacing:.1em;margin-bottom:6px">
            Formule du score total
            </div>
            <div style="font-size:1rem;color:#e2e8f0;font-weight:700">
            Score total = <span style="color:#60a5fa">Criticité lieu (1–4)</span>
            × <span style="color:#34d399">Pathogénicité (1–3)</span>
            × <span style="color:#fbbf24">Résistance (1–3)</span>
            × <span style="color:#f87171">Dissémination (1–3)</span>
            </div>
            <div style="font-size:.72rem;color:#64748b;margin-top:8px">
            Score minimum : 1×1×1×1 = <b style="color:#94a3b8">1</b>
            &nbsp;·&nbsp;
            Score maximum : 4×3×3×3 = <b style="color:#f87171">108</b>
            </div>
        </div>
        <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px">
            <div style="background:#f0fdf4;border:1.5px solid #86efac;border-radius:8px;padding:12px;text-align:center">
            <div style="font-size:1.1rem">✅</div>
            <div style="font-weight:800;color:#166534;font-size:.85rem;margin-top:4px">CONFORME</div>
            <div style="font-size:.78rem;color:#166534;margin-top:4px">Score &lt; seuil alerte</div>
            </div>
            <div style="background:#fffbeb;border:1.5px solid #fcd34d;border-radius:8px;padding:12px;text-align:center">
            <div style="font-size:1.1rem">⚠️</div>
            <div style="font-weight:800;color:#92400e;font-size:.85rem;margin-top:4px">ALERTE</div>
            <div style="font-size:.78rem;color:#92400e;margin-top:4px">Seuil alerte ≤ Score ≤ seuil action</div>
            </div>
            <div style="background:#fef2f2;border:1.5px solid #fca5a5;border-radius:8px;padding:12px;text-align:center">
            <div style="font-size:1.1rem">🚨</div>
            <div style="font-weight:800;color:#991b1b;font-size:.85rem;margin-top:4px">ACTION</div>
            <div style="font-size:.78rem;color:#991b1b;margin-top:4px">Score &gt; seuil action</div>
            </div>
        </div>
        </div>""", unsafe_allow_html=True)

        _seuil_alerte = st.session_state.get("_seuil_alerte", 24)
        _seuil_action = st.session_state.get("_seuil_action", 36)

        st.markdown("#### ⚙️ Modifier les seuils")
        if not can_edit:
            st.info("👁️ Mode lecture seule — connectez-vous pour modifier les seuils.")

        sc1, sc2, sc3 = st.columns([2, 2, 3])
        with sc1:
            new_seuil_alerte = st.number_input(
                "⚠️ Seuil ALERTE",
                min_value=1, max_value=80, value=int(_seuil_alerte), step=1,
                disabled=not can_edit,
                help="En dessous : conforme. À partir de ce score : alerte.",
                key="input_seuil_alerte")
        with sc2:
            new_seuil_action = st.number_input(
                "🚨 Seuil ACTION",
                min_value=1, max_value=108, value=int(_seuil_action), step=1,
                disabled=not can_edit,
                help="Au-dessus de ce score : action immédiate requise.",
                key="input_seuil_action")
        with sc3:
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            if can_edit:
                if st.button("💾 Sauvegarder les seuils", use_container_width=True,
                            key="save_seuils", type="primary"):
                    if new_seuil_alerte >= new_seuil_action:
                        st.error("❌ Le seuil d'alerte doit être strictement inférieur au seuil d'action.")
                    else:
                        st.session_state["_seuil_alerte"] = new_seuil_alerte
                        st.session_state["_seuil_action"] = new_seuil_action
                        _supa_upsert('seuils', json.dumps({
                            "alerte": new_seuil_alerte,
                            "action": new_seuil_action
                        }, ensure_ascii=False))
                        st.success(
                            f"✅ Seuils sauvegardés — Alerte : {new_seuil_alerte} · Action : {new_seuil_action}")
                        st.rerun()

        if new_seuil_alerte >= new_seuil_action:
            st.error("❌ Le seuil d'alerte doit être strictement inférieur au seuil d'action.")
        else:
            st.markdown(f"""
            <div style="background:#f8fafc;border:1.5px solid #e2e8f0;border-radius:10px;
            padding:14px 18px;margin-top:8px">
            <div style="font-size:.78rem;font-weight:700;color:#475569;margin-bottom:10px">
                Aperçu de la grille avec ces seuils
            </div>
            <div style="display:flex;gap:0;border-radius:8px;overflow:hidden;border:1px solid #e2e8f0">
                <div style="flex:1;background:#f0fdf4;padding:10px;text-align:center;border-right:1px solid #e2e8f0">
                <div style="font-size:.65rem;color:#166534;font-weight:700;text-transform:uppercase">✅ Conforme</div>
                <div style="font-size:1.1rem;font-weight:900;color:#166534;margin-top:2px">
                    Score &lt; {new_seuil_alerte}
                </div>
                </div>
                <div style="flex:1;background:#fffbeb;padding:10px;text-align:center;border-right:1px solid #e2e8f0">
                <div style="font-size:.65rem;color:#92400e;font-weight:700;text-transform:uppercase">⚠️ Alerte</div>
                <div style="font-size:1.1rem;font-weight:900;color:#92400e;margin-top:2px">
                    {new_seuil_alerte} – {new_seuil_action}
                </div>
                </div>
                <div style="flex:1;background:#fef2f2;padding:10px;text-align:center">
                <div style="font-size:.65rem;color:#991b1b;font-weight:700;text-transform:uppercase">🚨 Action</div>
                <div style="font-size:1.1rem;font-weight:900;color:#dc2626;margin-top:2px">
                    Score &gt; {new_seuil_action}
                </div>
                </div>
            </div>
            </div>""", unsafe_allow_html=True)

        st.divider()

        st.markdown("#### 📊 Tableau de référence — scores limites par criticité de lieu")
        st.caption(
            "Montre à quel score germe (pathogénicité × résistance × dissémination) "
            "les seuils sont déclenchés selon la criticité du lieu.")

        _sa     = new_seuil_alerte
        _sc_val = new_seuil_action

        st.markdown(
            "<div style='display:grid;grid-template-columns:1.5fr 1fr 1fr 1fr 2fr;"
            "gap:4px;background:#1e40af;border-radius:10px 10px 0 0;padding:10px 14px'>"
            "<div style='font-size:.72rem;font-weight:800;color:#fff'>Criticité lieu</div>"
            "<div style='font-size:.72rem;font-weight:800;color:#fff;text-align:center'>Score lieu</div>"
            "<div style='font-size:.72rem;font-weight:800;color:#fff;text-align:center'>Germe → ⚠️ Alerte</div>"
            "<div style='font-size:.72rem;font-weight:800;color:#fff;text-align:center'>Germe → 🚨 Action</div>"
            "<div style='font-size:.72rem;font-weight:800;color:#fff'>Exemples de lieux</div>"
            "</div>",
            unsafe_allow_html=True)

        # ── MISE À JOUR : criticité 1 à 4 ──────────────────────────────────────
        lc_examples = {
            1: "Couloirs, locaux techniques, zones administratives",
            2: "Préparations non stériles, zones annexes ZAC, vestiaires",
            3: "ZAC, salles blanches ISO C/D, zones de remplissage",
            4: "Salles blanches ISO A/B, isolateurs, zones aseptiques critiques",
        }
        lc_colors_ref = {
            1: "#22c55e",   # vert
            2: "#f59e0b",   # orange
            3: "#ef4444",   # rouge
            4: "#7c3aed",   # violet (criticité maximale)
        }

        for lci, loc_crit_val in enumerate([1, 2, 3, 4]):
            lc_lbl       = f"Nv.{loc_crit_val} — {LOC_CRIT_LABELS[str(loc_crit_val)]}"
            germe_alerte = _sa / loc_crit_val
            germe_action = _sc_val / loc_crit_val
            lc_col = lc_colors_ref[loc_crit_val]
            row_bg = "#f8fafc" if lci % 2 == 0 else "#ffffff"
            # Arrondir le bas uniquement sur la dernière ligne
            border_radius = "0 0 10px 10px" if loc_crit_val == 4 else "0"
            st.markdown(
                "<div style='display:grid;grid-template-columns:1.5fr 1fr 1fr 1fr 2fr;"
                f"gap:4px;background:{row_bg};border:1px solid #e2e8f0;border-top:none;"
                f"padding:10px 14px;align-items:center;border-radius:{border_radius}'>"
                f"<div style='font-size:.85rem;font-weight:700'>"
                f"<span style='color:{lc_col}'>●</span> {lc_lbl}</div>"
                f"<div style='text-align:center'>"
                f"<span style='background:{lc_col}22;color:{lc_col};"
                f"border:1px solid {lc_col}55;border-radius:6px;"
                f"padding:2px 10px;font-size:.82rem;font-weight:800'>× {loc_crit_val}</span></div>"
                f"<div style='text-align:center;font-size:.82rem;font-weight:700;color:#92400e'>"
                f"Score germe ≥ {germe_alerte:.1f}</div>"
                f"<div style='text-align:center;font-size:.82rem;font-weight:700;color:#dc2626'>"
                f"Score germe &gt; {germe_action:.1f}</div>"
                f"<div style='font-size:.7rem;color:#64748b;font-style:italic'>"
                f"{lc_examples[loc_crit_val]}</div>"
                "</div>",
                unsafe_allow_html=True)

        st.markdown(
            "<div style='background:#1e293b;border-radius:0 0 10px 10px;padding:8px 14px'>"
            f"<div style='font-size:.75rem;color:#94a3b8'>"
            f"Score germe = Pathogénicité × Résistance × Dissémination (min 1 · max 27) "
            f"· Seuil alerte : <b style='color:#fbbf24'>{_sa}</b> "
            f"· Seuil action : <b style='color:#f87171'>{_sc_val}</b>"
            f"</div></div>",
            unsafe_allow_html=True)

        st.divider()

        st.markdown("#### ↩️ Réinitialiser aux valeurs par défaut")
        st.markdown(
            "<div style='background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;"
            "padding:10px 14px;font-size:.78rem;color:#475569;margin-bottom:8px'>"
            "Les valeurs par défaut sont <b>Alerte : 24</b> et <b>Action : 36</b>.<br>"
            "Ces seuils correspondent à :<br>"
            "• Alerte dès qu'un germe de score 8 est trouvé en zone critique (3×8=24)<br>"
            "• Action dès qu'un germe de score 9 est trouvé en zone critique (4×9=36)"
            "</div>",
            unsafe_allow_html=True)
        if can_edit:
            if st.button("↩️ Remettre Alerte=24 / Action=36", key="reset_seuils"):
                st.session_state["_seuil_alerte"] = 24
                st.session_state["_seuil_action"] = 36
                _supa_upsert('seuils', json.dumps({"alerte": 24, "action": 36}, ensure_ascii=False))
                st.success("✅ Seuils réinitialisés — Alerte : 24 · Action : 36")
                st.rerun()

    # ══════════════════════════════════════════════════════════════════════════
    # OPÉRATEURS
    # ══════════════════════════════════════════════════════════════════════════
    with subtab_operateurs:
        ops = st.session_state.operators
        if not ops:
            st.info("Aucun opérateur enregistré.")
        else:
            st.markdown(
                f'<div style="background:#f0f9ff;border:1px solid #bae6fd;border-radius:10px;'
                f'padding:12px 16px;margin-bottom:16px">'
                f'<span style="font-size:.75rem;color:#0369a1;font-weight:700">'
                f'👥 {len(ops)} opérateur(s)</span></div>',
                unsafe_allow_html=True)
            for i, op in enumerate(ops):
                nom        = op.get('nom', '—')
                profession = op.get('profession', '—')
                oc1, oc2, oc3 = st.columns([5, 1, 1])
                with oc1:
                    st.markdown(f"""
                    <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;
                    padding:10px 14px;display:flex;gap:16px;align-items:center">
                      <div style="background:#2563eb;color:#fff;border-radius:50%;
                      width:36px;height:36px;display:flex;align-items:center;justify-content:center;
                      font-weight:700;font-size:.9rem;flex-shrink:0">
                        {nom[0].upper() if nom else '?'}
                      </div>
                      <div>
                        <div style="font-weight:700;font-size:.9rem;color:#0f172a">{nom}</div>
                        <div style="font-size:.72rem;color:#475569;margin-top:2px">👔 {profession}</div>
                      </div>
                    </div>""", unsafe_allow_html=True)
                with oc2:
                    if can_edit:
                        if st.button("✏️", key=f"edit_op_{i}"):
                            st.session_state._edit_operator = i
                            st.rerun()
                with oc3:
                    if can_edit:
                        if st.button("🗑️", key=f"del_op_{i}"):
                            ops.pop(i)
                            save_operators(ops, supa=True)
                            st.session_state.operators = ops
                            st.rerun()

        st.divider()
        p_opts = ["Préparateur en pharmacie hospitalière", "Pharmacien", "Interne de pharmacie"]

        if st.session_state.get('_edit_operator') is not None:
            idx = st.session_state._edit_operator
            op  = st.session_state.operators[idx]
            st.markdown(f"### ✏️ Modifier — {op.get('nom', '')}")
            ec1, ec2 = st.columns(2)
            with ec1:
                edit_nom = st.text_input("Nom *", value=op.get('nom', ''), key="op_edit_nom")
            with ec2:
                cur_p    = op.get('profession', '')
                p_idx    = p_opts.index(cur_p) if cur_p in p_opts else 0
                edit_pro = st.selectbox("Profession *", p_opts, index=p_idx, key="op_edit_prof")
            eb1, eb2 = st.columns(2)
            with eb1:
                if st.button("✅ Enregistrer", use_container_width=True, key="op_save_edit"):
                    if edit_nom.strip():
                        st.session_state.operators[idx] = {
                            "nom": edit_nom.strip(), "profession": edit_pro}
                        save_operators(st.session_state.operators, supa=True)
                        st.session_state._edit_operator = None
                        st.success("✅ Mis à jour")
                        st.rerun()
                    else:
                        st.error("Le nom est obligatoire.")
            with eb2:
                if st.button("Annuler", use_container_width=True, key="op_cancel_edit"):
                    st.session_state._edit_operator = None
                    st.rerun()
        elif can_edit:
            st.markdown("### ➕ Ajouter un opérateur")
            nc1, nc2 = st.columns(2)
            with nc1:
                new_nom = st.text_input("Nom *", placeholder="Ex: Marie Dupont", key="op_new_nom")
            with nc2:
                new_pro = st.selectbox("Profession *", p_opts, key="op_new_prof")
            if st.button("➕ Ajouter", key="op_add"):
                if not new_nom.strip():
                    st.error("Le nom est obligatoire.")
                elif any(o['nom'].lower() == new_nom.strip().lower() for o in st.session_state.operators):
                    st.error("Cet opérateur existe déjà.")
                else:
                    st.session_state.operators.append({"nom": new_nom.strip(), "profession": new_pro})
                    save_operators(st.session_state.operators, supa=True)
                    st.success(f"✅ **{new_nom}** ajouté")
                    st.rerun()

    # ══════════════════════════════════════════════════════════════════════════
    # SAUVEGARDE
    # ══════════════════════════════════════════════════════════════════════════
    with subtab_backup:
        st.markdown("### 💾 Sauvegarde & Restauration")
        supa_connected = get_supabase_client() is not None
        if supa_connected:
            st.success("✅ **Supabase actif** — données persistantes dans le cloud.")
        else:
            st.warning("⚠️ **Supabase non configuré** — données perdues au redémarrage.")

        st.markdown("""
        <div style="background:#fffbeb;border:1.5px solid #fcd34d;border-radius:12px;
        padding:16px 20px;margin:12px 0">
          <div style="font-weight:800;color:#92400e;font-size:.95rem;margin-bottom:8px">
            📋 Pourquoi sauvegarder ?
          </div>
          <div style="font-size:.82rem;color:#78350f;line-height:1.8">
            Chaque modification du code provoque un redémarrage. Sans Supabase, toutes
            les données locales sont <strong>effacées</strong>.<br>
            ✅ <strong>Solution 1</strong> : configurer Supabase (onglet ☁️).<br>
            ✅ <strong>Solution 2</strong> : exporter avant chaque update, réimporter après.
          </div>
        </div>""", unsafe_allow_html=True)

        st.divider()
        st.markdown("#### ⬇️ Exporter toutes les données")
        backup_data     = export_all_data()
        backup_json     = json.dumps(backup_data, ensure_ascii=False, indent=2)
        backup_filename = f"backup_URC_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        b1, b2, b3, b4 = st.columns(4)
        b1.metric("🦠 Germes",          len(backup_data.get("germs", [])))
        b2.metric("🧪 Prélèvements",    len(backup_data.get("prelevements", [])))
        b3.metric("📅 Lectures planif.", len(backup_data.get("schedules", [])))
        b4.metric("📋 Analyse",       len(backup_data.get("surveillance", [])))
        st.download_button(
            label=f"⬇️ Télécharger ({len(backup_json)//1024 + 1} Ko)",
            data=backup_json, file_name=backup_filename,
            mime="application/json",
            use_container_width=True, key="main_export_btn")

        st.divider()
        st.markdown("#### ⬆️ Restaurer depuis une sauvegarde")
        st.markdown("""
        <div style="background:#fef2f2;border:1.5px solid #fca5a5;border-radius:10px;
        padding:12px 16px;margin-bottom:12px">
          <span style="color:#dc2626;font-weight:700;font-size:.82rem">
            ⚠️ La restauration remplace TOUTES les données sans possibilité d'annulation.
          </span>
        </div>""", unsafe_allow_html=True)

        uploaded_backup = st.file_uploader(
            "Fichier de sauvegarde (.json)", type=["json"], key="backup_uploader")
        if uploaded_backup is not None:
            try:
                backup_content = json.loads(uploaded_backup.read().decode("utf-8"))
                meta = backup_content.get("_meta", {})
                st.markdown(f"""
                <div style="background:#f0fdf4;border:1px solid #86efac;border-radius:10px;
                padding:14px 18px;margin-bottom:12px">
                  <div style="font-weight:700;color:#166534;font-size:.85rem;margin-bottom:8px">
                    📁 Contenu détecté
                  </div>
                  <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:8px;
                  font-size:.75rem;color:#0f172a">
                    <div>🦠 Germes : <strong>{len(backup_content.get("germs",[]))}</strong></div>
                    <div>🧪 Prélèvements : <strong>{len(backup_content.get("prelevements",[]))}</strong></div>
                    <div>📅 Lectures : <strong>{len(backup_content.get("schedules",[]))}</strong></div>
                    <div>👤 Opérateurs : <strong>{len(backup_content.get("operators",[]))}</strong></div>
                    <div>📍 Points : <strong>{len(backup_content.get("points",[]))}</strong></div>
                    <div>📋 Analyse : <strong>{len(backup_content.get("surveillance",[]))}</strong></div>
                  </div>
                  <div style="font-size:.68rem;color:#475569;margin-top:8px">
                    Exporté le : {meta.get("exported_at","—")[:19].replace("T"," ")}
                    | Version : {meta.get("version","?")}
                  </div>
                </div>""", unsafe_allow_html=True)

                if st.session_state.get("confirm_restore", False):
                    st.error("🚨 Dernière confirmation — toutes les données seront remplacées.")
                    rc1, rc2 = st.columns(2)
                    with rc1:
                        if st.button("✅ OUI — Restaurer maintenant",
                                     use_container_width=True, key="confirm_restore_yes"):
                            ok, msg = import_all_data(backup_content)
                            st.session_state.confirm_restore = False
                            if ok: st.success(f"✅ {msg}"); st.rerun()
                            else:  st.error(msg)
                    with rc2:
                        if st.button("❌ Annuler", use_container_width=True, key="confirm_restore_no"):
                            st.session_state.confirm_restore = False
                            st.rerun()
                else:
                    if can_edit:
                        if st.button("⬆️ Restaurer ces données",
                                     use_container_width=True, key="restore_btn"):
                            st.session_state.confirm_restore = True
                            st.rerun()
            except json.JSONDecodeError:
                st.error("❌ Fichier JSON invalide.")
            except Exception as e:
                st.error(f"❌ Erreur : {e}")

    # ══════════════════════════════════════════════════════════════════════════
    # SUPABASE
    # ══════════════════════════════════════════════════════════════════════════
    with subtab_supabase:
        st.markdown("### ☁️ Configuration Supabase")
        supa_ok = get_supabase_client() is not None
        if supa_ok:
            st.success("✅ **Supabase connecté** — modifications synchronisées en temps réel.")
        else:
            st.error("🔴 **Supabase non connecté** — sauvegarde locale uniquement.")

        st.markdown("""
        <div style="background:#f8fafc;border:1.5px solid #e2e8f0;border-radius:12px;
        padding:20px;margin-top:16px">
          <div style="font-size:.95rem;font-weight:700;color:#0f172a;margin-bottom:12px">
            📋 Comment configurer Supabase
          </div>
          <div style="font-size:.82rem;color:#1e293b;line-height:1.8">
            <strong>1.</strong> Créez un compte sur <strong>supabase.com</strong><br>
            <strong>2.</strong> Créez un nouveau projet<br>
            <strong>3.</strong> Dans l'éditeur SQL, exécutez le code ci-dessous
          </div>
        </div>""", unsafe_allow_html=True)

        st.code("""CREATE TABLE IF NOT EXISTS app_state (
    key TEXT PRIMARY KEY,
    value TEXT NOT NULL,
    updated_at TIMESTAMP DEFAULT NOW()
);
ALTER TABLE app_state ENABLE ROW LEVEL SECURITY;
CREATE POLICY "allow_all" ON app_state FOR ALL USING (true) WITH CHECK (true);""",
                language="sql")

        st.markdown("""
        <div style="background:#f8fafc;border:1.5px solid #e2e8f0;border-radius:12px;
        padding:20px;margin-top:12px">
          <div style="font-size:.82rem;color:#1e293b;line-height:1.8">
            <strong>4.</strong> Dans <em>Project Settings → API</em>, copiez :<br>
            &nbsp;&nbsp;• <strong>Project URL</strong> → <code>SUPABASE_URL</code><br>
            &nbsp;&nbsp;• <strong>anon/public key</strong> → <code>SUPABASE_KEY</code>
          </div>
        </div>""", unsafe_allow_html=True)

        st.code("""SUPABASE_URL = "https://xxxxx.supabase.co"
SUPABASE_KEY = "eyJhbGci..."  # votre clé anon""", language="toml")

        if supa_ok:
            st.divider()
            st.markdown("### 🔄 Actions Supabase")
            syn1, syn2 = st.columns(2)
            with syn1:
                if can_edit:
                    if st.button("🔄 Forcer la synchronisation", use_container_width=True):
                        save_germs(st.session_state.germs)
                        save_prelevements(st.session_state.prelevements, supa=True)
                        save_schedules(st.session_state.schedules, supa=True)
                        save_surveillance(st.session_state.surveillance)
                        save_points(st.session_state.points, supa=True)
                        save_operators(st.session_state.operators, supa=True)
                        save_pending_identifications(st.session_state.pending_identifications, supa=True)
                        save_origin_measures(st.session_state.origin_measures, supa=True)
                        save_faq(st.session_state.faq_items, supa=True)
                        st.session_state["_mesures_modifiees"] = False
                        st.success("✅ Toutes les données synchronisées !")
            with syn2:
                if can_edit:
                    if st.button("🔃 Recharger depuis Supabase", use_container_width=True):
                        st.session_state.germs                   = load_germs()[0]
                        st.session_state.prelevements            = load_prelevements()
                        st.session_state.schedules               = load_schedules()
                        st.session_state.surveillance            = load_surveillance()
                        st.session_state.points                  = load_points()
                        st.session_state.operators               = load_operators()
                        st.session_state.pending_identifications = load_pending_identifications()
                        st.session_state.origin_measures         = load_origin_measures()
                        st.session_state.faq_items               = load_faq()
                        st.success("✅ Données rechargées depuis Supabase !")
                        st.rerun()

    # ══════════════════════════════════════════════════════════════════════════
    # FAQ
    # ══════════════════════════════════════════════════════════════════════════
    with subtab_faq:
        faq_items = st.session_state.get("faq_items", [])

        cats_count = {}
        for f in faq_items:
            c = f.get("category", "Général")
            cats_count[c] = cats_count.get(c, 0) + 1

        cols_stat = st.columns(min(len(cats_count) + 1, 5))
        with cols_stat[0]:
            st.metric("Total Q&R", len(faq_items))
        for i, (cat, cnt) in enumerate(list(cats_count.items())[:4], 1):
            with cols_stat[i]:
                st.metric(cat[:14], cnt)

        st.divider()

        edit_idx = st.session_state.get("_faq_edit_idx")

        if can_edit and st.session_state.get("_faq_show_form", False):
            is_edit  = edit_idx is not None
            existing = faq_items[edit_idx] if is_edit else {}
            form_bg  = "#eff6ff" if is_edit else "#f0fdf4"
            form_bdr = "#93c5fd" if is_edit else "#86efac"
            form_ttl = "✏️ Modifier la question" if is_edit else "➕ Nouvelle question"

            st.markdown(
                f"<div style='background:{form_bg};border:1.5px solid {form_bdr};"
                f"border-radius:12px;padding:18px;margin-bottom:16px'>",
                unsafe_allow_html=True)
            st.markdown(f"#### {form_ttl}")

            fc1, fc2 = st.columns([3, 1])
            with fc1:
                faq_q = st.text_input(
                    "Question *", value=existing.get("question", ""),
                    placeholder="Ex: Comment ajouter un point de prélèvement ?",
                    key="faq_form_question")
            with fc2:
                cur_cat = existing.get("category", "Général")
                faq_c = st.selectbox(
                    "Catégorie", FAQ_CATEGORIES,
                    index=FAQ_CATEGORIES.index(cur_cat) if cur_cat in FAQ_CATEGORIES else 0,
                    key="faq_form_category")

            faq_a = st.text_area(
                "Réponse * (Markdown supporté)", value=existing.get("answer", ""),
                height=150,
                placeholder="Décrivez la réponse. **Gras**, *italique*, listes…",
                key="faq_form_answer")
            if faq_a.strip():
                with st.expander("👁️ Aperçu", expanded=False):
                    st.markdown(faq_a)

            fb1, fb2 = st.columns(2)
            with fb1:
                if st.button(
                    "✔️ Mettre à jour" if is_edit else "✅ Ajouter",
                    use_container_width=True, type="primary", key="faq_form_submit"):
                    if not faq_q.strip():
                        st.error("La question est obligatoire.")
                    elif not faq_a.strip():
                        st.error("La réponse est obligatoire.")
                    else:
                        if is_edit:
                            faq_items[edit_idx].update(
                                question=faq_q.strip(), answer=faq_a.strip(), category=faq_c)
                        else:
                            faq_items.append({
                                "id":       f"faq_{int(datetime.now().timestamp())}",
                                "category": faq_c,
                                "question": faq_q.strip(),
                                "answer":   faq_a.strip(),
                                "order":    len(faq_items),
                            })
                        save_faq(faq_items, supa=True)
                        st.session_state["faq_items"]      = faq_items
                        st.session_state["_faq_show_form"] = False
                        st.session_state["_faq_edit_idx"]  = None
                        st.success("✅ FAQ mise à jour !")
                        st.rerun()
            with fb2:
                if st.button("✕ Annuler", use_container_width=True, key="faq_form_cancel"):
                    st.session_state["_faq_show_form"] = False
                    st.session_state["_faq_edit_idx"]  = None
                    st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)

        elif can_edit and not st.session_state.get("_faq_show_form", False):
            if st.button("➕ Ajouter une question", key="faq_add_btn", use_container_width=True):
                st.session_state["_faq_show_form"] = True
                st.session_state["_faq_edit_idx"]  = None
                st.rerun()

        if not faq_items:
            st.markdown(
                "<div style='background:#f8fafc;border:1.5px dashed #cbd5e1;"
                "border-radius:12px;padding:32px;text-align:center;margin-top:12px'>"
                "<div style='font-size:2.5rem;margin-bottom:8px'>❓</div>"
                "<div style='font-weight:700;color:#475569'>Aucune question définie</div>"
                "<div style='font-size:.8rem;color:#94a3b8;margin-top:4px'>"
                "Cliquez sur ➕ Ajouter une question ci-dessus</div></div>",
                unsafe_allow_html=True)
        else:
            all_cats_tab = ["Toutes"] + sorted(set(f.get("category", "Général") for f in faq_items))
            faq_filter_cat = st.selectbox(
                "Filtrer", all_cats_tab, key="faq_tab_cat_filter", label_visibility="collapsed")

            st.markdown(
                "<div style='display:grid;grid-template-columns:2fr 1fr;"
                "gap:4px;background:#1e40af;border-radius:10px 10px 0 0;"
                "padding:10px 14px;margin-top:8px'>"
                "<div style='font-size:.72rem;font-weight:800;color:#fff'>Question</div>"
                "<div style='font-size:.72rem;font-weight:800;color:#fff;text-align:center'>Catégorie</div>"
                "</div>",
                unsafe_allow_html=True)

            CAT_COL = {
                "Général":             "#2563eb",
                "Score & Seuils":      "#7c3aed",
                "Prélèvements":        "#0891b2",
                "Paramètres":          "#059669",
                "Données":             "#d97706",
                "Mesures correctives": "#dc2626",
            }

            displayed = [
                (i, f) for i, f in enumerate(faq_items)
                if faq_filter_cat == "Toutes" or f.get("category") == faq_filter_cat
            ]

            if not displayed:
                st.markdown(
                    "<div style='background:#f8fafc;border:1px solid #e2e8f0;border-top:none;"
                    "border-radius:0 0 10px 10px;padding:20px;text-align:center;"
                    "color:#94a3b8;font-size:.82rem'>Aucune question dans cette catégorie</div>",
                    unsafe_allow_html=True)
            else:
                for dp, (ri, item) in enumerate(displayed):
                    cc     = CAT_COL.get(item.get("category", "Général"), "#475569")
                    row_bg = "#f8fafc" if dp % 2 == 0 else "#ffffff"

                    rc1, rc2 = st.columns([6, 1])
                    with rc1:
                        st.markdown(
                            f"<div style='display:grid;grid-template-columns:2fr 1fr;"
                            f"gap:4px;background:{row_bg};border:1px solid #e2e8f0;"
                            f"border-top:none;padding:10px 14px;align-items:center'>"
                            f"<div style='font-size:.82rem;font-weight:600;color:#0f172a'>"
                            f"{item['question']}</div>"
                            f"<div style='text-align:center'>"
                            f"<span style='background:{cc}18;color:{cc};"
                            f"border:1px solid {cc}44;border-radius:12px;"
                            f"padding:2px 10px;font-size:.65rem;font-weight:700'>"
                            f"{item.get('category','Général')}</span></div></div>",
                            unsafe_allow_html=True)
                    with rc2:
                        a1, a2, a3, a4 = st.columns(4)
                        with a1:
                            if can_edit and ri > 0:
                                if st.button("↑", key=f"faq_up_{ri}", help="Monter"):
                                    faq_items[ri], faq_items[ri-1] = faq_items[ri-1], faq_items[ri]
                                    for k, f in enumerate(faq_items): f["order"] = k
                                    save_faq(faq_items, supa=True)
                                    st.session_state["faq_items"] = faq_items
                                    st.rerun()
                        with a2:
                            if can_edit and ri < len(faq_items) - 1:
                                if st.button("↓", key=f"faq_dn_{ri}", help="Descendre"):
                                    faq_items[ri], faq_items[ri+1] = faq_items[ri+1], faq_items[ri]
                                    for k, f in enumerate(faq_items): f["order"] = k
                                    save_faq(faq_items, supa=True)
                                    st.session_state["faq_items"] = faq_items
                                    st.rerun()
                        with a3:
                            if can_edit:
                                if st.button("✏️", key=f"faq_edit_{ri}"):
                                    st.session_state["_faq_edit_idx"]  = ri
                                    st.session_state["_faq_show_form"] = True
                                    st.rerun()
                        with a4:
                            if can_edit:
                                if st.button("🗑️", key=f"faq_del_{ri}"):
                                    faq_items.pop(ri)
                                    for k, f in enumerate(faq_items): f["order"] = k
                                    save_faq(faq_items, supa=True)
                                    st.session_state["faq_items"] = faq_items
                                    st.rerun()

                st.markdown(
                    f"<div style='background:#1e293b;border-radius:0 0 10px 10px;"
                    f"padding:8px 14px'><div style='font-size:.75rem;color:#94a3b8'>"
                    f"{len(faq_items)} question(s) · {len(displayed)} affichée(s)"
                    f"</div></div>",
                    unsafe_allow_html=True)

        st.divider()

        if can_edit:
            st.markdown("#### ↩️ Réinitialiser la FAQ")
            st.caption("Recharge les questions prédéfinies (efface les modifications personnalisées).")
            if st.button("↩️ Remettre les questions par défaut", key="faq_reset"):
                st.session_state["faq_items"] = [dict(f) for f in DEFAULT_FAQ]
                save_faq(st.session_state["faq_items"], supa=True)
                st.success("✅ FAQ réinitialisée.")
                st.rerun()